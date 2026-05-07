# -*- coding: utf-8 -*-
"""Эвристический разбор .xlsx (КС-2): период документа, таблица работ, сводка по месяцам, экспорт."""

from __future__ import annotations

import calendar
import hashlib
import importlib
import json
import os
import pickle
import re
import sys
import traceback
from dataclasses import dataclass, field
from decimal import Decimal
from datetime import date, datetime
from typing import Any, Iterable

_DATE_CELL = re.compile(
    r"(?<!\d)(?:0?[1-9]|[12]\d|3[01])[\./](?:0?[1-9]|1[0-2])[\./](?:\d{2}|\d{4})(?!\d)",
    re.UNICODE,
)
_KS2 = re.compile(r"к\s*с\s*[-−–]*\s*2", re.IGNORECASE)

HEADER_MAX_ROW = 30
HEADER_MAX_COL = 40
BODY_MAX_ROW_SHEET1 = 220
BODY_MAX_COL = 44
MAX_SHEETS_DATE_SCAN = 10
MAX_ROW_OTHER_SHEETS = 120
MAX_TABLE_SCAN_ROW = 220
MAX_TABLE_ROWS = 600
EMPTY_STREAK_STOP = 6

KEYS_NAME = ("наимен", "содержан", "работы", "работа")
KEYS_UNIT = ("ед.изм", "ед изм", "единиц")
KEYS_QTY = ("колич", "кол-во", "кол во", "объём", "объем", "объё", "кол-в")
KEYS_PRICE = ("цена", "тариф")
# В подзаголовках КС‑2 есть «Кол.: всего» — отдельное «всего» нельзя относить к деньгам.
KEYS_SUM = ("сумма", "стоим")
KEYS_SUM_HEADER = KEYS_SUM + ("общая",)
KEYS_DATE = ("дата", "период", "месяц")

UNKNOWN_MONTH_KEY = (0, 0)
_CACHE_VER = 13
_CACHE_DIR = os.path.join(os.path.dirname(__file__), "Данные", "cache_ks2")
_STORE_VER = 11
_STORE_PATH = os.path.join(os.path.dirname(__file__), "Данные", "ks2_results_store.pkl")
_DEBUG_LOG_PATH = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "debug-1b1da5.log"))
_AI_ASSIST_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "Окраска_по_PDF"))
_MIN_PLAUSIBLE_YEAR = 2010
_MAX_PLAUSIBLE_YEAR = 2100


def _дата_правдоподобна(dt: tuple[int, int, int]) -> bool:
    y, m, d = dt
    if y < _MIN_PLAUSIBLE_YEAR or y > _MAX_PLAUSIBLE_YEAR:
        return False
    if m < 1 or m > 12:
        return False
    if d < 1 or d > 31:
        return False
    return True


def _dbg_log(hypothesis_id: str, location: str, message: str, data: dict | None = None, run_id: str = "pre-fix") -> None:
    #region agent log
    try:
        payload = {
            "sessionId": "1b1da5",
            "runId": run_id,
            "hypothesisId": hypothesis_id,
            "location": location,
            "message": message,
            "data": data or {},
            "timestamp": int(__import__("time").time() * 1000),
        }
        with open(_DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        pass
    #endregion


def _получить_ai_модуль():
    try:
        if _AI_ASSIST_DIR not in sys.path:
            sys.path.insert(0, _AI_ASSIST_DIR)
        return importlib.import_module("ассистент_llm")
    except Exception:
        return None

_VSEGO_PO_AKT = re.compile(r"все\s*го\s*по\s*акт", re.IGNORECASE)
_PERIOD_RANGE_HEAD = re.compile(
    r"""
    (?:с|от)\s*
    (?P<d1>\d{1,2})[.](?P<m1>\d{1,2})[.](?P<y1>\d{2,4})
    \s*(?:г\.?)?\s*
    (?:по|до|-|–|—)
    \s*
    (?P<d2>\d{1,2})[.](?P<m2>\d{1,2})[.](?P<y2>\d{2,4})
    """,
    re.IGNORECASE | re.VERBOSE,
)
_PERIOD_RANGE_HEAD_SLASH = re.compile(
    r"""
    (?:с|от)\s*
    (?P<d1>\d{1,2})[/](?P<m1>\d{1,2})[/](?P<y1>\d{2,4})
    \s*(?:г\.?)?\s*
    (?:по|до|-|–|—)
    \s*
    (?P<d2>\d{1,2})[/](?P<m2>\d{1,2})[/](?P<y2>\d{2,4})
    """,
    re.IGNORECASE | re.VERBOSE,
)


def _norm_header(s: str) -> str:
    t = s.lower().replace("ё", "е")
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _дата_из_совпадения(m: re.Match[str]) -> tuple[int, int, int] | None:
    d_s, mo_s, y_s = m.group(0).replace("/", ".").split(".")
    try:
        d, mo = int(d_s), int(mo_s)
        y = int(y_s)
        if y < 100:
            y += 2000
        if mo < 1 or mo > 12 or d < 1 or d > 31:
            return None
        return y, mo, d
    except Exception:
        return None


def _даты_из_текста(s: str) -> list[tuple[int, int, int]]:
    out: list[tuple[int, int, int]] = []
    for m in _DATE_CELL.finditer(s):
        p = _дата_из_совпадения(m)
        if p:
            out.append(p)
    return out


def _дата_из_значения(val: object) -> list[tuple[int, int, int]]:
    if val is None:
        return []
    if isinstance(val, datetime):
        return [(val.year, val.month, val.day)]
    if isinstance(val, date):
        return [(val.year, val.month, val.day)]
    s = str(val).strip()
    if not s:
        return []
    return _даты_из_текста(s)


def _год_полный(y: int) -> int:
    if y < 100:
        return y + 2000
    return y


def _тройка_дд_мм_гг(d_s: str, m_s: str, y_s: str) -> tuple[int, int, int] | None:
    try:
        d, mo, yy = int(d_s), int(m_s), _год_полный(int(y_s))
        if mo < 1 or mo > 12 or d < 1 or d > 31:
            return None
        last = calendar.monthrange(yy, mo)[1]
        return yy, mo, min(d, last)
    except (ValueError, TypeError):
        return None


def _период_с_по_из_блоба(blob: str) -> tuple[tuple[int, int, int], tuple[int, int, int]] | None:
    one_line = re.sub(r"\s+", " ", blob.replace("\n", " ").replace("\xa0", " ")).strip()
    for rx in (_PERIOD_RANGE_HEAD, _PERIOD_RANGE_HEAD_SLASH):
        m = rx.search(one_line)
        if m:
            t1 = _тройка_дд_мм_гг(m.group("d1"), m.group("m1"), m.group("y1"))
            t2 = _тройка_дд_мм_гг(m.group("d2"), m.group("m2"), m.group("y2"))
            if t1 and t2:
                return t1, t2
    return None


def _диапазон_из_строки_несколько_дат(header_lines: list[str]) -> tuple[tuple[int, int, int], tuple[int, int, int]] | None:
    """Строки вроде «7981 … 31.03.2026 01.03.2026 31.03.2026» — последние две даты «с/по»."""
    for i, line in enumerate(header_lines):
        ds: list[tuple[int, int, int]] = []
        for m in _DATE_CELL.finditer(line):
            p = _дата_из_совпадения(m)
            if p:
                ds.append(p)
        if len(ds) >= 3:
            return ds[-2], ds[-1]
        if len(ds) == 2 and i > 0:
            prev = _norm_header(header_lines[i - 1])
            pl = "".join(prev.split())
            if ("отчетн" in prev) or prev.strip().lower() in {"с по", "спо"} or (
                "с" in pl and "по" in pl and len(prev) < 52
            ):
                return ds[0], ds[1]
    return None


def _месяцы_календарные_между(
    a: tuple[int, int, int],
    b: tuple[int, int, int],
) -> list[tuple[int, int]]:
    """Список уникальных (год, месяц) от начальной до конечной даты включительно."""
    da = date(a[0], a[1], a[2])
    db = date(b[0], b[1], b[2])
    if da > db:
        da, db = db, da
    out: list[tuple[int, int]] = []
    y, mo = da.year, da.month
    end_y, end_m = db.year, db.month
    while (y, mo) <= (end_y, end_m):
        out.append((y, mo))
        if mo == 12:
            y, mo = y + 1, 1
        else:
            mo += 1
    return out


def _число_из_строки_локаль(s0: str) -> float | None:
    """Парсит суммы из Excel при отображении строкой (РФ/ЕС): «24 000,50», «24.000», «1.234.567,89»."""
    s = s0.replace("\xa0", "").replace("\u202f", "").strip().replace(" ", "")
    s = (
        s.replace("₽", "")
        .replace("руб.", "")
        .replace("руб", "")
        .replace("р.", "")
        .replace("р", "")
        .replace("'", "")
    )
    s = re.sub(r"[^\d,.\-−]", "", s)
    neg = False
    if s.startswith("−"):
        neg = True
        s = s[1:]
    elif s.startswith("-"):
        neg = True
        s = s[1:]
    if not s:
        return None

    if "," in s:
        if "." in s:
            # Точки — группы тысяч, запятая — десятичная
            s = s.replace(".", "").replace(",", ".")
        else:
            parts = s.split(",")
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                if len(parts[1]) <= 2:
                    s = parts[0] + "." + parts[1]
                else:
                    s = "".join(parts)
            else:
                s = s.replace(",", ".")
    elif "." in s:
        parts = s.split(".")
        if len(parts) >= 2 and all(p.isdigit() for p in parts):
            if len(parts) == 2:
                a, b = parts
                if len(b) <= 2:
                    s = f"{a}.{b}"
                elif len(b) == 3:
                    if len(a) > 3:
                        s = f"{a}.{b}"
                    else:
                        s = a + b
                else:
                    try:
                        return float(s.replace(".", ""))
                    except ValueError:
                        return None
            else:
                s = "".join(parts)
        else:
            try:
                return float(s)
            except ValueError:
                return None

    try:
        x = float(s)
        return -x if neg else x
    except ValueError:
        return None


def _число_из_ячейки(val: object) -> float | None:
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, Decimal):
        try:
            return float(val)
        except Exception:
            return None
    if isinstance(val, (int, float)):
        try:
            return float(val)
        except Exception:
            return None
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    return _число_из_строки_локаль(s)


def _категории_заголовка_ячейки(text: str) -> set[str]:
    t = _norm_header(text)
    if len(t) < 2:
        return set()
    found: set[str] = set()
    if any(k in t for k in KEYS_NAME):
        found.add("name")
    if any(k in t for k in KEYS_UNIT):
        found.add("unit")
    if any(k in t for k in KEYS_QTY):
        found.add("qty")
    if any(k in t for k in KEYS_PRICE):
        found.add("price")
    if any(k in t for k in KEYS_SUM_HEADER):
        found.add("sum")
    if any(k in t for k in KEYS_DATE):
        found.add("date")
    return found


def _роль_колонки_приоритет(text: str) -> str | None:
    """Одна роль на ячейку заголовка (порядок важен). Для КС‑2 с двумя строками: «общая» — сумма, «на ед.» — цена."""
    t = _norm_header(text.replace("\n", " "))
    if not t:
        return None
    # Объединённые ячейки «Сметная стоимость…» задают блок; конкретные колонки — в следующей строке.
    if len(t) > 42 and ("стоим" in t or "сметн" in t):
        return None
    if t == "общая" or t.startswith("общая ") or (" общая " in f" {t} "):
        return "sum"
    if "на ед" in t or "наединицу" in t.replace(" ", ""):
        return "price"
    if any(k in t for k in KEYS_SUM):
        return "sum"
    if any(k in t for k in KEYS_DATE):
        return "date"
    if any(k in t for k in KEYS_PRICE):
        return "price"
    if any(k in t for k in KEYS_QTY):
        return "qty"
    if any(k in t for k in KEYS_UNIT):
        return "unit"
    if any(k in t for k in KEYS_NAME):
        return "name"
    return None


def _строка_похожа_на_итог(text: str) -> bool:
    t = _norm_header(text)
    if not t:
        return False
    if "итог" in t or "всего" in t or "всего к оплате" in t:
        return True
    return False


def _глобальный_конец_табличной_части_кс(name: str) -> bool:
    """После блока табличных строк идёт «Итого прямые…», затем трудозатраты, НДС — останавливаем сбор."""
    tl = "".join(ch for ch in _norm_header(name.replace("\n", " ")) if not ch.isspace())
    needles = (
        "итогопрямезатратыпоакт",
        "итогипоакту:",
        "итоготрудозатрат",
        "стоимость1чел.",
        "стоимость1чел/",
        "стоимость(тз+тзм)",
        "итогосучетомдоп",
        "ндс22%",
        "ндс20%",
        "ндс18%",
        "всегопоакту",
    )
    return any(k in tl for k in needles) or tl == "материалы" or tl.startswith("материалы")


def _собрать_роли_из_мульти_заголовка(
    rows: list[list[object]],
    header_idx: int,
    *,
    max_subrows: int = 6,
) -> dict[int, str]:
    col_roles: dict[int, str] = {}
    hi = min(header_idx + max_subrows, len(rows))
    for i in range(header_idx, hi):
        row = rows[i]
        for j, val in enumerate(row):
            if val is None:
                continue
            role = _роль_колонки_приоритет(str(val))
            if role:
                col_roles[j] = role
    return col_roles


def _ячейка_похожа_на_номер_позиции(val: object) -> bool:
    if val is None:
        return False
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            x = float(val)
            return x.is_integer() and 1 <= x <= 99999
        except (TypeError, ValueError, OverflowError):
            return False
    s = _norm_header(str(val).strip().replace("\n", ""))
    if re.match(r"^\d{1,5}[а-яa-z]?$", s):
        return True
    return bool(re.match(r"^\d{1,5}$", s))


def _строка_номеров_колонок_кс(row: list[object]) -> bool:
    """Строка «1, 2, 3 …» под многострочным заголовком КС‑2 — не строка позиций сметы."""
    expect = 1
    count = 0
    for j in range(min(20, len(row))):
        v = row[j]
        x = _число_из_ячейки(v)
        if x is None:
            continue
        if abs(x - expect) > 0.51:
            return False
        expect += 1
        count += 1
        if count >= 10:
            return True
    return False


def _индекс_первой_строки_данных(rows: list[list[object]], header_idx: int) -> int:
    """Пропуск строк под заголовком КС‑2 («по порядку», нумерация колонок) до первой позиции сметы."""
    for off in range(1, 14):
        i = header_idx + off
        if i >= len(rows):
            break
        row = rows[i]
        if not row:
            continue
        if _строка_номеров_колонок_кс(row):
            continue
        for j in range(min(6, len(row))):
            if _ячейка_похожа_на_номер_позиции(row[j]):
                return i
    return header_idx + 1


def _извлечь_всего_по_акту(rows: list[list[object]]) -> float | None:
    """Строка подвала «ВСЕГО по акту» — берём наибольшую сумму в хвостовых столбцах строки."""
    start = max(0, len(rows) - 220)
    label_needles = (
        "всегопоакту",
        "итогипоакту",
        "итогопоакту",
        "всегокоплате",
        "всегокоплатесндс",
        "итогосучетомндс",
    )
    for i in range(len(rows) - 1, start - 1, -1):
        row = rows[i]
        if not row:
            continue
        label_parts: list[str] = []
        for j in range(min(16, len(row))):
            if row[j] is None:
                continue
            s = str(row[j]).strip()
            if s:
                label_parts.append(_norm_header(s.replace("\n", " ")))
        merged = "".join(label_parts).replace(" ", "")
        if not (_VSEGO_PO_AKT.search(merged) or any(n in merged for n in label_needles)):
            continue

        vals: list[float] = []
        for j in range(2, len(row)):
            x = _число_из_ячейки(row[j])
            if x is None:
                continue
            ax = abs(x)
            if ax < 0.009:
                continue
            vals.append(ax)
        if not vals:
            continue
        big = [v for v in vals if v >= 99.5 or abs(v - round(v)) > 1e-6]
        if big:
            return max(big)
        return max(vals)
    return None


def _fallback_собрать_строки_листа(
    rows: list[list[object]],
    *,
    имя_файла: str,
    имя_листа: str,
    период_дата: tuple[int, int, int] | None,
) -> list[СтрокаТаблицы]:
    """Резервный сбор строк, когда заголовок таблицы не распознан."""
    def _похоже_на_наименование(s: str) -> bool:
        tl = _norm_header(s.replace("\n", " "))
        if len(tl) < 4:
            return False
        if not re.search(r"[а-яa-z]", tl):
            return False
        bad = (
            "унифицированная",
            "форма кс-2",
            "подрядчик",
            "заказчик",
            "локальный сметный расчет",
            "составлен(а) в текущих",
            "итоги по акту",
            "всего по акту",
            "стоимость в текущих",
        )
        if any(b in tl for b in bad):
            return False
        return True

    if период_дата:
        py, pm = период_дата[0], период_дата[1]
    else:
        py, pm = UNKNOWN_MONTH_KEY
    out: list[СтрокаТаблицы] = []
    empty_run = 0
    for i, row in enumerate(rows[: max(len(rows), MAX_TABLE_ROWS)]):
        if not row:
            empty_run += 1
            if empty_run >= EMPTY_STREAK_STOP + 2:
                if out:
                    break
                continue
            continue
        name_raw = ""
        for j in range(min(12, len(row))):
            v = row[j]
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            if _число_из_ячейки(v) is not None:
                continue
            if len(s) >= 3:
                name_raw = s
                break
        if name_raw and (_строка_похожа_на_итог(name_raw) or _глобальный_конец_табличной_части_кс(name_raw)):
            if out:
                break
            continue
        max_sum = _максимальная_сумма_в_ряду(row)
        if not name_raw and max_sum is None:
            empty_run += 1
            continue
        empty_run = 0
        if max_sum is None:
            continue
        if abs(max_sum) < 0.01:
            continue
        if name_raw and not _похоже_на_наименование(name_raw):
            continue
        if max_sum > 5e10:
            continue
        out.append(
            СтрокаТаблицы(
                имя_файла=имя_файла,
                лист=имя_листа,
                строка_excel=i + 1,
                год=py,
                месяц=pm,
                наименование=name_raw[:500] if name_raw else "(авто)",
                сумма=max_sum,
            )
        )
        if len(out) >= MAX_TABLE_ROWS:
            break
    # Оставляем только содержательные наборы, чтобы не ловить шум.
    return out if len(out) >= 5 else []


def _максимальная_сумма_в_ряду(row: list[object]) -> float | None:
    vals: list[float] = []
    for j in range(4, min(40, len(row))):
        x = _число_из_ячейки(row[j])
        if x is None:
            continue
        ax = abs(x)
        if ax < 0.009:
            continue
        vals.append(ax)
    if not vals:
        return None
    big = [v for v in vals if v >= 99.5 or abs(v - round(v)) > 1e-6]
    return max(big) if big else max(vals)


def _лейбл_ряда_слева(row: list[object]) -> str:
    parts: list[str] = []
    for j in range(min(6, len(row))):
        if row[j] is None:
            continue
        s = str(row[j]).strip()
        if s:
            parts.append(s)
    return " ".join(parts)


def _извлечь_суммы_ндс_подвал(rows: list[list[object]]) -> tuple[float | None, float | None, float | None]:
    """(доход без НДС, сумма НДС, всего с НДС); подвал типовой КС‑2."""
    с_ндс = _извлечь_всего_по_акту(rows)
    nds: float | None = None
    без: float | None = None
    start = max(0, len(rows) - 120)
    for i in range(len(rows) - 1, start - 1, -1):
        row = rows[i]
        if not row:
            continue
        raw = _лейбл_ряда_слева(row)
        if not raw:
            continue
        tl = _norm_header(raw.replace("\n", " "))
        tls = "".join(ch for ch in tl if not ch.isspace())
        money = _максимальная_сумма_в_ряду(row)
        if money is None:
            continue
        if _VSEGO_PO_AKT.search(tls):
            continue
        if "ндс" in tls and "%" in raw:
            nds = money if nds is None else max(nds, money)
        if "итого" in tls and ("учетом" in tls or "учётом" in tl) and "ндс" not in tls:
            без = money if без is None else max(без, money)
    if без is None and с_ндс is not None and nds is not None:
        без = max(0.0, с_ндс - nds)
    return без, nds, с_ндс


def _локатор_таблицы(rows: list[list[object]]) -> tuple[tuple[int, dict[int, str], int] | None, list[str]]:
    warns: list[str] = []
    best_i = -1
    best_score = 0
    lim = min(len(rows), MAX_TABLE_SCAN_ROW)
    for ii in range(lim):
        row = rows[ii]
        cats: set[str] = set()
        for val in row:
            if val is None:
                continue
            s = str(val).strip()
            cats |= _категории_заголовка_ячейки(s)
        score = len(cats)
        if score >= 2 and score > best_score:
            best_score = score
            best_i = ii
    if best_i < 0:
        warns.append("таблица работ не распознана (нет строки заголовка с >=2 категориями колонок)")
        return None, warns
    roles = _собрать_роли_из_мульти_заголовка(rows, best_i)
    if not any(r == "sum" for r in roles.values()) and not any(r == "name" for r in roles.values()):
        warns.append("заголовок таблицы найден, но нет колонок «наименование»/«сумма»")
        return None, warns
    fd = _индекс_первой_строки_данных(rows, best_i)
    return (best_i, roles, fd), warns


def _строка_с_номером_позиции(row: list[object]) -> bool:
    for jj in range(min(5, len(row))):
        if _ячейка_похожа_на_номер_позиции(row[jj]):
            return True
    return False


@dataclass
class УзелПозицииКс2:
    """Строка сметной части для дерева файлов."""

    дерево_id: str
    родитель_id: str | None
    строка_excel: int
    тип: str
    название: str
    единица: str
    количество: str
    сумма: float | None


def _собрать_узлы_позиций(
    rows: list[list[object]],
    *,
    first_data_row_idx: int,
    col_roles: dict[int, str],
) -> list[УзелПозицииКс2]:
    """Иерархия «раздел → работа → материалы» для обзора в дереве файлов."""
    _CONF_HIGH = 0.78
    _CONF_MID = 0.58
    _re_unit_from_template = re.compile(r"(?<!\w)(?:\d+(?:[.,]\d+)?)\s*(м2|м3|м\.п\.|мп|пм|м|шт|кг|т|компл|комплект)(?!\w)", re.IGNORECASE)
    _re_num_token = re.compile(r"-?\d+(?:[.,]\d+)?")
    _re_clean_number_line = re.compile(r"^\s*-?\d+(?:[.,]\d+)?\s*$")

    def _похоже_на_единицу(s: str) -> bool:
        t = _norm_header((s or "").replace(".", "").replace(" ", ""))
        if not t:
            return False
        known = {
            "шт",
            "м",
            "м2",
            "м3",
            "мп",
            "т",
            "кг",
            "компл",
            "комплект",
            "пм",
            "челч",
            "машч",
        }
        return t in known or t.startswith("м2") or t.startswith("м3")

    def _единица_из_шаблона(text: str) -> str:
        m = _re_unit_from_template.search(text or "")
        if not m:
            return ""
        unit = (m.group(1) or "").strip()
        return unit if _похоже_на_единицу(unit) else ""

    def _первое_число_из_текста(v: object) -> float | None:
        x = _число_из_ячейки(v)
        if x is not None:
            return x
        s = str(v or "").replace("\r", "\n")
        if not s.strip():
            return None
        # В ячейках КС-2 часто хранится:
        #   1,99516
        #   (609*1,24)
        # Берём только "чистые" числовые строки, расчёты не используем.
        for raw_line in s.split("\n"):
            line = raw_line.strip()
            if not line:
                continue
            if any(op in line for op in ("*", "+", "/", "=")):
                continue
            # Скобки в строке — обычно формула/пояснение, пропускаем.
            if "(" in line or ")" in line:
                continue
            if _re_clean_number_line.match(line):
                return _число_из_строки_локаль(line)
        # Безопасный fallback: если строка без формул, можно взять первое число.
        if not any(op in s for op in ("*", "+", "/", "=")):
            m = _re_num_token.search(s)
            if m:
                return _число_из_строки_локаль(m.group(0))
        return None

    def _infer_qty_unit(row: list[object], *, name_col_idx: int | None, sum_col_idx: int | None) -> tuple[str, str]:
        qty_out = ""
        unit_out = ""
        for j, v in enumerate(row):
            if v is None:
                continue
            if sum_col_idx is not None and j == sum_col_idx:
                continue
            if name_col_idx is not None and j == name_col_idx:
                continue
            s = str(v).strip()
            if s and not unit_out:
                u = _единица_из_шаблона(s)
                if u:
                    unit_out = u
                    continue
            if s and not unit_out and _похоже_на_единицу(s):
                unit_out = s
        return qty_out, unit_out

    def _норм_qty(x: float) -> str:
        return f"{float(x):,.4f}".rstrip("0").rstrip(".").replace(",", " ").replace(".", ",")

    def _extract_with_conf(
        row: list[object],
        *,
        unit_idx: int | None,
        qty_idx: int | None,
        name_col_idx: int | None,
        row_pos_num: float | None,
    ) -> tuple[str, str, float]:
        unit_s = ""
        qty_s = ""
        conf = 0.0
        unit_direct = False
        qty_direct = False

        if qty_idx is not None and qty_idx < len(row):
            qv = row[qty_idx]
            qx = _первое_число_из_текста(qv)
            if qx is not None:
                ax = abs(float(qx))
                if 0.0001 <= ax <= 1_000_000:
                    qty_s = _норм_qty(float(qx))
                    qty_direct = True
                    conf += 0.45
                    if abs(float(qx) - round(float(qx))) > 0.0001:
                        conf += 0.18
                    elif 1 <= int(round(ax)) <= 20:
                        conf -= 0.22
                    if row_pos_num is not None and abs(float(qx) - float(row_pos_num)) < 0.0001:
                        conf -= 0.35

        if unit_idx is not None and unit_idx < len(row):
            uv = row[unit_idx]
            us = str(uv).strip() if uv is not None else ""
            if us and _похоже_на_единицу(us):
                unit_s = us
                unit_direct = True
                conf += 0.35
            elif us:
                ut = _единица_из_шаблона(us)
                if ut:
                    unit_s = ut
                    conf += 0.28

        if not unit_s:
            for j, v in enumerate(row):
                if name_col_idx is not None and j == name_col_idx:
                    pass
                if v is None:
                    continue
                ut = _единица_из_шаблона(str(v))
                if ut:
                    unit_s = ut
                    conf += 0.22
                    break

        if qty_direct and unit_direct:
            conf += 0.08
        conf = max(0.0, min(1.0, conf))
        return unit_s, qty_s, conf

    def _колонка_похожа_на_нумерацию(rows_part: list[list[object]], col_idx: int) -> bool:
        vals: list[int] = []
        for row in rows_part:
            if col_idx >= len(row):
                continue
            x = _первое_число_из_текста(row[col_idx])
            if x is None:
                continue
            xi = int(round(float(x)))
            if abs(float(x) - xi) > 0.0001:
                continue
            vals.append(xi)
        if len(vals) < 3:
            return False
        small = [v for v in vals if 1 <= v <= 30]
        if len(small) < max(3, int(len(vals) * 0.65)):
            return False
        # Явная последовательность 1,2,3... или почти такая.
        consec = 0
        for i in range(1, len(small)):
            if small[i] == small[i - 1] + 1:
                consec += 1
        if consec >= max(2, int((len(small) - 1) * 0.7)):
            return True
        nondec = 0
        for i in range(1, len(small)):
            if small[i] >= small[i - 1]:
                nondec += 1
        return nondec >= max(3, int((len(small) - 1) * 0.72))

    name_cols = [j for j, r in col_roles.items() if r == "name"]
    sum_cols = [j for j, r in col_roles.items() if r == "sum"]
    unit_cols = [j for j, r in col_roles.items() if r == "unit"]
    qty_cols = [j for j, r in col_roles.items() if r == "qty"]
    name_col = min(name_cols) if name_cols else None
    sum_col = min(sum_cols) if sum_cols else None
    unit_col = min(unit_cols) if unit_cols else None
    qty_col = min(qty_cols) if qty_cols else None

    # Жёсткое правило по вашему требованию:
    # Кол-во брать только из 7-го столбца табличной части (1-based), индекс Python = 6.
    # Ед. обычно в 6-м (индекс 5), но для Ед. остаётся fallback.
    preferred_unit_col = 5
    preferred_qty_col = 6
    hi_pref = min(len(rows), first_data_row_idx + 260)
    pref_unit_hits = 0
    pref_qty_hits = 0
    for i in range(first_data_row_idx, hi_pref):
        row = rows[i]
        if preferred_unit_col < len(row):
            v = row[preferred_unit_col]
            if v is not None and _похоже_на_единицу(str(v).strip()):
                pref_unit_hits += 1
        if preferred_qty_col < len(row):
            v = row[preferred_qty_col]
            x = _первое_число_из_текста(v) if v is not None else None
            if x is not None and 0.0001 <= abs(float(x)) <= 1_000_000:
                pref_qty_hits += 1
    if pref_unit_hits >= 1:
        unit_col = preferred_unit_col
    qty_col = preferred_qty_col

    qty_candidates: list[int] = []
    unit_candidates: list[int] = []

    # Если роли unit/qty не определились по заголовку, доуточняем их
    # по фактическому содержимому колонок табличной части.
    if unit_col is None or qty_col is None:
        hi = min(len(rows), first_data_row_idx + 260)
        rows_part = rows[first_data_row_idx:hi]
        max_w = max((len(r) for r in rows_part), default=0)
        best_qty_col = qty_col
        best_qty_score = -1
        best_unit_col = unit_col
        best_unit_score = -1
        col_stats: dict[int, tuple[int, int, int]] = {}  # col -> (num_hits, frac_hits, small_int_hits)
        for j in range(max_w):
            if name_col is not None and j == name_col:
                continue
            if sum_col is not None and j == sum_col:
                continue
            if _колонка_похожа_на_нумерацию(rows_part, j):
                continue
            num_hits = 0
            frac_hits = 0
            small_int_hits = 0
            unit_hits = 0
            for i in range(first_data_row_idx, hi):
                row = rows[i]
                if j >= len(row):
                    continue
                v = row[j]
                if v is None:
                    continue
                x = _первое_число_из_текста(v)
                if x is not None:
                    ax = abs(float(x))
                    if 0.0001 <= ax <= 100000:
                        num_hits += 1
                        if abs(float(x) - round(float(x))) > 0.0001:
                            frac_hits += 1
                        elif 1 <= int(round(ax)) <= 20:
                            small_int_hits += 1
                    continue
                s = str(v).strip()
                if not s:
                    continue
                u_from_tpl = _единица_из_шаблона(s)
                if u_from_tpl:
                    unit_hits += 1
                    continue
                if _похоже_на_единицу(s):
                    unit_hits += 1

            # Кол-во: колонка с максимальным числом "правдоподобных" чисел.
            qty_score = num_hits * 10 + frac_hits * 4
            col_stats[j] = (num_hits, frac_hits, small_int_hits)
            if qty_score > best_qty_score and num_hits >= 2:
                best_qty_score = qty_score
                best_qty_col = j
            # Ед.: колонка с текстами-единицами.
            if unit_hits > best_unit_score and unit_hits >= 1:
                best_unit_score = unit_hits
                best_unit_col = j

        if qty_col is None and best_qty_col is not None:
            qty_col = best_qty_col
        # Валидация qty: если колонка "похожа на нумерацию", ищем более правдоподобную альтернативу.
        if qty_col is not None and qty_col in col_stats:
            num_hits, frac_hits, small_int_hits = col_stats[qty_col]
            if num_hits >= 4 and small_int_hits >= max(3, int(num_hits * 0.7)):
                alt_col = qty_col
                alt_score = -1
                for j, (n2, f2, s2) in col_stats.items():
                    if j == qty_col:
                        continue
                    # Хотим колонку с меньшей "нумерацией" и большей дробной составляющей.
                    score = (n2 * 8) + (f2 * 10) - (s2 * 6)
                    if score > alt_score and n2 >= 2 and s2 <= max(1, int(n2 * 0.6)):
                        alt_score = score
                        alt_col = j
                if alt_col != qty_col:
                    qty_col = alt_col
        if unit_col is None:
            # В КС-2 "Ед." обычно рядом с "Кол-во".
            near_unit = None
            if qty_col is not None:
                for cand in (qty_col - 1, qty_col + 1):
                    if cand < 0:
                        continue
                    hits = 0
                    for i in range(first_data_row_idx, hi):
                        row = rows[i]
                        if cand >= len(row):
                            continue
                        v = row[cand]
                        if v is None:
                            continue
                        s = str(v).strip()
                        if s and _похоже_на_единицу(s):
                            hits += 1
                    if hits >= 1:
                        near_unit = cand
                        break
            unit_col = near_unit if near_unit is not None else best_unit_col

        # Кандидаты для построчного выбора лучшей пары unit/qty.
        qty_candidates = [preferred_qty_col]
        unit_candidates = []
        if unit_col is not None:
            unit_candidates.append(unit_col)
        if qty_col is not None:
            for c in (qty_col - 1, qty_col + 1):
                if c >= 0 and c not in unit_candidates:
                    unit_candidates.append(c)
        if best_unit_col is not None and best_unit_col not in unit_candidates:
            unit_candidates.append(best_unit_col)
    else:
        qty_candidates = [preferred_qty_col]
        if unit_col is not None:
            unit_candidates = [unit_col]

    узлы: list[УзелПозицииКс2] = []
    if name_col is None and sum_col is None:
        return узлы

    current_section_id: str | None = None
    last_work_id: str | None = None
    last_unit_global: str = ""
    last_unit_by_work: dict[str, str] = {}
    last_unit_by_section: dict[str, str] = {}
    data_row = 0
    row_conf_by_excel: dict[int, float] = {}

    for ri in range(first_data_row_idx, len(rows)):
        if data_row >= MAX_TABLE_ROWS:
            break
        row = rows[ri]

        def cell(j: int) -> object:
            return row[j] if j < len(row) else None

        name_raw = ""
        if name_col is not None:
            v = cell(name_col)
            if v is not None:
                name_raw = str(v).strip()
        if not name_raw:
            for j in range(min(8, len(row))):
                if name_col is not None and j == name_col:
                    continue
                v = cell(j)
                if v is None:
                    continue
                s = str(v).strip()
                if s and not _число_из_ячейки(v):
                    name_raw = s
                    break

        if name_raw and _глобальный_конец_табличной_части_кс(name_raw):
            break

        if name_raw and _строка_похожа_на_итог(name_raw):
            continue

        sum_val: float | None = None
        if sum_col is not None:
            sum_val = _число_из_ячейки(cell(sum_col))

        pos_num = _число_из_ячейки(cell(0)) if len(row) > 0 else None
        unit_v = cell(unit_col) if unit_col is not None else None
        qty_v = cell(qty_col) if qty_col is not None else None
        unit_s = str(unit_v).strip() if unit_col is not None and unit_v is not None else ""
        qty_s = ""
        if qty_col is not None and qty_v is not None:
            qn_direct = _первое_число_из_текста(qty_v)
            if qn_direct is not None:
                qty_s = _норм_qty(float(qn_direct))
        _, _, base_conf = _extract_with_conf(
            row,
            unit_idx=unit_col,
            qty_idx=qty_col,
            name_col_idx=name_col,
            row_pos_num=pos_num,
        )
        if unit_s.lower() in ("none", "nan"):
            unit_s = ""
        if qty_s.lower() in ("none", "nan"):
            qty_s = ""
        if not unit_s:
            # Ищем unit в шаблонах вида "100 м2 ...", "1 м3 ..." по всей строке.
            for v in row:
                if v is None:
                    continue
                u = _единица_из_шаблона(str(v))
                if u:
                    unit_s = u
                    break
        # Защита от ложного попадания "номера позиции" в Ед./Кол-во:
        # если "Ед." получилась числом и не похожа на единицу измерения — очищаем.
        if unit_s:
            if _число_из_ячейки(unit_s) is not None and not _похоже_на_единицу(unit_s):
                unit_s = ""
        # Если в обе колонки попало одно и то же маленькое число (1..20), это почти
        # всегда номер/индекс, а не корректные Ед./Кол-во.
        if unit_s and qty_s and unit_s == qty_s:
            qn = _число_из_ячейки(qty_s)
            if qn is not None and abs(float(qn)) <= 20:
                unit_s = ""
        if not qty_s or not unit_s:
            qty_fb, unit_fb = _infer_qty_unit(row, name_col_idx=name_col, sum_col_idx=sum_col)
            if not qty_s and qty_fb:
                qty_s = qty_fb
            if not unit_s and unit_fb:
                unit_s = unit_fb

        # Построчный выбор лучшего кандидата (максимум заполнения при контроле риска).
        best_unit = unit_s
        best_qty = qty_s
        best_conf = base_conf
        for qc in (qty_candidates or ([qty_col] if qty_col is not None else [])):
            for uc in (unit_candidates or ([unit_col] if unit_col is not None else [])):
                cu, cq, cc = _extract_with_conf(
                    row,
                    unit_idx=uc,
                    qty_idx=qc,
                    name_col_idx=name_col,
                    row_pos_num=pos_num,
                )
                if cc > best_conf and (cq or cu):
                    best_unit, best_qty, best_conf = cu, cq, cc
        unit_s, qty_s = best_unit, best_qty
        # Пороговая стратегия: low-confidence не записываем, mid/high — заполняем.
        if best_conf < _CONF_MID:
            qty_s = ""
            unit_s = ""
        elif best_conf < _CONF_HIGH:
            # Заполняем только те части, что прошли базовую валидацию.
            if unit_s and not _похоже_на_единицу(unit_s):
                unit_s = ""
        row_conf_by_excel[ri + 1] = best_conf

        if not name_raw and sum_val is None:
            continue

        nlow = _norm_header(name_raw.replace("\n", " "))
        is_divider = nlow.startswith("раздел ") or (
            len(name_raw) < 92 and ("раздел" in nlow) and ("раздел " in nlow or nlow.startswith("раздел"))
        )
        has_posnum = _строка_с_номером_позиции(row)
        tree_id = f"r{ri + 1}"
        parent_node: str | None = None
        kind = "прочее"
        nm = name_raw[:800] if name_raw else "(без названия)"

        if is_divider:
            kind = "раздел"
            parent_node = None
            current_section_id = tree_id
            last_work_id = None
        elif has_posnum and len(name_raw) >= 8:
            kind = "работа"
            parent_node = current_section_id
            last_work_id = tree_id
        elif name_raw.strip() and last_work_id:
            kind = "материал"
            parent_node = last_work_id
        elif name_raw.strip():
            kind = "строка"
            parent_node = current_section_id

        # Часто единица измерения в КС-2 пропущена в строке, но есть в соседних
        # строках той же работы/раздела. Берём ближайший осмысленный контекст.
        if not unit_s and qty_s:
            if kind == "материал" and last_work_id and last_work_id in last_unit_by_work:
                unit_s = last_unit_by_work.get(last_work_id, "")
            elif current_section_id and current_section_id in last_unit_by_section:
                unit_s = last_unit_by_section.get(current_section_id, "")
            elif last_unit_global:
                unit_s = last_unit_global

        if unit_s:
            last_unit_global = unit_s
            if current_section_id:
                last_unit_by_section[current_section_id] = unit_s
            if kind in ("работа", "материал") and last_work_id:
                last_unit_by_work[last_work_id] = unit_s

        узлы.append(
            УзелПозицииКс2(
                дерево_id=tree_id,
                родитель_id=parent_node,
                строка_excel=ri + 1,
                тип=kind,
                название=nm,
                единица=unit_s,
                количество=qty_s.replace("\n", " "),
                сумма=sum_val,
            ),
        )
        data_row += 1

    # Второй проход: если в строке есть количество, но нет единицы,
    # пробуем взять единицу из ближайшего контекста (соседние строки).
    for i, node in enumerate(узлы):
        if node.единица or not (node.количество or "").strip() or (node.количество or "").strip() == "—":
            continue
        ctx_unit = ""
        # 1) ищем назад
        j = i - 1
        while j >= 0:
            cand = (узлы[j].единица or "").strip()
            if cand and cand != "—":
                ctx_unit = cand
                break
            j -= 1
        # 2) если назад не нашли — ищем вперёд
        if not ctx_unit:
            j = i + 1
            while j < len(узлы):
                cand = (узлы[j].единица or "").strip()
                if cand and cand != "—":
                    ctx_unit = cand
                    break
                j += 1
        if ctx_unit:
            node.единица = ctx_unit

    # AI-fallback: для "проблемных" документов, где много qty без unit.
    qty_rows = [n for n in узлы if (n.количество or "").strip() and (n.количество or "").strip() != "—"]
    missing_unit_rows = [n for n in qty_rows if not (n.единица or "").strip() or (n.единица or "").strip() == "—"]
    if qty_rows and len(missing_unit_rows) >= 3 and (len(missing_unit_rows) / max(1, len(qty_rows))) >= 0.25:
        ai_mod = _получить_ai_модуль()
        post_chat_json = getattr(ai_mod, "post_chat_json", None) if ai_mod is not None else None
        if callable(post_chat_json):
            try:
                sample_nodes = []
                for n in узлы[:180]:
                    rc = float(row_conf_by_excel.get(int(n.строка_excel), 0.0))
                    if rc >= _CONF_MID:
                        continue
                    sample_nodes.append(
                        {
                            "excel_row": int(n.строка_excel),
                            "name": (n.название or "")[:220],
                            "qty": (n.количество or "").strip(),
                            "unit": (n.единица or "").strip(),
                            "confidence": round(rc, 3),
                        }
                    )
                prompt = {
                    "task": "Заполни только пропуски unit (единица измерения) для строк, где есть qty. Не выдумывай: если не уверен — пропусти.",
                    "format": {"fills": [{"excel_row": 0, "unit": "м2"}]},
                    "rows": sample_nodes,
                }
                system = (
                    "Ты помощник по разбору КС-2. Верни ТОЛЬКО JSON-объект. "
                    "Заполняй только поле unit. Не меняй qty. "
                    "Если в строке unit неизвестен, не добавляй её в fills."
                )
                ai_out, ai_err = post_chat_json(
                    _AI_ASSIST_DIR,
                    system=system,
                    user_content=json.dumps(prompt, ensure_ascii=False),
                    temperature=0.0,
                    timeout_sec=45.0,
                )
                if not ai_err and isinstance(ai_out, dict):
                    fills = ai_out.get("fills")
                    if isinstance(fills, list):
                        by_row: dict[int, str] = {}
                        for it in fills:
                            if not isinstance(it, dict):
                                continue
                            try:
                                er = int(it.get("excel_row"))
                            except Exception:
                                continue
                            unit = str(it.get("unit") or "").strip()
                            if unit and _похоже_на_единицу(unit):
                                by_row[er] = unit
                        if by_row:
                            for n in узлы:
                                if (not (n.единица or "").strip() or (n.единица or "").strip() == "—") and n.строка_excel in by_row:
                                    # Принимаем AI только для строк низкой уверенности.
                                    if float(row_conf_by_excel.get(int(n.строка_excel), 0.0)) < _CONF_MID:
                                        n.единица = by_row[n.строка_excel]
                    _dbg_log(
                        "H1",
                        "кс2_разбор._собрать_узлы_позиций",
                        "ai_fill_units_applied",
                        {"fills": len(ai_out.get("fills") or []) if isinstance(ai_out, dict) else 0},
                    )
            except Exception as e:
                _dbg_log(
                    "H1",
                    "кс2_разбор._собрать_узлы_позиций",
                    "ai_fill_units_error",
                    {"err": str(e)[:240]},
                )

    return узлы


@dataclass
class СтрокаТаблицы:
    имя_файла: str
    лист: str
    строка_excel: int
    год: int
    месяц: int
    наименование: str
    сумма: float | None


@dataclass
class РезультатДокумента:
    путь: str
    имя: str
    листов: int
    есть_кс2: bool
    период_дата: tuple[int, int, int] | None
    все_даты: list[tuple[int, int, int]]
    фрагмент: str
    строки: list[СтрокаТаблицы] = field(default_factory=list)
    сумма_таблицы: float | None = None
    сумма_всего_по_акту: float | None = None
    выручка_по_документу: float | None = None
    месяцы_распределения_выручки: list[tuple[int, int]] = field(default_factory=list)
    доход_без_ндс: float | None = None
    сумма_ндс: float | None = None
    доход_с_ндс: float | None = None
    узлы_позиций: list[УзелПозицииКс2] = field(default_factory=list)
    предупреждения: list[str] = field(default_factory=list)
    ошибка: str | None = None


@dataclass
class СводкаМесяца:
    год: int
    месяц: int
    файлов: int
    имена_файлов: list[str]
    сумма_строк: float


def _кэш_путь_документа(path: str) -> str:
    key = hashlib.sha1(os.path.normcase(os.path.normpath(path)).encode("utf-8", errors="ignore")).hexdigest()
    return os.path.join(_CACHE_DIR, f"{key}.pkl")


def удалить_кэш_документа(path: str) -> None:
    cp = _кэш_путь_документа(path)
    try:
        if os.path.isfile(cp):
            os.remove(cp)
    except Exception:
        pass


def _прочитать_из_кэша(path: str) -> РезультатДокумента | None:
    try:
        st = os.stat(path)
    except OSError:
        return None
    cp = _кэш_путь_документа(path)
    if not os.path.isfile(cp):
        return None
    try:
        with open(cp, "rb") as f:
            payload = pickle.load(f)
        if not isinstance(payload, dict):
            return None
        if payload.get("ver") != _CACHE_VER:
            return None
        if int(payload.get("size", -1)) != int(st.st_size):
            return None
        if int(payload.get("mtime_ns", -1)) != int(st.st_mtime_ns):
            return None
        res = payload.get("result")
        if isinstance(res, РезультатДокумента):
            return res
    except Exception:
        return None
    return None


def _сохранить_в_кэш(path: str, res: РезультатДокумента) -> None:
    if res.ошибка:
        return
    try:
        st = os.stat(path)
        os.makedirs(_CACHE_DIR, exist_ok=True)
        payload = {
            "ver": _CACHE_VER,
            "size": int(st.st_size),
            "mtime_ns": int(st.st_mtime_ns),
            "result": res,
        }
        with open(_кэш_путь_документа(path), "wb") as f:
            pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception:
        pass


def _is_invalid_zero_result(res: РезультатДокумента) -> bool:
    zero_rows = len(res.строки) == 0
    no_money = (
        (res.сумма_всего_по_акту is None or abs(float(res.сумма_всего_по_акту)) < 0.01)
        and (res.сумма_таблицы is None or abs(float(res.сумма_таблицы)) < 0.01)
        and (res.выручка_по_документу is None or abs(float(res.выручка_по_документу)) < 0.01)
    )
    return bool(zero_rows and no_money)


def _dict_строка_таблицы(v: СтрокаТаблицы) -> dict[str, Any]:
    return {
        "имя_файла": v.имя_файла,
        "лист": v.лист,
        "строка_excel": int(v.строка_excel),
        "год": int(v.год),
        "месяц": int(v.месяц),
        "наименование": v.наименование,
        "сумма": float(v.сумма) if v.сумма is not None else None,
    }


def _dict_узел_позиции(v: УзелПозицииКс2) -> dict[str, Any]:
    return {
        "дерево_id": v.дерево_id,
        "родитель_id": v.родитель_id,
        "строка_excel": int(v.строка_excel),
        "тип": v.тип,
        "название": v.название,
        "единица": v.единица,
        "количество": v.количество,
        "сумма": float(v.сумма) if v.сумма is not None else None,
    }


def _dict_результат(res: РезультатДокумента) -> dict[str, Any]:
    return {
        "путь": res.путь,
        "имя": res.имя,
        "листов": int(res.листов),
        "есть_кс2": bool(res.есть_кс2),
        "период_дата": list(res.период_дата) if res.период_дата else None,
        "все_даты": [list(x) for x in res.все_даты],
        "фрагмент": res.фрагмент,
        "строки": [_dict_строка_таблицы(x) for x in res.строки],
        "сумма_таблицы": float(res.сумма_таблицы) if res.сумма_таблицы is not None else None,
        "сумма_всего_по_акту": float(res.сумма_всего_по_акту) if res.сумма_всего_по_акту is not None else None,
        "выручка_по_документу": float(res.выручка_по_документу) if res.выручка_по_документу is not None else None,
        "месяцы_распределения_выручки": [list(x) for x in res.месяцы_распределения_выручки],
        "доход_без_ндс": float(res.доход_без_ндс) if res.доход_без_ндс is not None else None,
        "сумма_ндс": float(res.сумма_ндс) if res.сумма_ндс is not None else None,
        "доход_с_ндс": float(res.доход_с_ндс) if res.доход_с_ндс is not None else None,
        "узлы_позиций": [_dict_узел_позиции(x) for x in res.узлы_позиций],
        "предупреждения": list(res.предупреждения),
        "ошибка": res.ошибка,
    }


def _результат_из_dict(raw: dict[str, Any]) -> РезультатДокумента:
    строки = [
        СтрокаТаблицы(
            имя_файла=str(x.get("имя_файла") or ""),
            лист=str(x.get("лист") or ""),
            строка_excel=int(x.get("строка_excel") or 0),
            год=int(x.get("год") or 0),
            месяц=int(x.get("месяц") or 0),
            наименование=str(x.get("наименование") or ""),
            сумма=(float(x["сумма"]) if x.get("сумма") is not None else None),
        )
        for x in (raw.get("строки") or [])
        if isinstance(x, dict)
    ]
    узлы = [
        УзелПозицииКс2(
            дерево_id=str(x.get("дерево_id") or ""),
            родитель_id=(str(x["родитель_id"]) if x.get("родитель_id") else None),
            строка_excel=int(x.get("строка_excel") or 0),
            тип=str(x.get("тип") or "прочее"),
            название=str(x.get("название") or ""),
            единица=(str(x["единица"]) if x.get("единица") else None),
            количество=(str(x["количество"]) if x.get("количество") else None),
            сумма=(float(x["сумма"]) if x.get("сумма") is not None else None),
        )
        for x in (raw.get("узлы_позиций") or [])
        if isinstance(x, dict)
    ]
    период = raw.get("период_дата")
    все_даты = raw.get("все_даты") or []
    мес = raw.get("месяцы_распределения_выручки") or []
    return РезультатДокумента(
        путь=str(raw.get("путь") or ""),
        имя=str(raw.get("имя") or ""),
        листов=int(raw.get("листов") or 0),
        есть_кс2=bool(raw.get("есть_кс2")),
        период_дата=(tuple(int(v) for v in период) if isinstance(период, (list, tuple)) and len(период) == 3 else None),  # type: ignore[arg-type]
        все_даты=[tuple(int(v) for v in x) for x in все_даты if isinstance(x, (list, tuple)) and len(x) == 3],  # type: ignore[list-item]
        фрагмент=str(raw.get("фрагмент") or ""),
        строки=строки,
        сумма_таблицы=(float(raw["сумма_таблицы"]) if raw.get("сумма_таблицы") is not None else None),
        сумма_всего_по_акту=(float(raw["сумма_всего_по_акту"]) if raw.get("сумма_всего_по_акту") is not None else None),
        выручка_по_документу=(float(raw["выручка_по_документу"]) if raw.get("выручка_по_документу") is not None else None),
        месяцы_распределения_выручки=[tuple(int(v) for v in x) for x in мес if isinstance(x, (list, tuple)) and len(x) == 2],  # type: ignore[list-item]
        доход_без_ндс=(float(raw["доход_без_ндс"]) if raw.get("доход_без_ндс") is not None else None),
        сумма_ндс=(float(raw["сумма_ндс"]) if raw.get("сумма_ндс") is not None else None),
        доход_с_ндс=(float(raw["доход_с_ндс"]) if raw.get("доход_с_ндс") is not None else None),
        узлы_позиций=узлы,
        предупреждения=[str(x) for x in (raw.get("предупреждения") or []) if isinstance(x, str)],
        ошибка=(str(raw["ошибка"]) if raw.get("ошибка") else None),
    )


def _хэш_результата(res: РезультатДокумента) -> str:
    payload = _dict_результат(res)
    payload["путь"] = ""
    body = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha1(body.encode("utf-8", errors="ignore")).hexdigest()


def получить_хэш_результата(res: РезультатДокумента) -> str:
    return _хэш_результата(res)


def _empty_store() -> dict[str, Any]:
    return {"items_by_hash": {}, "index_by_path": {}, "meta_by_hash": {}}


def _загрузить_store() -> dict[str, Any]:
    out = _empty_store()
    if not os.path.isfile(_STORE_PATH):
        return out
    try:
        with open(_STORE_PATH, "rb") as f:
            payload = pickle.load(f)
        if not isinstance(payload, dict):
            return out
        ver = int(payload.get("ver", -1))
        if ver == _STORE_VER:
            raw_items = payload.get("items_by_hash")
            raw_index = payload.get("index_by_path")
            raw_meta = payload.get("meta_by_hash")
            if not isinstance(raw_items, dict):
                return out
            items: dict[str, РезультатДокумента] = {}
            for k, v in raw_items.items():
                if isinstance(k, str) and isinstance(v, РезультатДокумента):
                    items[k] = v
            index: dict[str, str] = {}
            if isinstance(raw_index, dict):
                for p, h in raw_index.items():
                    if isinstance(p, str) and isinstance(h, str) and h in items:
                        index[os.path.normpath(p)] = h
            meta: dict[str, dict[str, Any]] = {}
            if isinstance(raw_meta, dict):
                for h, m in raw_meta.items():
                    if isinstance(h, str) and isinstance(m, dict):
                        meta[h] = m
            out["items_by_hash"] = items
            out["index_by_path"] = index
            out["meta_by_hash"] = meta
            return out
        # Миграция: старый формат ver=2 с ключом path -> result
        raw_old = payload.get("items")
        if isinstance(raw_old, dict):
            for p, v in raw_old.items():
                if not isinstance(p, str) or not isinstance(v, РезультатДокумента):
                    continue
                h = _хэш_результата(v)
                out["items_by_hash"][h] = v
                out["index_by_path"][os.path.normpath(p)] = h
                out["meta_by_hash"][h] = {
                    "source": "legacy_path_store",
                    "imported_at": datetime.now().isoformat(timespec="seconds"),
                    "name": v.имя or os.path.basename(p),
                }
            _сохранить_store(out)
            return out
    except Exception:
        return out
    return out


def _сохранить_store(store: dict[str, Any]) -> None:
    try:
        os.makedirs(os.path.dirname(_STORE_PATH), exist_ok=True)
        payload = {
            "ver": _STORE_VER,
            "items_by_hash": store.get("items_by_hash", {}),
            "index_by_path": store.get("index_by_path", {}),
            "meta_by_hash": store.get("meta_by_hash", {}),
        }
        with open(_STORE_PATH, "wb") as f:
            pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception:
        pass


def получить_все_сохраненные() -> list[РезультатДокумента]:
    store = _загрузить_store()
    items = store.get("items_by_hash", {})
    if not isinstance(items, dict):
        return []
    out: list[РезультатДокумента] = []
    dirty = False
    for h, res in list(items.items()):
        if not isinstance(h, str) or not isinstance(res, РезультатДокумента):
            continue
        if _is_invalid_zero_result(res):
            items.pop(h, None)
            dirty = True
            continue
        out.append(res)
    if dirty:
        _сохранить_store(store)
    return out


def получить_сохраненный_результат(path: str) -> РезультатДокумента | None:
    key = os.path.normpath(path)
    store = _загрузить_store()
    items = store.get("items_by_hash", {})
    index = store.get("index_by_path", {})
    h = index.get(key) if isinstance(index, dict) else None
    res = items.get(h) if isinstance(items, dict) and isinstance(h, str) else None
    if isinstance(res, РезультатДокумента) and _is_invalid_zero_result(res):
        if isinstance(items, dict):
            items.pop(h, None)
        if isinstance(index, dict):
            index.pop(key, None)
        _сохранить_store(store)
        #region agent log
        _dbg_log("H1", "кс2_разбор.получить_сохраненный_результат", "store_invalidated_zero_result", {"path": key})
        #endregion
        res = None
    #region agent log
    _dbg_log("H1", "кс2_разбор.получить_сохраненный_результат", "store_lookup", {"path": key, "hit": bool(res), "rows": len(res.строки) if res else 0})
    #endregion
    return res


def сохранить_результат_по_хэшу(res: РезультатДокумента, *, source: str = "app", source_path: str | None = None) -> str:
    if res.ошибка:
        return ""
    store = _загрузить_store()
    items = store.get("items_by_hash", {})
    index = store.get("index_by_path", {})
    meta = store.get("meta_by_hash", {})
    if not isinstance(items, dict) or not isinstance(index, dict) or not isinstance(meta, dict):
        store = _empty_store()
        items = store["items_by_hash"]
        index = store["index_by_path"]
        meta = store["meta_by_hash"]
    h = _хэш_результата(res)
    items[h] = res
    src_p = source_path or res.путь
    if src_p:
        index[os.path.normpath(src_p)] = h
    if res.путь:
        index[os.path.normpath(res.путь)] = h
    meta[h] = {
        "source": source,
        "imported_at": datetime.now().isoformat(timespec="seconds"),
        "name": res.имя or os.path.basename(src_p or res.путь or h),
    }
    _сохранить_store(store)
    return h


def сохранить_сохраненный_результат(path: str, res: РезультатДокумента) -> None:
    if res.ошибка:
        return
    сохранить_результат_по_хэшу(res, source="app", source_path=path)


def удалить_сохраненный_результат(path: str) -> None:
    key = os.path.normpath(path)
    store = _загрузить_store()
    index = store.get("index_by_path", {})
    items = store.get("items_by_hash", {})
    meta = store.get("meta_by_hash", {})
    if not isinstance(index, dict):
        return
    h = index.pop(key, None)
    if isinstance(h, str) and isinstance(items, dict):
        # если хэш больше не связан ни с одним путём — удаляем запись полностью
        still_used = any(v == h for v in index.values())
        if not still_used:
            items.pop(h, None)
            if isinstance(meta, dict):
                meta.pop(h, None)
    _сохранить_store(store)


def удалить_сохраненный_по_хэшу(content_hash: str) -> None:
    if not content_hash:
        return
    store = _загрузить_store()
    items = store.get("items_by_hash", {})
    index = store.get("index_by_path", {})
    meta = store.get("meta_by_hash", {})
    if isinstance(items, dict):
        items.pop(content_hash, None)
    if isinstance(index, dict):
        for p in [k for k, v in index.items() if v == content_hash]:
            index.pop(p, None)
    if isinstance(meta, dict):
        meta.pop(content_hash, None)
    _сохранить_store(store)


def сбросить_сохраненные_результаты() -> int:
    store = _загрузить_store()
    items = store.get("items_by_hash", {})
    n = len(items) if isinstance(items, dict) else 0
    try:
        if os.path.isfile(_STORE_PATH):
            os.remove(_STORE_PATH)
    except Exception:
        pass
    return n


def импортировать_результат_из_json(json_path: str, *, source: str = "chat") -> str:
    with open(json_path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    if not isinstance(payload, dict):
        raise ValueError("JSON должен содержать объект с ключом 'result'.")
    raw = payload.get("result")
    if not isinstance(raw, dict):
        raise ValueError("JSON не содержит корректный объект 'result'.")
    res = _результат_из_dict(raw)
    if not res.имя:
        res.имя = os.path.basename(json_path)
    return сохранить_результат_по_хэшу(res, source=source)


def экспортировать_результат_в_json(res: РезультатДокумента, out_path: str) -> None:
    payload = {"schema": "ks2_result_v1", "content_hash": _хэш_результата(res), "result": _dict_результат(res)}
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def сбросить_кэш_документов() -> None:
    try:
        if os.path.isdir(_CACHE_DIR):
            for name in os.listdir(_CACHE_DIR):
                p = os.path.join(_CACHE_DIR, name)
                if os.path.isfile(p):
                    try:
                        os.remove(p)
                    except Exception:
                        pass
    except Exception:
        pass


def разобрать_или_взять_сохраненный(path: str) -> РезультатДокумента:
    saved = получить_сохраненный_результат(path)
    if saved is not None:
        #region agent log
        _dbg_log("H1", "кс2_разбор.разобрать_или_взять_сохраненный", "use_store", {"path": os.path.normpath(path), "rows": len(saved.строки)})
        #endregion
        return saved
    #region agent log
    _dbg_log("H1", "кс2_разбор.разобрать_или_взять_сохраненный", "store_miss_parse", {"path": os.path.normpath(path)})
    #endregion
    res = разобрать_документ(path)
    if not res.ошибка:
        сохранить_сохраненный_результат(path, res)
    return res


def _собрать_строки_листа(
    rows: list[list[object]],
    *,
    имя_файла: str,
    имя_листа: str,
    first_data_row_idx: int,
    col_roles: dict[int, str],
    период_гм: tuple[int, int],
) -> list[СтрокаТаблицы]:
    name_cols = [j for j, r in col_roles.items() if r == "name"]
    sum_cols = [j for j, r in col_roles.items() if r == "sum"]
    date_cols = [j for j, r in col_roles.items() if r == "date"]
    name_col = min(name_cols) if name_cols else None
    sum_col = min(sum_cols) if sum_cols else None

    out: list[СтрокаТаблицы] = []
    if sum_col is None and not name_col:
        return out

    empty_run = 0
    data_row = 0
    for i in range(first_data_row_idx, len(rows)):
        if data_row >= MAX_TABLE_ROWS:
            break
        row = rows[i] if i < len(rows) else []
        # расширить row до нужной длины
        def cell(j: int) -> object:
            return row[j] if j < len(row) else None

        name_raw = ""
        if name_col is not None:
            v = cell(name_col)
            if v is not None:
                name_raw = str(v).strip()
        if not name_raw:
            for j in range(min(8, len(row))):
                if name_col is not None and j == name_col:
                    continue
                v = cell(j)
                if v is None:
                    continue
                s = str(v).strip()
                if s and not _число_из_ячейки(v):
                    name_raw = s
                    break

        if name_raw and _глобальный_конец_табличной_части_кс(name_raw):
            break

        if name_raw and _строка_похожа_на_итог(name_raw):
            continue

        sum_val: float | None = None
        if sum_col is not None:
            sum_val = _число_из_ячейки(cell(sum_col))

        if not name_raw and sum_val is None:
            empty_run += 1
            if empty_run >= EMPTY_STREAK_STOP:
                break
            continue
        empty_run = 0

        if not name_raw and sum_val is None:
            continue

        y, m = период_гм
        if date_cols:
            for dj in date_cols:
                ds = _дата_из_значения(cell(dj))
                if ds:
                    y, m, _d = ds[0]
                    break

        out.append(
            СтрокаТаблицы(
                имя_файла=имя_файла,
                лист=имя_листа,
                строка_excel=i + 1,
                год=y,
                месяц=m,
                наименование=name_raw[:500] if name_raw else "",
                сумма=sum_val,
            )
        )
        data_row += 1

    return out


def _найти_таблицу_на_листе(
    rows: list[list[object]],
    *,
    имя_файла: str,
    имя_листа: str,
    период_дата: tuple[int, int, int] | None,
) -> tuple[list[СтрокаТаблицы], list[УзелПозицииКс2], list[str]]:
    period_gm = (
        (период_дата[0], период_дата[1])
        if период_дата
        else UNKNOWN_MONTH_KEY
    )
    ctx, warns = _локатор_таблицы(rows)
    if not ctx:
        fb = _fallback_собрать_строки_листа(
            rows,
            имя_файла=имя_файла,
            имя_листа=имя_листа,
            период_дата=период_дата,
        )
        if fb:
            warns.append(f"таблица распознана fallback-режимом: {len(fb)} строк")
            return fb, [], warns
        return [], [], warns

    _, col_roles, fd = ctx
    lines = _собрать_строки_листа(
        rows,
        имя_файла=имя_файла,
        имя_листа=имя_листа,
        first_data_row_idx=fd,
        col_roles=col_roles,
        период_гм=period_gm,
    )
    узлы = _собрать_узлы_позиций(rows, first_data_row_idx=fd, col_roles=col_roles)
    if not lines:
        fb = _fallback_собрать_строки_листа(
            rows,
            имя_файла=имя_файла,
            имя_листа=имя_листа,
            период_дата=период_дата,
        )
        if fb:
            warns.append(f"под заголовком пусто; применён fallback: {len(fb)} строк")
            return fb, узлы, warns
        warns.append("под заголовком таблицы не найдено ни одной содержательной строки")
    return lines, узлы, warns


def _лист_в_матрицу(ws: object, max_r: int, max_c: int) -> list[list[object]]:
    out: list[list[object]] = []
    mr = min(getattr(ws, "max_row", None) or max_r, max_r)
    mc = min(getattr(ws, "max_column", None) or max_c, max_c)
    for row in ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc, values_only=True):
        out.append(list(row))
    return out


def разобрать_документ(path: str) -> РезультатДокумента:
    #region agent log
    _dbg_log("H2", "кс2_разбор.разобрать_документ", "enter", {"path": os.path.normpath(path)})
    #endregion
    имя = os.path.basename(path)
    if not os.path.isfile(path):
        return РезультатДокумента(
            путь=path,
            имя=имя,
            листов=0,
            есть_кс2=False,
            период_дата=None,
            все_даты=[],
            фрагмент="—",
            ошибка="файл не найден",
        )

    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        return РезультатДокумента(
            путь=path,
            имя=имя,
            листов=0,
            есть_кс2=False,
            период_дата=None,
            все_даты=[],
            фрагмент="—",
            ошибка="формат .xls: сохраните как .xlsx",
        )
    if ext != ".xlsx":
        return РезультатДокумента(
            путь=path,
            имя=имя,
            листов=0,
            есть_кс2=False,
            период_дата=None,
            все_даты=[],
            фрагмент="—",
            ошибка="ожидается .xlsx",
        )

    cached = _прочитать_из_кэша(path)
    if cached is not None:
        zero_rows = len(cached.строки) == 0
        no_money = (
            (cached.сумма_всего_по_акту is None or abs(float(cached.сумма_всего_по_акту)) < 0.01)
            and (cached.сумма_таблицы is None or abs(float(cached.сумма_таблицы)) < 0.01)
            and (cached.выручка_по_документу is None or abs(float(cached.выручка_по_документу)) < 0.01)
        )
        if zero_rows and no_money:
            #region agent log
            _dbg_log("H3", "кс2_разбор.разобрать_документ", "doc_cache_invalidated_zero_result", {"path": os.path.normpath(path)})
            #endregion
            удалить_кэш_документа(path)
        else:
            #region agent log
            _dbg_log("H3", "кс2_разбор.разобрать_документ", "doc_cache_hit", {"path": os.path.normpath(path), "rows": len(cached.строки)})
            #endregion
            return cached

    try:
        from openpyxl import load_workbook
    except ImportError:
        return РезультатДокумента(
            путь=path,
            имя=имя,
            листов=0,
            есть_кс2=False,
            период_дата=None,
            все_даты=[],
            фрагмент="—",
            ошибка="не установлен openpyxl",
        )

    все_даты: list[tuple[int, int, int]] = []
    ks2 = False
    chunk: list[str] = []
    header_lines: list[str] = []
    первая_в_шапке: tuple[int, int, int] | None = None
    предупреждения: list[str] = []

    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception as e:
        return РезультатДокумента(
            путь=path,
            имя=имя,
            листов=0,
            есть_кс2=False,
            период_дата=None,
            все_даты=[],
            фрагмент="—",
            ошибка=f"ошибка чтения: {e}",
        )

    sheets = wb.sheetnames
    n_sh = len(sheets)

    матрицы_листов: list[tuple[str, list[list[object]]]] = []
    try:
        for si, sheet_name in enumerate(sheets[:MAX_SHEETS_DATE_SCAN]):
            ws = wb[sheet_name]
            sheet_max_r = BODY_MAX_ROW_SHEET1 if si == 0 else MAX_ROW_OTHER_SHEETS
            max_r = min(getattr(ws, "max_row", None) or sheet_max_r, sheet_max_r)
            max_c = min(getattr(ws, "max_column", None) or BODY_MAX_COL, BODY_MAX_COL)
            ridx = 0
            mat_sheet: list[list[object]] = []
            for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c, values_only=True):
                ridx += 1
                row_list = list(row)
                mat_sheet.append(row_list)
                row_bits: list[str] = []
                for val in row_list:
                    if val is None:
                        continue
                    s = str(val).strip()
                    if not s:
                        continue
                    row_bits.append(re.sub(r"\s+", " ", s))
                    if _KS2.search(s):
                        ks2 = True
                    ds_here = _дата_из_значения(val)
                    if si == 0 and ridx <= HEADER_MAX_ROW:
                        if ds_here and первая_в_шапке is None:
                            первая_в_шапке = ds_here[0]
                        if len(chunk) < 14 and len(s) < 200:
                            chunk.append(s[:120])
                    elif len(chunk) < 18 and len(s) < 200:
                        chunk.append(s[:120])
                    for dt in ds_here:
                        if _дата_правдоподобна(dt):
                            все_даты.append(dt)
                if si == 0 and ridx <= HEADER_MAX_ROW and row_bits:
                    header_lines.append(" ".join(row_bits)[:500])
            матрицы_листов.append((sheet_name, mat_sheet))
            # Не прерываемся досрочно: в ряде КС-2 табличная часть лежит на следующих листах.
    finally:
        wb.close()

    период_дата = первая_в_шапке if (первая_в_шапке and _дата_правдоподобна(первая_в_шапке)) else None
    if период_дата is None and все_даты:
        период_дата = max(все_даты)
    if период_дата is None:
        предупреждения.append("дата периода не найдена (ДД.ММ.ГГГГ в шапке или в теле)")

    месяцы_распределения_выручки: list[tuple[int, int]] = []
    rng_head = None
    for hl in header_lines:
        rng_head = _период_с_по_из_блоба(hl)
        if rng_head:
            break
    if rng_head is None:
        blob_h = "\n".join(header_lines)
        if blob_h.strip():
            rng_head = _период_с_по_из_блоба(blob_h)
    if rng_head is None and header_lines:
        rng_head = _диапазон_из_строки_несколько_дат(header_lines)
    if rng_head:
        ta, tb = rng_head
        if date(ta[0], ta[1], ta[2]) > date(tb[0], tb[1], tb[2]):
            ta, tb = tb, ta
        месяцы_распределения_выручки = _месяцы_календарные_между(ta, tb)
        if not месяцы_распределения_выручки:
            месяцы_распределения_выручки = [(tb[0], tb[1])]
        период_дата = (tb[0], tb[1], tb[2])

    frag = " · ".join(chunk[:5]) if chunk else "—"

    # Разбор таблицы по собранным матрицам листов (без второго чтения файла).
    строки: list[СтрокаТаблицы] = []
    сумма_всего_по_акту: float | None = None
    доход_без_ндс: float | None = None
    сумма_ндс: float | None = None
    доход_с_ндс: float | None = None
    узлы_raw: list[УзелПозицииКс2] = []
    best_warns: list[str] = []
    try:
        if матрицы_листов:
            best_lines: list[СтрокаТаблицы] = []
            best_nodes: list[УзелПозицииКс2] = []
            best_score: float = -1.0
            best_footer: tuple[float | None, float | None, float | None] = (None, None, None)
            non_empty_seen = False
            for sheet_name, mat in матрицы_листов:
                try:
                    без_nds, nds_val, с_nds = _извлечь_суммы_ндс_подвал(mat)

                    таб, узлы, wtab = _найти_таблицу_на_листе(
                        mat,
                        имя_файла=имя,
                        имя_листа=sheet_name,
                        период_дата=период_дата,
                    )
                    #region agent log
                    _dbg_log(
                        "H4",
                        "кс2_разбор.sheet_eval",
                        "sheet_candidate",
                        {
                            "file": имя,
                            "sheet": sheet_name,
                            "rows": len(таб),
                            "sum_tab": round(sum((ln.сумма or 0.0) for ln in таб), 2),
                            "sum_footer": с_nds,
                            "warns": wtab[:2],
                        },
                    )
                    #endregion
                    tab_sum = sum((ln.сумма or 0.0) for ln in таб)
                    fallback_used = any("fallback" in _norm_header(w) for w in wtab)
                    score = float(len(таб))
                    if с_nds is not None and таб:
                        ratio = (tab_sum / с_nds) if с_nds else 0.0
                        if 0.3 <= ratio <= 3.0:
                            score += 180.0
                        elif 0.1 <= ratio <= 6.0:
                            score += 80.0
                    if с_nds is not None:
                        score += 25.0
                    if fallback_used:
                        score -= 20.0

                    # Если есть хоть один лист с реальной таблицей, предпочитаем только такие листы.
                    if таб:
                        if (not non_empty_seen) or (score > best_score):
                            non_empty_seen = True
                            best_score = score
                            best_lines = таб
                            best_nodes = узлы
                            best_warns = wtab
                            best_footer = (без_nds, nds_val, с_nds)
                        continue

                    if not non_empty_seen and score > best_score:
                        best_score = score
                        best_lines = таб
                        best_nodes = узлы
                        best_warns = wtab
                        best_footer = (без_nds, nds_val, с_nds)
                except Exception as sheet_err:
                    #region agent log
                    _dbg_log(
                        "H4",
                        "кс2_разбор.sheet_eval",
                        "sheet_error_continue",
                        {
                            "file": имя,
                            "sheet": sheet_name,
                            "err": str(sheet_err)[:240],
                            "tb": traceback.format_exc()[-1200:],
                        },
                    )
                    #endregion
                    continue
            строки.extend(best_lines)
            узлы_raw = best_nodes
            предупреждения.extend(best_warns)
            доход_без_ндс, сумма_ндс, доход_с_ндс = best_footer
            сумма_всего_по_акту = доход_с_ндс
    except Exception as e:
        предупреждения.append(f"разбор табличной части: {e}")

    сумма_таблицы: float | None = None
    parts: list[float] = []
    for ln in строки:
        if ln.сумма is not None:
            parts.append(ln.сумма)
    if parts:
        сумма_таблицы = sum(parts)

    выручка_по_документу = сумма_всего_по_акту if сумма_всего_по_акту is not None else сумма_таблицы

    result = РезультатДокумента(
        путь=path,
        имя=имя,
        листов=n_sh,
        есть_кс2=ks2,
        период_дата=период_дата,
        все_даты=все_даты,
        фрагмент=frag[:240],
        строки=строки,
        сумма_таблицы=сумма_таблицы,
        сумма_всего_по_акту=сумма_всего_по_акту,
        выручка_по_документу=выручка_по_документу,
        месяцы_распределения_выручки=месяцы_распределения_выручки,
        доход_без_ндс=доход_без_ндс,
        сумма_ндс=сумма_ндс,
        доход_с_ндс=доход_с_ндс,
        узлы_позиций=узлы_raw,
        предупреждения=предупреждения,
        ошибка=None,
    )
    #region agent log
    _dbg_log(
        "H5",
        "кс2_разбор.разобрать_документ",
        "result_summary",
        {
            "path": os.path.normpath(path),
            "rows": len(result.строки),
            "sum_table": result.сумма_таблицы,
            "sum_act": result.сумма_всего_по_акту,
            "revenue": result.выручка_по_документу,
            "warns": result.предупреждения[:3],
        },
    )
    #endregion
    _сохранить_в_кэш(path, result)
    return result


def разобрать_пакет(paths: Iterable[str]) -> list[РезультатДокумента]:
    return [разобрать_документ(p) for p in paths]


def агрегировать_по_месяцам(документы: list[РезультатДокумента]) -> list[СводкаМесяца]:
    """Выручка каждого акта делится поровну на все месяцы периода «с … по …» из шапки; иначе один месяц по полю периода."""
    files_by_month: dict[tuple[int, int], list[str]] = {}
    sum_by_month: dict[tuple[int, int], float] = {}

    for d in документы:
        if d.ошибка:
            continue
        amt = d.выручка_по_документу if d.выручка_по_документу is not None else 0.0
        amt = float(amt)
        months_used = (
            list(d.месяцы_распределения_выручки)
            if d.месяцы_распределения_выручки
            else (
                [(d.период_дата[0], d.период_дата[1])]
                if d.период_дата
                else [UNKNOWN_MONTH_KEY]
            )
        )
        portion = amt / len(months_used) if months_used else 0.0
        for key in months_used:
            files_by_month.setdefault(key, []).append(d.имя)
            sum_by_month[key] = sum_by_month.get(key, 0.0) + portion

    keys = set(files_by_month.keys()) | set(sum_by_month.keys())
    out: list[СводкаМесяца] = []
    for gy, mo in sorted(keys, key=lambda t: (t[0] if t[0] else 9999, t[1] if t[1] else 13)):
        raw_names = files_by_month.get((gy, mo), [])
        names = sorted(set(raw_names))
        out.append(
            СводкаМесяца(
                год=gy,
                месяц=mo,
                файлов=len(names),
                имена_файлов=names,
                сумма_строк=sum_by_month.get((gy, mo), 0.0),
            )
        )
    return out


def _метка_месяца(г: int, m: int) -> str:
    if (г, m) == UNKNOWN_MONTH_KEY:
        return "неизвестный период"
    return f"{m:02d}.{г:04d}"


_MONTHS_RU = (
    "",
    "янв",
    "фев",
    "мар",
    "апр",
    "май",
    "июн",
    "июл",
    "авг",
    "сен",
    "окт",
    "ноя",
    "дек",
)


def подпись_месяца_краткая(год: int, месяц: int) -> str:
    """Читаемая подпись оси для дашборда (акты по периоду)."""
    if (год, месяц) == UNKNOWN_MONTH_KEY:
        return "без периода"
    if 1 <= месяц <= 12:
        return f"{_MONTHS_RU[месяц]} {год}"
    return _метка_месяца(год, месяц)


def всего_выручка_по_сводке(сводка: list[СводкаМесяца]) -> float:
    """Сумма выручки по месяцам (итог по акту или сумма строк «общая»)."""
    return float(sum(s.сумма_строк for s in сводка))


def текстовый_отчёт(документы: list[РезультатДокумента], сводка: list[СводкаМесяца]) -> str:
    lines: list[str] = []
    lines.append(
        "━━ Сводка по месяцам (период «с … по …» из шапки — выручка акта делится поровну по месяцам; "
        "иначе один месяц по дате периода) ━━\n\n",
    )
    if not сводка:
        lines.append("Нет данных для сводки.\n\n")
    else:
        for s in сводка:
            label = _метка_месяца(s.год, s.месяц)
            lines.append(f"• {label} — документов: {s.файлов}")
            if s.сумма_строк:
                lines.append(f"  Сумма по актам за месяц: {s.сумма_строк:,.2f}".replace(",", " "))
            if s.имена_файлов:
                show = ", ".join(s.имена_файлов[:12])
                if len(s.имена_файлов) > 12:
                    show += f" … (+{len(s.имена_файлов) - 12})"
                lines.append(f"  Файлы: {show}")
            lines.append("")

    lines.append("━━ По документам (сортировка по дате периода, затем по имени) ━━\n\n")

    def sort_key(dd: РезультатДокумента) -> tuple:
        if dd.ошибка:
            return (9999, 99, 99, dd.имя.lower())
        if dd.период_дата:
            return (*dd.период_дата, dd.имя.lower())
        return (9999, 99, 99, dd.имя.lower())

    sorted_docs = sorted(документы, key=sort_key)
    for i, d in enumerate(sorted_docs, start=1):
        lines.append(f"{i}. {d.имя}")
        if d.ошибка:
            lines.append(f"   Ошибка: {d.ошибка}\n")
            continue
        if d.период_дата:
            p = d.период_дата
            ds = f"{p[2]:02d}.{p[1]:02d}.{p[0]:04d}"
        else:
            ds = "нет"
        lines.append(
            f"   Листов: {d.листов} | Период: {ds} | «КС-2» в тексте: {'да' if d.есть_кс2 else 'нет'}",
        )
        if d.выручка_по_документу is not None:
            lines.append(
                f"   Выручка по документу: {d.выручка_по_документу:,.2f}".replace(",", " ")
                + (
                    " (строка «ВСЕГО по акту»)"
                    if d.сумма_всего_по_акту is not None
                    else " (сумма колонки «общая» по строкам)"
                ),
            )
        if len(d.месяцы_распределения_выручки) > 1:
            lab = ", ".join(_метка_месяца(y, m) for y, m in d.месяцы_распределения_выручки[:24])
            if len(d.месяцы_распределения_выручки) > 24:
                lab += " …"
            lines.append(f"   По графику (доля месяца = 1/{len(d.месяцы_распределения_выручки)} суммы): {lab}")
        if d.строки:
            st = f"{len(d.строки)} строк"
            if d.сумма_таблицы is not None:
                st += f", сумма «общая» {d.сумма_таблицы:,.2f}".replace(",", " ")
            lines.append(f"   Табличная часть: {st}")
        else:
            lines.append("   Табличная часть: не распознана")
        if d.предупреждения:
            for w in d.предупреждения:
                lines.append(f"   [!] {w}")
        lines.append(f"   Фрагмент: {d.фрагмент}\n")

    lines.append("━━ Табличная часть (все извлечённые строки) ━━\n\n")
    total_lines = sum(len(d.строки) for d in документы if not d.ошибка)
    if total_lines == 0:
        lines.append("Строк таблицы не найдены — возможен нестандартный макет КС-2.\n")
    else:
        lines.append(f"Всего строк: {total_lines}\n")
        for d in sorted_docs:
            if d.ошибка or not d.строки:
                continue
            lines.append(f"— {d.имя} ({len(d.строки)} стр.)")
            for ln in d.строки[:25]:
                sm = f"{ln.сумма:,.2f}".replace(",", " ") if ln.сумма is not None else "—"
                nm = ln.наименование[:72] + ("…" if len(ln.наименование) > 72 else "")
                lines.append(f"   {ln.строка_excel}: [{_метка_месяца(ln.год, ln.месяц)}] сумма={sm}  {nm}")
            if len(d.строки) > 25:
                lines.append(f"   … ещё {len(d.строки) - 25} строк")
            lines.append("")

    return "".join(lines)


def сохранить_сводку_xlsx(path: str, документы: list[РезультатДокумента], сводка: list[СводкаМесяца]) -> None:
    from openpyxl import Workbook

    wb = Workbook(write_only=False)
    ws0 = wb.active
    ws0.title = "По_месяцам"
    ws0.append(["Год", "Месяц", "Период", "Документов", "Сумма_выручки", "Файлы"])
    for s in сводка:
        label = _метка_месяца(s.год, s.месяц)
        files_cell = "; ".join(s.имена_файлов[:40])
        if len(s.имена_файлов) > 40:
            files_cell += f" …(+{len(s.имена_файлов) - 40})"
        ws0.append([s.год, s.месяц, label, s.файлов, s.сумма_строк or 0.0, files_cell])

    ws1 = wb.create_sheet("Документы")
    ws1.append(
        [
            "Файл",
            "Дата_периода",
            "КС-2_в_тексте",
            "Листов",
            "Строк_таблицы",
            "ВСЕГО_по_акту",
            "Сумма_общая_по_строкам",
            "Выручка_для_сводки",
            "Замечания",
        ],
    )
    for d in sorted(документы, key=lambda x: (x.имя.lower())):
        if d.ошибка:
            ws1.append([d.имя, "", "", "", "", "", "", "", d.ошибка])
            continue
        dp = ""
        if d.период_дата:
            p = d.период_дата
            dp = f"{p[2]:02d}.{p[1]:02d}.{p[0]:04d}"
        warn = "; ".join(d.предупреждения) if d.предупреждения else ""
        ws1.append(
            [
                d.имя,
                dp,
                "да" if d.есть_кс2 else "нет",
                d.листов,
                len(d.строки),
                d.сумма_всего_по_акту if d.сумма_всего_по_акту is not None else "",
                d.сумма_таблицы if d.сумма_таблицы is not None else "",
                d.выручка_по_документу if d.выручка_по_документу is not None else "",
                warn,
            ],
        )

    rows_any = any(d.строки for d in документы if not d.ошибка)
    if rows_any:
        ws2 = wb.create_sheet("Строки")
        ws2.append(
            [
                "Файл",
                "Лист",
                "Строка",
                "Год",
                "Месяц",
                "Период",
                "Наименование",
                "Сумма",
            ],
        )
        for d in документы:
            if d.ошибка:
                continue
            for ln in d.строки:
                ws2.append(
                    [
                        ln.имя_файла,
                        ln.лист,
                        ln.строка_excel,
                        ln.год,
                        ln.месяц,
                        _метка_месяца(ln.год, ln.месяц),
                        ln.наименование,
                        ln.сумма if ln.сумма is not None else "",
                    ],
                )

    wb.save(path)
