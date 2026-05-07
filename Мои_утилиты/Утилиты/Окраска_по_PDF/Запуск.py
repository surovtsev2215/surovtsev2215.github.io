# -*- coding: utf-8 -*-
"""
Калькулятор АКЗ по PDF: площадь антикоррозионной защиты (окраски) по ведомости и PDF.
Экспорт результата в Excel (.xlsx). Сессию при необходимости можно сохранить в JSON (меню «Ещё»).
Зоны (оси) задаются через чат ассистента и пункты меню «Ещё» на вкладке «Расчёт АКЗ»; база КМ/КМД — справочник и шаблоны имён файлов.
Вкладка «Ассистент» — выбор PDF проекта и чат с LLM (OpenAI-совместимый API): оси, режим, анализ, сохранение.
Рядом с PDF автоматически ищется файл проекта *_окраска_проект.json (или общий для нескольких PDF в папке).
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
import sys
import threading
import tkinter as tk
from collections import Counter, defaultdict
from datetime import datetime
from tkinter import filedialog, messagebox, scrolledtext, ttk
from typing import Any, Callable

_OPENPYXL_OK = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    _OPENPYXL_OK = True
except ImportError:
    Workbook = None  # type: ignore[misc, assignment]

_APP_VERSION = "3.5.0"
_APP_DISPLAY_NAME = "Калькулятор АКЗ по PDF"

# Ширины колонок результата (металл): id как в Treeview (ширина — у колонки, не у отдельной строки).
_METAL_COL_DEFAULT_WIDTHS_FLAT: dict[str, int] = {
    "mk": 118,
    "pos": 52,
    "prof": 300,
    "lmm": 86,
    "qty": 76,
    "shp": 52,
    "m2kit": 104,
    "m2pc": 88,
    "st": 40,
}
_METAL_COL_LABELS_FLAT: tuple[tuple[str, str], ...] = (
    ("mk", "Марка"),
    ("pos", "Поз."),
    ("prof", "Сечение"),
    ("lmm", "Длина, мм"),
    ("qty", "Шт."),
    ("shp", "Отпр."),
    ("m2kit", "м2 итого"),
    ("m2pc", "м2 за 1шт."),
    ("st", "Ошибка"),
)
_METAL_COL_DEFAULT_WIDTHS_TREE: dict[str, int] = {
    "#0": 210,
    "prof": 300,
    "lmm": 86,
    "qty": 64,
    "m2kit": 104,
    "m2pc": 88,
    "st": 40,
}
_METAL_COL_LABELS_TREE: tuple[tuple[str, str], ...] = (
    ("#0", "Марка"),
    ("prof", "Сечение"),
    ("lmm", "Длина, мм"),
    ("qty", "Шт."),
    ("m2kit", "м2 итого"),
    ("m2pc", "м2 за 1шт."),
    ("st", "Ошибка"),
)

_утил_dir = os.path.dirname(os.path.abspath(__file__))


def _metal_table_prefs_file() -> str:
    return os.path.join(_утил_dir, "metal_table_prefs.json")


def _merge_metal_col_widths(saved: dict[str, Any] | None) -> tuple[dict[str, int], dict[str, int]]:
    flat = dict(_METAL_COL_DEFAULT_WIDTHS_FLAT)
    tree_w = dict(_METAL_COL_DEFAULT_WIDTHS_TREE)
    if not isinstance(saved, dict):
        return flat, tree_w

    def _clip_w(x: Any) -> int:
        try:
            return max(28, min(900, int(float(x))))
        except (TypeError, ValueError):
            return 72

    sf = saved.get("flat_widths")
    if isinstance(sf, dict):
        for key in flat:
            if key in sf:
                flat[key] = _clip_w(sf[key])
    st = saved.get("tree_widths")
    if isinstance(st, dict):
        for key in tree_w:
            if key in st:
                tree_w[key] = _clip_w(st[key])
    return flat, tree_w


def _read_metal_col_widths_from_disk() -> tuple[dict[str, int], dict[str, int]]:
    try:
        path = _metal_table_prefs_file()
        if not os.path.isfile(path):
            return _merge_metal_col_widths(None)
        with open(path, encoding="utf-8") as f:
            raw = json.load(f)
        return _merge_metal_col_widths(raw if isinstance(raw, dict) else None)
    except (OSError, json.JSONDecodeError, TypeError):
        return _merge_metal_col_widths(None)


def _write_metal_col_widths_disk(flat_m: dict[str, int], tree_m: dict[str, int]) -> None:
    blob = {
        "flat_widths": {k: int(flat_m.get(k, _METAL_COL_DEFAULT_WIDTHS_FLAT[k])) for k in _METAL_COL_DEFAULT_WIDTHS_FLAT},
        "tree_widths": {k: int(tree_m.get(k, _METAL_COL_DEFAULT_WIDTHS_TREE[k])) for k in _METAL_COL_DEFAULT_WIDTHS_TREE},
    }
    with open(_metal_table_prefs_file(), "w", encoding="utf-8") as f:
        json.dump(blob, f, ensure_ascii=False, indent=2)

if _утил_dir not in sys.path:
    sys.path.insert(0, _утил_dir)
_ЯДРО_ХАБ = os.path.normpath(os.path.join(_утил_dir, "..", "..", "Ядро"))
if _ЯДРО_ХАБ not in sys.path:
    sys.path.insert(0, _ЯДРО_ХАБ)
import hub_theme as _hub_theme  # type: ignore
import оформление_утилиты as _shell  # type: ignore

try:
    import ассистент_llm as _assistant_llm  # type: ignore
except ImportError:
    _assistant_llm = None  # type: ignore[misc, assignment]

_UI_ACCENT = _hub_theme.ACCENT
_UI_ACCENT_ACTIVE = _hub_theme.ACCENT_HOVER
_UI_ACCENT_FG = _hub_theme.ACCENT_FG
_UI_TEXT = _hub_theme.TEXT
_UI_TEXT_DIM = _hub_theme.TEXT_DIM
_UI_BORDER = _hub_theme.BORDER_STRONG
_UI_PROGRESS_TROUGH = _hub_theme.PROGRESS_TROUGH
_UI_CARD_BG = _hub_theme.CARD_ALT
_UI_LIST_BG = _hub_theme.CARD
_UI_LIST_SELECT = _hub_theme.SIDEBAR_SELECT_BG
_UI_SUCCESS = _hub_theme.SUCCESS

# Допуск сверки масс/площадей со ведомостью (доля %).
# 0.1% для реальных КМ/КМД почти всегда даёт ложные срабатывания; 2% — разумный компромисс с reconcile.
_VALIDATION_TOLERANCE_STRICT_PCT = 2.0
# Межфайловое сравнение марок: допуск по относительному расхождению м² и минимум по абсолюту (м²).
_CROSS_FILE_MARK_TOLERANCE_STRICT_PCT = 35.0
_CROSS_FILE_MARK_MIN_ABS_DELTA_M2 = 0.75

try:
    import каталог_удалённый as _catalog_remote
except ImportError:
    _catalog_remote = None  # type: ignore[misc, assignment]

_DEPS_OK = False
try:
    import fitz  # PyMuPDF

    from авто_разбор import AnalyzeResult, PageResult, analyze_pdf_document
    from ведомость_металл import (
        cross_file_mark_checks,
        default_catalog_path,
        default_catalog_paths,
        metal_lines_missing_catalog_hints,
        profile_section_display_str,
        summarize_metal_validation,
        _clean_profile_cell,
        _unify_profile_chars,
        _normalize_profile_key,
    )

    _DEPS_OK = True
except ImportError:
    fitz = None
    AnalyzeResult = None  # type: ignore[misc, assignment]
    PageResult = None  # type: ignore[misc, assignment]
    analyze_pdf_document = None  # type: ignore[misc, assignment]
    default_catalog_path = lambda: ""  # type: ignore[misc, assignment]
    default_catalog_paths = lambda: []  # type: ignore[misc, assignment]
    cross_file_mark_checks = None  # type: ignore[misc, assignment]
    metal_lines_missing_catalog_hints = lambda _lines: []  # type: ignore[misc, assignment]
    summarize_metal_validation = lambda _v: {"count_total": 0, "by_severity": {}, "codes_top": []}  # type: ignore[misc, assignment]
    profile_section_display_str = lambda _r: None  # type: ignore[misc, assignment]

    def _clean_profile_cell_fallback(x: str) -> str:

        return str(x or "").strip()

    def _unify_profile_chars_fallback(x: str) -> str:

        return str(x or "").strip()

    _clean_profile_cell = _clean_profile_cell_fallback  # type: ignore[misc, assignment]

    _unify_profile_chars = _unify_profile_chars_fallback  # type: ignore[misc, assignment]

    def _normalize_profile_key(s: str) -> str:  # type: ignore[misc]
        s = (s or "").strip().lower().replace("х", "x")
        trans = str.maketrans(
            {
                "K": "к",
                "k": "к",
                "B": "б",
                "b": "б",
                "M": "м",
                "m": "м",
                "P": "п",
                "p": "п",
                "C": "с",
                "c": "с",
                "Y": "у",
                "y": "у",
                "T": "т",
                "t": "т",
            }
        )
        s = s.translate(trans)
        m = __import__("re").search(r"(\d{1,5})\s*(?:x|[-–—])\s*(\d{1,3})", s, __import__("re").I)
        if m:
            return f"{int(m.group(1))}x{int(m.group(2))}"
        return __import__("re").sub(r"[^a-zа-яё0-9]+", "", s, flags=__import__("re").I)


def _metal_line_profile_display(ml: dict[str, Any]) -> str:
    raw = str(ml.get("profile_raw") or "").strip().rstrip(",").strip()
    if raw:
        return raw
    return str(ml.get("profile_key") or "")


# Ячейка целиком — только лист B×т (в колонке «Профиль / состав» не показываем).
_RE_PROFILE_PLATE_SIZE_ONLY = re.compile(
    r"^\s*\d{1,5}\s*[x×х]\s*\d{1,3}\s*$",
    re.IGNORECASE,
)
_RE_PROFILE_PLATE_TRIPLE_IN_TEXT = re.compile(
    r"(?<![0-9])\d{1,5}\s*[x×х]\s*\d{1,5}\s*[x×х]\s*\d{1,5}(?!\d)",
    re.IGNORECASE,
)


def _strip_plate_sizes_from_profile_label(s: str) -> str:
    """Удаляет фрагменты вида 194×10; прокат (ДВУТАВР40К2 и т.д.) оставляет."""
    s = (s or "").strip()
    if not s:
        return ""
    if _RE_PROFILE_PLATE_TRIPLE_IN_TEXT.search(s):
        return s
    _wxh_token = re.compile(r"(?<![0-9])\d{1,5}\s*[x×х]\s*\d{1,3}(?![0-9])", re.IGNORECASE)
    parts = re.split(r"\s*;\s*", s)
    kept: list[str] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if _RE_PROFILE_PLATE_SIZE_ONLY.fullmatch(p):
            continue
        p2 = _wxh_token.sub("", p)
        p2 = re.sub(r"\s+", " ", p2).strip(" ,;").strip()
        if p2 and not _RE_PROFILE_PLATE_SIZE_ONLY.fullmatch(p2):
            kept.append(p2)
    return "; ".join(kept).strip()


def _bom_section_visual_from_ml(ml: dict[str, Any]) -> str:
    """Текст ячейки «Сечение» как из КМД: нормализация × и пробелов без подмены на «Лист …»."""
    raw = str(ml.get("profile_raw") or "").strip()
    if not raw:
        return ""
    cc = _clean_profile_cell(raw)  # type: ignore[misc]
    u = _unify_profile_chars(cc)  # type: ignore[misc]
    u = re.sub(r"\s+", " ", u).strip()
    return u.replace("x", "×").replace("х", "×").replace("X", "×")


def _metal_line_profile_display_ui(ml: dict[str, Any]) -> str:
    """Приоритет: как в спецификации («Сечение»), затем расчётная подпись «Лист …»."""
    ps = str(ml.get("profile_section_display") or "").strip()
    if ps:
        out = ps
    else:
        bom = _bom_section_visual_from_ml(ml)
        if bom:
            out = bom
        else:
            pd = str(ml.get("profile_dims_mm") or "").strip()
            if pd:
                out = pd
            else:
                base = _strip_plate_sizes_from_profile_label(_metal_line_profile_display(ml)).strip()
                out = base if base else str(ml.get("profile_key") or "").strip()

    lone = bool(out and len(out) <= 4 and bool(re.fullmatch(r"\d{1,4}", out.strip())))
    if lone:
        raw = str(ml.get("profile_raw") or "").strip()
        alt = profile_section_display_str(raw)  # type: ignore[misc]
        if isinstance(alt, str) and alt.strip():
            return alt.strip()
        pk = str(ml.get("profile_key") or "").strip()
        if pk and re.fullmatch(r"\d{1,5}x\d{1,5}$", pk.strip().lower().replace("х", "x")):
            return pk.lower().replace("х", "×").replace("x", "×")
        if len(raw) >= len(out) + 3:
            return raw[:240]
        return raw if raw else out
    return out


def _metal_line_profile_signature(ml: dict[str, Any]) -> str:
    pk = ml.get("profile_key")
    if pk is not None and str(pk).strip():
        return str(pk).strip().lower()
    return (_normalize_profile_key(str(ml.get("profile_raw") or "")) or "").lower()


def _metal_line_position_sort(ml: dict[str, Any]) -> int:
    p = str(ml.get("position") or "")
    return int(p) if p.isdigit() else 0


def _metal_line_area_m2_one_assembly(ml: dict[str, Any]) -> float:
    """м² позиции для одного комплекта марки (как в спецификации), без множителя отправки."""
    v = ml.get("area_m2_one_assembly")
    if v is not None:
        return float(v)
    sq = max(1, int(ml.get("shipment_qty") or 1))
    return float(ml.get("area_m2") or 0) / sq


def _metal_line_area_m2_per_piece(ml: dict[str, Any]) -> float:
    """м² окраски одной детали по колонке «шт» ведомости (строка = qty шт. на 1 комплект марки)."""
    v = ml.get("area_m2_per_piece")
    if v is not None:
        return float(v)
    q = ml.get("qty")
    try:
        qn = max(1, int(float(q))) if q is not None and str(q).strip() != "" else 1
    except (TypeError, ValueError):
        qn = 1
    return _metal_line_area_m2_one_assembly(ml) / float(qn)


def _metal_filter_matches_row(mark: str, ml: dict[str, Any], needle: str) -> bool:
    n = (needle or "").strip().lower()
    if not n:
        return True
    parts = (
        mark,
        str(ml.get("position") or ""),
        str(ml.get("profile_raw") or ""),
        str(ml.get("profile_key") or ""),
        str(ml.get("steel") or ""),
        str(ml.get("note") or ""),
        str(ml.get("profile_section_display") or ""),
        _metal_line_profile_display_ui(ml),
    )
    blob = " ".join(parts).lower()
    return n in blob


def _путь_установки() -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "УСТАНОВКА.txt")


def _apply_окраска_ui_style(master: tk.Misc) -> ttk.Style:
    """Единые ttk-стили (палитра хаба)."""
    st = ttk.Style(master)
    _hub_theme.apply_hub_ttk(st, card_bg=_UI_CARD_BG)
    return st


def _catalog_display_for_report(cat_path: str | None) -> str:
    if cat_path and str(cat_path).strip():
        p = os.path.normpath(str(cat_path).strip())
        bn = os.path.basename(p)
        if bn and bn not in (".", "..", ".csv"):
            return bn
        tail = p.replace("\\", "/")
        return tail[-56:] if len(tail) > 56 else tail
    try:
        ps = default_catalog_paths()
    except Exception:
        ps = []
    if ps:
        if len(ps) == 1:
            line = os.path.basename(ps[0])
        else:
            line = " + ".join(os.path.basename(p) for p in ps)
        if _catalog_remote is not None:
            m = _catalog_remote.read_meta()
            if m and isinstance(m, dict) and m.get("version"):
                line += f" · облако v{m.get('version')}"
        return line
    return "профили_база.csv + облако (кеш) + профили_м2_на_пм.csv"


def _short_text(s: str, n: int = 400) -> str:
    t = (s or "").strip().replace("\r", " ")
    if len(t) <= n:
        return t
    return t[: n - 1] + "…"


def _путь_проекта_для_pdf(pdf_path: str) -> str:
    base = os.path.splitext(pdf_path)[0]
    return base + "_окраска_проект.json"


def _путь_проекта_multi(first_pdf: str) -> str:
    d = os.path.dirname(os.path.abspath(first_pdf))
    return os.path.join(d, "_окраска_несколько_pdf_проект.json")


def _axis_groups_json_path() -> str:
    return os.path.join(_утил_dir, "осевые_группы.json")


def _km_kmd_base_path() -> str:
    return os.path.join(_утил_dir, "км_кмд_база.json")


def _recent_projects_json_path() -> str:
    return os.path.join(_утил_dir, "недавние_проекты.json")


def _recent_projects_load_paths() -> list[str]:
    path = _recent_projects_json_path()
    if not os.path.isfile(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        raw = data.get("paths") if isinstance(data, dict) else data
        if not isinstance(raw, list):
            return []
        out: list[str] = []
        for p in raw:
            if isinstance(p, str) and p.strip():
                out.append(os.path.normpath(p.strip()))
        return out
    except Exception:
        return []


def _recent_projects_save_paths(paths: list[str]) -> None:
    path = _recent_projects_json_path()
    uniq: list[str] = []
    seen: set[str] = set()
    for p in paths:
        n = os.path.normpath(p)
        if n not in seen:
            seen.add(n)
            uniq.append(n)
    tmp = path + ".tmp"
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump({"paths": uniq[:24]}, f, ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except Exception:
        pass


def _recent_projects_touch(project_json_path: str) -> None:
    p = os.path.normpath(project_json_path.strip())
    cur = _recent_projects_load_paths()
    cur = [p] + [x for x in cur if x != p]
    _recent_projects_save_paths(cur)


def _normalize_axis_groups_json(rows: Any) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    if not isinstance(rows, list):
        return out
    for row in rows:
        if not isinstance(row, dict):
            continue
        name = str(row.get("name") or "").strip()
        fr = str(row.get("file_re") or "").strip()
        if not name or not fr:
            continue
        try:
            re.compile(fr, re.I)
        except re.error:
            continue
        mr = str(row.get("mark_re") or "").strip()
        if mr:
            try:
                re.compile(mr, re.I)
            except re.error:
                mr = ""
        out.append({"name": name, "file_re": fr, "mark_re": mr})
    return out


_LONE_NUMERIC_PROFILE_RAW = re.compile(r"^\s*\d+(?:[.,]\d+)?\s*$")


def _metal_export_row_signals(metal_lines: list[dict[str, Any]] | None) -> dict[str, Any]:
    """Признак ошибки колонок разбора: в «Сечение» попало одно число без размерности профиля."""
    n = 0
    for ml in metal_lines or []:
        raw = str((ml or {}).get("profile_raw") or "")
        if raw.strip() and _LONE_NUMERIC_PROFILE_RAW.fullmatch(raw):
            n += 1
    return {"profile_raw_lone_numeric_token_rows": n}


class ОкраскаПоPdfПанель:
    def __init__(self, родитель: tk.Misc) -> None:
        self._root = родитель
        self._rep = getattr(родитель, "report_tab_error", lambda *_: None)

        self._pdf_paths: list[str] = []
        self._results: dict[str, AnalyzeResult] = {}  # type: ignore[misc, valid-type]
        self._busy = False
        self._axis_groups: list[dict[str, str]] = []
        self._axes_grp_tree: ttk.Treeview | None = None
        self._axes_sum_tree: ttk.Treeview | None = None
        self._lbl_axes_hint: tk.Label | None = None
        self._km_kmd_cache: dict[str, Any] | None = None
        self._load_axis_groups_disk()
        self._assistant_after_analyze_cb: Callable[[], None] | None = None
        self._current_project_path: str | None = None
        self._assistant_busy = False
        self._assistant_chat: scrolledtext.ScrolledText | None = None
        self._assistant_status_lbl: tk.Label | None = None
        self._assistant_msg_entry: tk.Entry | None = None
        def _metal_catalog_available() -> bool:
            try:
                return len(default_catalog_paths()) > 0
            except Exception:
                p0 = default_catalog_path()
                return bool(p0 and os.path.isfile(p0))

        self._goal = tk.StringVar(value=("metal" if _metal_catalog_available() else "sheet"))
        self._catalog_path: str | None = None
        self._cross_file_validation: list[dict[str, Any]] = []
        self._report_title_override = tk.StringVar(value="")
        ac0: dict[str, Any] = (
            _assistant_llm.прочитать_конфиг(_утил_dir) if _assistant_llm is not None else {}
        )
        # Принудительный LLM (меню «Ещё»). Фоновый авто‑режим в авто_разбор.py не зависит от этого.
        self._merged_llm_force = tk.BooleanVar(value=False)
        self._allow_vision_ui = tk.BooleanVar(value=bool(ac0.get("allow_vision_default")))
        self._metal_flat_list = tk.BooleanVar(value=False)
        self._metal_filter_var = tk.StringVar(value="")
        self._metal_tool_widgets: list[Any] = []
        self._metal_table_after_id: str | None = None

        self._banner_canvas: tk.Canvas | None = None
        self._shimmer_after: str | None = None
        self._shimmer_phase = 0.0

        self._build_ui()
        self._sync_metal_result_tools(False)

    def _build_ui(self) -> None:
        bg = getattr(self._root, "cget", lambda **_k: "#f8fafc")("bg") if hasattr(self._root, "cget") else "#f8fafc"
        self._panel_bg = bg
        try:
            self._root.configure(bg=bg)
        except Exception:
            pass
        _apply_окраска_ui_style(self._root)

        bh = tk.Frame(self._root, bg=bg)
        bh.pack(fill=tk.X)
        self._banner_canvas = tk.Canvas(
            bh,
            height=72,
            highlightthickness=0,
            borderwidth=0,
            bd=0,
        )
        self._banner_canvas.pack(fill=tk.X)
        self._banner_canvas.bind("<Configure>", lambda _e: self._paint_banner())

        pad_out = {"padx": 8, "pady": 6}
        self._notebook = ttk.Notebook(self._root)
        self._notebook.pack(fill=tk.BOTH, expand=True)
        tab_assist = tk.Frame(self._notebook, bg=bg)
        tab_calc = tk.Frame(self._notebook, bg=bg)
        self._notebook.add(tab_assist, text="Ассистент")
        self._notebook.add(tab_calc, text="Расчёт АКЗ")
        self._build_assistant_tab(tab_assist, bg)
        pad_btn = 4

        top = tk.Frame(tab_calc, bg=bg)
        top.pack(fill=tk.X, **pad_out)

        row1 = tk.Frame(top, bg=bg)
        row1.pack(fill=tk.X, pady=(0, 6))
        files_lf = ttk.LabelFrame(row1, text="Файлы", padding=(8, 6))
        files_lf.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))
        files_inner = tk.Frame(files_lf, bg=_UI_CARD_BG)
        files_inner.pack(fill=tk.BOTH, expand=True)
        ttk.Button(
            files_inner,
            text="Добавить PDF…",
            command=self._add_pdfs,
            style="Secondary.TButton",
        ).pack(side=tk.LEFT, padx=(0, pad_btn))
        ttk.Button(
            files_inner,
            text="Очистить список",
            command=self._clear_paths,
            style="Secondary.TButton",
        ).pack(side=tk.LEFT, padx=(0, pad_btn))
        self._btn_analyze = ttk.Button(
            files_inner,
            text="Анализ",
            command=self._start_analyze,
            style="Accent.TButton",
        )
        self._btn_analyze.pack(side=tk.LEFT, padx=(8, 0))
        more = tk.Menubutton(
            files_inner,
            text="Ещё ▾",
            font=_hub_theme.FONT_SM,
            bg=_UI_CARD_BG,
            fg=_UI_TEXT,
            relief=tk.FLAT,
            bd=0,
            highlightthickness=1,
            highlightbackground=_UI_BORDER,
            padx=10,
            pady=4,
            cursor="hand2",
        )
        more_menu = tk.Menu(more, tearoff=0, font=_hub_theme.FONT_SM)
        more_menu.add_command(label="Открыть сессию (JSON)…", command=self._load_project_dialog)
        more_menu.add_command(label="Сохранить сессию (JSON)…", command=self._save_project)
        more_menu.add_separator()
        more_menu.add_command(label="Зоны: пресет оси 1–3…", command=self._axis_preset_123)
        more_menu.add_command(label="Зоны: шаблоны КМ / КМД…", command=self._apply_km_kmd_presets_dialog)
        more_menu.add_command(label="Зоны: сохранить общий шаблон…", command=self._save_axis_groups_disk)
        more_menu.add_command(label="Зоны: экспорт свода (CSV)…", command=self._export_axes_csv)
        more_menu.add_command(label="Зоны: добавить зону…", command=self._axis_group_add_dialog)
        more_menu.add_separator()
        more_menu.add_command(label="Каталог из интернета: настройка…", command=self._remote_catalog_settings)
        more_menu.add_command(label="Обновить каталог из интернета", command=self._remote_catalog_fetch_now)
        more_menu.add_command(label="Ключи без м²/м в каталоге (CSV)…", command=self._export_missing_keys_csv)
        more_menu.add_separator()
        sub_kmd = tk.Menu(more_menu, tearoff=0, font=_hub_theme.FONT_SM)
        sub_kmd.add_checkbutton(
            label="Принудительно LLM при ≥4 стр.",
            variable=self._merged_llm_force,
            onvalue=True,
            offvalue=False,
        )
        sub_kmd.add_checkbutton(
            label="Vision при сыром тексте",
            variable=self._allow_vision_ui,
            onvalue=True,
            offvalue=False,
        )
        more_menu.add_cascade(label="Объединённые PDF (КМД)", menu=sub_kmd)
        more_menu.add_separator()
        more_menu.add_command(label="Отладка merged: сохранить JSON…", command=self._export_merged_debug_json)
        more.config(menu=more_menu)
        more.pack(side=tk.LEFT, padx=(12, 0))

        self._lbl_count = tk.Label(row1, text="Файлов: 0", bg=bg, font=("Segoe UI", 10, "bold"), fg=_UI_TEXT)
        self._lbl_count.pack(side=tk.RIGHT, padx=8)

        mid_outer = tk.Frame(tab_calc, bg=bg)
        mid_outer.pack(fill=tk.BOTH, expand=True, padx=8, pady=(6, 4))
        pw_h = ttk.Panedwindow(mid_outer, orient=tk.HORIZONTAL)
        pw_h.pack(fill=tk.BOTH, expand=True)

        left = tk.Frame(pw_h, bg=bg, width=240)
        pw_h.add(left, weight=0)
        right = tk.Frame(pw_h, bg=bg)
        pw_h.add(right, weight=1)

        left_pad = tk.Frame(left, bg=_UI_CARD_BG, highlightthickness=1, highlightbackground=_UI_BORDER)
        left_pad.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        tk.Label(left_pad, text="Загруженные PDF", font=("Segoe UI", 9, "bold"), bg=_UI_CARD_BG, fg=_UI_TEXT).pack(anchor=tk.W, pady=(0, 4))
        self._list = tk.Listbox(
            left_pad,
            height=10,
            font=("Segoe UI", 9),
            selectmode=tk.EXTENDED,
            activestyle="none",
            bg=_UI_LIST_BG,
            fg=_UI_TEXT,
            selectbackground=_UI_LIST_SELECT,
            selectforeground=_hub_theme.SIDEBAR_SELECT_FG,
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=_UI_BORDER,
        )
        self._list.pack(fill=tk.BOTH, expand=True)

        pw_v = ttk.Panedwindow(right, orient=tk.VERTICAL)
        pw_v.pack(fill=tk.BOTH, expand=True)
        upper = tk.Frame(pw_v, bg=bg)
        pw_v.add(upper, weight=1)

        prog_fr = tk.Frame(upper, bg=bg)
        prog_fr.pack(fill=tk.X, pady=(0, 8))

        self._lbl_stage = tk.Label(prog_fr, text="Этап: —", bg=bg, fg=_UI_TEXT_DIM, font=("Segoe UI", 9))
        self._lbl_stage.pack(anchor=tk.W)
        row_pb = tk.Frame(prog_fr, bg=bg)
        row_pb.pack(fill=tk.X, pady=4)
        try:
            self._pb = ttk.Progressbar(row_pb, orient=tk.HORIZONTAL, length=400, mode="determinate", maximum=100, style="Hub.Horizontal.TProgressbar")
        except tk.TclError:
            self._pb = ttk.Progressbar(row_pb, orient=tk.HORIZONTAL, length=400, mode="determinate", maximum=100)
        self._pb.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._lbl_pct = tk.Label(row_pb, text="0%", width=6, bg=bg, font=("Segoe UI", 11, "bold"), fg=_UI_ACCENT)
        self._lbl_pct.pack(side=tk.LEFT, padx=(8, 0))

        self._lbl_strategy = tk.Label(
            upper,
            text="Способ расчёта: —",
            bg=bg,
            fg=_UI_TEXT,
            font=("Segoe UI", 9),
            wraplength=560,
            justify=tk.LEFT,
        )
        self._lbl_strategy.pack(anchor=tk.W, pady=(0, 4))
        self._lbl_reason = tk.Label(
            upper,
            text="",
            bg=bg,
            fg=_UI_TEXT_DIM,
            font=("Segoe UI", 9),
            wraplength=560,
            justify=tk.LEFT,
        )
        self._lbl_reason.pack(anchor=tk.W, pady=(0, 6))

        sum_row = tk.Frame(upper, bg=bg)
        sum_row.pack(fill=tk.X, pady=(0, 4))
        self._lbl_suspicion = tk.Label(
            sum_row,
            text="",
            bg=bg,
            fg=_UI_TEXT_DIM,
            font=("Segoe UI", 9),
            wraplength=720,
            justify=tk.LEFT,
        )
        self._lbl_suspicion.pack(anchor=tk.W, pady=(2, 0))

        tree_hdr = tk.Frame(upper, bg=bg)
        tree_hdr.pack(fill=tk.X, pady=(8, 2))
        tk.Label(tree_hdr, text="Результаты расчёта", font=("Segoe UI", 10, "bold"), bg=bg, fg=_UI_TEXT).pack(side=tk.LEFT)
        tree_tools = tk.Frame(tree_hdr, bg=bg)
        tree_tools.pack(side=tk.RIGHT)
        ttk.Button(tree_tools, text="Развернуть всё", command=self._tree_expand_all).pack(side=tk.LEFT, padx=(0, pad_btn))
        ttk.Button(tree_tools, text="Свернуть марки", command=self._tree_collapse_marks).pack(side=tk.LEFT, padx=(0, pad_btn))
        ttk.Button(tree_tools, text="Копировать сводку", command=self._copy_summary).pack(side=tk.LEFT, padx=(4, 0))
        self._btn_metal_columns = ttk.Button(
            tree_tools,
            text="Ширина колонок…",
            command=self._metal_open_column_widths_dialog,
            state=tk.DISABLED,
        )
        self._btn_metal_columns.pack(side=tk.LEFT, padx=(6, 0))

        tree_tune = tk.Frame(upper, bg=bg)
        tree_tune.pack(fill=tk.X, pady=(0, 6))
        cb_flat = ttk.Checkbutton(
            tree_tune,
            text="Таблица одним списком (без групп марок)",
            variable=self._metal_flat_list,
            command=self._schedule_metal_table_refresh,
        )
        cb_flat.pack(side=tk.LEFT, padx=(0, 14))
        tk.Label(tree_tune, text="Фильтр:", bg=bg, fg=_UI_TEXT_DIM, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 6))
        ent_flt = ttk.Entry(tree_tune, textvariable=self._metal_filter_var, width=36)
        ent_flt.pack(side=tk.LEFT, padx=(0, 8))
        ent_flt.bind("<KeyRelease>", self._schedule_metal_table_refresh)
        tk.Label(
            tree_tune,
            text="двойной щелчок по строке — копировать",
            bg=bg,
            fg=_UI_TEXT_DIM,
            font=("Segoe UI", 8),
        ).pack(side=tk.LEFT, padx=(4, 0))
        self._metal_tool_widgets = [cb_flat, ent_flt, self._btn_metal_columns]

        tree_row = tk.Frame(upper, bg=bg, highlightthickness=1, highlightbackground=_UI_BORDER)
        tree_row.pack(fill=tk.BOTH, expand=True)
        tree_row.grid_rowconfigure(0, weight=1)
        tree_row.grid_columnconfigure(0, weight=1)
        self._tree = ttk.Treeview(tree_row, columns=("file", "page", "area", "detail"), show="headings", height=17)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb = ttk.Scrollbar(tree_row, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        vsb.grid(row=0, column=1, sticky="ns")
        self._tree_hscroll = ttk.Scrollbar(tree_row, orient="horizontal", command=self._tree.xview)
        self._tree.configure(xscrollcommand=self._tree_hscroll.set)
        self._tree_hscroll.grid(row=1, column=0, sticky="ew", columnspan=2)
        try:
            _hub_theme.configure_results_tree_tags(self._tree)
        except tk.TclError:
            pass
        self._tree.bind("<Button-3>", self._tree_context_menu)
        self._tree.bind("<Double-Button-1>", self._tree_double_click_copy_row)

        ttk.Separator(upper, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(10, 0))
        total_card = tk.Frame(upper, bg=_UI_CARD_BG, highlightthickness=1, highlightbackground=_UI_BORDER)
        total_card.pack(fill=tk.X, pady=(8, 0))
        inner_total = tk.Frame(total_card, bg=_UI_CARD_BG)
        inner_total.pack(fill=tk.X, padx=12, pady=10)
        row_tot = tk.Frame(inner_total, bg=_UI_CARD_BG)
        row_tot.pack(fill=tk.X)
        tk.Label(row_tot, text="Итого по таблице", bg=_UI_CARD_BG, fg=_UI_TEXT_DIM, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 12))
        val_fr = tk.Frame(row_tot, bg=_UI_CARD_BG)
        val_fr.pack(side=tk.LEFT)
        self._lbl_total_value = tk.Label(val_fr, text="—", font=("Segoe UI", 14, "bold"), bg=_UI_CARD_BG, fg=_UI_TEXT)
        self._lbl_total_value.pack(side=tk.LEFT, anchor=tk.S)
        self._lbl_total_unit = tk.Label(val_fr, text="", bg=_UI_CARD_BG, fg=_UI_TEXT_DIM, font=("Segoe UI", 9))
        self._lbl_total_unit.pack(side=tk.LEFT, anchor=tk.S, padx=(6, 0), pady=(0, 2))
        self._lbl_total_sub = tk.Label(
            inner_total,
            text="",
            anchor=tk.W,
            bg=_UI_CARD_BG,
            fg=_UI_TEXT_DIM,
            font=("Segoe UI", 9),
            justify=tk.LEFT,
            wraplength=720,
        )
        self._lbl_total_sub.pack(fill=tk.X, pady=(8, 0))

        save_row = tk.Frame(inner_total, bg=_UI_CARD_BG)
        save_row.pack(fill=tk.X, pady=(14, 0))
        tk.Label(
            save_row,
            text="Название для Excel-отчёта",
            bg=_UI_CARD_BG,
            fg=_UI_TEXT_DIM,
            font=("Segoe UI", 9),
        ).pack(side=tk.LEFT, padx=(0, 10))
        self._entry_report_title = ttk.Entry(save_row, textvariable=self._report_title_override, width=36)
        self._entry_report_title.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 12))
        self._btn_save_result = ttk.Button(
            save_row,
            text="Сохранить мой результат…",
            command=self._export_xlsx,
            style="Success.TButton",
        )
        self._btn_save_result.pack(side=tk.RIGHT)

        self._hint = tk.Label(
            tab_calc,
            text="Площадь АКЗ: режим «металл» — ведомость + каталог м²/п.м (CSV); режим «лист» — по штампу или геометрии PDF. Итог сохраняется в Excel.",
            anchor=tk.W,
            bg=bg,
            fg=_UI_TEXT_DIM,
            font=("Segoe UI", 9),
            wraplength=720,
            justify=tk.LEFT,
        )
        self._hint.pack(fill=tk.X, padx=8, pady=(0, 6))

    def _build_assistant_tab(self, host: tk.Frame, bg: str) -> None:
        intro = (
            "«PDF проекта…» задаёт состав файлов (тот же список, что на «Расчёт АКЗ»). "
            "Рядом с PDF может лежать JSON сессии (*_окраска_проект.json) — он появляется после «Анализа» или «Сохранить сессию»; "
            "если его ещё нет, это не ошибка: ассистент всё равно получает имена PDF и зоны. "
            "Если JSON есть, подтянутся цель расчёта, каталог и сохранённые результаты для выбранных PDF. "
            "Опишите задачу по-русски (например: «режим металла, оси 1–5, пересчитай»). "
            "В запрос к API уходит только список имён PDF и зоны — содержимое страниц PDF не отправляется. "
            "Нужен ключ в assistant_config.json (см. УСТАНОВКА.txt)."
        )
        tk.Label(host, text=intro, bg=bg, fg=_UI_TEXT_DIM, font=_hub_theme.FONT_SM, wraplength=920, justify=tk.LEFT).pack(
            anchor=tk.W, padx=10, pady=(8, 6)
        )
        top = tk.Frame(host, bg=bg)
        top.pack(fill=tk.X, padx=10, pady=4)
        ttk.Button(
            top,
            text="PDF проекта…",
            command=self._assistant_pick_pdfs,
            style="Accent.TButton",
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(
            top,
            text="Настройки API…",
            command=self._assistant_settings_dialog,
            style="Secondary.TButton",
        ).pack(side=tk.LEFT)
        self._assistant_status_lbl = tk.Label(
            host,
            text="PDF не выбраны.",
            bg=bg,
            fg=_UI_TEXT_DIM,
            font=_hub_theme.FONT_SM,
            anchor=tk.W,
            wraplength=920,
            justify=tk.LEFT,
        )
        self._assistant_status_lbl.pack(fill=tk.X, padx=10, pady=(2, 4))

        mid = tk.Frame(host, bg=bg)
        mid.pack(fill=tk.BOTH, expand=True, padx=10, pady=(6, 4))
        self._assistant_chat = scrolledtext.ScrolledText(
            mid,
            height=16,
            wrap=tk.WORD,
            font=_hub_theme.FONT_BASE,
            state=tk.DISABLED,
            bg=_UI_LIST_BG,
            fg=_UI_TEXT,
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=_UI_BORDER,
        )
        self._assistant_chat.pack(fill=tk.BOTH, expand=True)

        tk.Label(
            host,
            text="Сообщение — только в поле ниже; область выше — журнал ответов (туда текст не отправляется).",
            bg=bg,
            fg=_UI_TEXT_DIM,
            font=_hub_theme.FONT_SM,
            wraplength=920,
            justify=tk.LEFT,
            anchor=tk.W,
        ).pack(fill=tk.X, padx=10, pady=(2, 0))

        bot = tk.Frame(host, bg=bg)
        bot.pack(fill=tk.X, padx=10, pady=(4, 10))
        self._assistant_msg_var = tk.StringVar(value="")
        self._assistant_msg_entry = tk.Entry(
            bot,
            textvariable=self._assistant_msg_var,
            font=_hub_theme.FONT_BASE,
            bg=_UI_LIST_BG,
            fg=_UI_TEXT,
            insertbackground=_UI_TEXT,
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=_UI_BORDER,
        )
        self._assistant_msg_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        self._assistant_msg_entry.bind("<Return>", lambda _e: self._assistant_send())
        self._assistant_msg_entry.bind("<Control-v>", self._assistant_paste_from_clipboard)
        self._assistant_msg_entry.bind("<Control-V>", self._assistant_paste_from_clipboard)
        self._assistant_msg_entry.bind("<Shift-Insert>", self._assistant_paste_from_clipboard)

        self._assistant_btn_send = ttk.Button(
            bot,
            text="Отправить",
            command=self._assistant_send,
            style="Accent.TButton",
        )
        self._assistant_btn_send.pack(side=tk.LEFT)

    def _assistant_pick_pdfs(self) -> None:
        paths = filedialog.askopenfilenames(filetypes=[("PDF", "*.pdf"), ("Все файлы", "*.*")])
        if not paths:
            return
        norm = [os.path.normpath(p) for p in paths if os.path.isfile(p)]
        if not norm:
            messagebox.showwarning("PDF", "Не удалось прочитать выбранные файлы.")
            return
        self._hydrate_session_from_pdfs_and_optional_json(norm)

    def _sidecar_project_path_for_current_pdfs(self) -> str:
        if len(self._pdf_paths) == 1:
            return _путь_проекта_для_pdf(self._pdf_paths[0])
        return _путь_проекта_multi(self._pdf_paths[0])

    def _assistant_update_project_status_label(self) -> None:
        lbl = self._assistant_status_lbl
        if lbl is None:
            return
        if not self._pdf_paths:
            lbl.config(text="PDF не выбраны — нажмите «PDF проекта…».")
            return
        jp = self._sidecar_project_path_for_current_pdfs()
        n = len(self._pdf_paths)
        if os.path.isfile(jp):
            lbl.config(text=f"PDF: {n} шт. · JSON проекта: найден ({os.path.basename(jp)})")
        else:
            lbl.config(
                text=(
                    f"PDF: {n} шт. — список принят. Файла «{os.path.basename(jp)}» рядом с PDF ещё нет "
                    "(это не означает, что PDF «не читается»): он появляется после «Анализ» на вкладке «Расчёт АКЗ» "
                    "или «Сохранить сессию» в меню «Ещё ▾». Пока зоны берутся из «{os.path.basename(_axis_groups_json_path())}». "
                    "Чат с ассистентом можно открыть: в запрос уходят имена PDF и зоны, не текст со страниц."
                )
            )

    def _reset_session_after_pdf_pick_no_sidecar_restore(self) -> None:
        self._results.clear()
        self._cross_file_validation.clear()
        self._load_axis_groups_disk()
        self._refresh_axes_grp_list()
        self._refresh_axes_summary()
        self._lbl_strategy.pack(anchor=tk.W, pady=(0, 4))
        self._lbl_reason.pack(anchor=tk.W, pady=(0, 6))
        self._lbl_strategy.config(text="Способ расчёта: —")
        self._lbl_reason.config(text="")
        self._lbl_total_sub.config(text="")
        self._refresh_table()
        self._set_progress(0, "—")

    def _hydrate_session_from_pdfs_and_optional_json(self, picked_norm: list[str]) -> None:
        self._pdf_paths = list(picked_norm)
        self._sync_list()
        self._current_project_path = self._sidecar_project_path_for_current_pdfs()
        jp = self._current_project_path
        try:
            if jp and os.path.isfile(jp):
                with open(jp, "r", encoding="utf-8") as f:
                    data = json.load(f)
                ver = int(data.get("version", 1))
                if ver < 2:
                    messagebox.showwarning(
                        "Проект",
                        "Это проект старого формата (ручные полигоны). Он не поддерживается в авто-режиме. "
                        "Список PDF обновлён; зоны — из шаблона на диске.",
                    )
                    self._reset_session_after_pdf_pick_no_sidecar_restore()
                elif (mode := self._apply_loaded_project_payload(
                    data,
                    pdf_paths_override=list(self._pdf_paths),
                    json_source_path=jp,
                    show_loaded_dialog=False,
                )) == "none":
                    self._reset_session_after_pdf_pick_no_sidecar_restore()
                elif mode == "full":
                    self._assistant_append_chat("sys", f"Подтянуты данные из JSON: {jp}")
                elif mode == "paths_only":
                    self._assistant_append_chat(
                        "sys",
                        f"Параметры из JSON применены ({jp}). Разбор PDF недоступен — проверьте зависимости (см. УСТАНОВКА.txt).",
                    )
            else:
                self._reset_session_after_pdf_pick_no_sidecar_restore()
                self._assistant_append_chat(
                    "sys",
                    f"Список PDF обновлён: выбрано {len(self._pdf_paths)} файл(ов). "
                    f"Файла проекта «{os.path.basename(jp) if jp else '—'}» рядом с PDF пока нет — это штатно для нового набора файлов. "
                    f"Зоны по умолчанию из «{os.path.basename(_axis_groups_json_path())}». "
                    "Чтобы разобрать чертежи и создать/обновить JSON: вкладка «Расчёт АКЗ» → «Анализ». "
                    "Ассистенту для ответа нужны имена файлов и зоны (текст со страниц PDF в чат не отправляется).",
                )
            self._rep(False)
        except Exception as e:
            self._rep(True)
            messagebox.showerror("Проект", str(e))
        self._assistant_update_project_status_label()

    def _assistant_append_chat(self, role: str, text: str) -> None:
        w = self._assistant_chat
        if w is None:
            return
        prefix = {"user": "Вы: ", "assistant": "Ассистент: ", "sys": "⚙ ", "err": "Ошибка: "}.get(role, "")
        try:
            w.configure(state=tk.NORMAL)
            w.insert(tk.END, prefix + (text or "").strip() + "\n\n")
            w.configure(state=tk.DISABLED)
            w.see(tk.END)
        except tk.TclError:
            pass

    def _assistant_settings_dialog(self) -> None:
        if _assistant_llm is None:
            messagebox.showinfo("Ассистент", "Модуль ассистент_llm недоступен.")
            return
        cfg = _assistant_llm.прочитать_конфиг(_утил_dir)
        top = tk.Toplevel(self._root)
        top.title("Настройки LLM")
        top.transient(self._root)
        top.grab_set()
        pad = {"padx": 10, "pady": 6}
        tk.Label(
            top,
            text="Ключ не коммитьте в git. Файл assistant_config.json рядом с утилитой.",
            fg=_UI_TEXT_DIM,
            font=("Segoe UI", 9),
        ).pack(anchor=tk.W, **pad)
        api_var = tk.StringVar(value=str(cfg.get("api_key") or ""))
        url_var = tk.StringVar(value=str(cfg.get("base_url") or "https://api.openai.com/v1"))
        model_var = tk.StringVar(value=str(cfg.get("model") or "gpt-4o-mini"))
        fr = tk.Frame(top)
        fr.pack(fill=tk.X, **pad)
        tk.Label(fr, text="API key:").grid(row=0, column=0, sticky=tk.W)
        api_e = tk.Entry(fr, textvariable=api_var, width=56, show="•")
        api_e.grid(row=0, column=1, sticky=tk.EW)
        self._assistant_bind_entry_paste(api_e)
        tk.Label(fr, text="Base URL:").grid(row=1, column=0, sticky=tk.W, pady=(6, 0))
        url_e = tk.Entry(fr, textvariable=url_var, width=56)
        url_e.grid(row=1, column=1, sticky=tk.EW, pady=(6, 0))
        self._assistant_bind_entry_paste(url_e)
        tk.Label(fr, text="Model:").grid(row=2, column=0, sticky=tk.W, pady=(6, 0))
        model_e = tk.Entry(fr, textvariable=model_var, width=56)
        model_e.grid(row=2, column=1, sticky=tk.EW, pady=(6, 0))
        self._assistant_bind_entry_paste(model_e)
        fr.columnconfigure(1, weight=1)

        def save() -> None:
            _assistant_llm.сохранить_конфиг(
                _утил_dir,
                {
                    "api_key": api_var.get().strip(),
                    "base_url": url_var.get().strip(),
                    "model": model_var.get().strip(),
                },
            )
            top.destroy()
            messagebox.showinfo("Ассистент", "Сохранено в assistant_config.json")

        bf = tk.Frame(top)
        bf.pack(fill=tk.X, **pad)
        ttk.Button(bf, text="Сохранить", command=save, style="Accent.TButton").pack(side=tk.RIGHT, padx=4)
        ttk.Button(bf, text="Отмена", command=top.destroy, style="Secondary.TButton").pack(side=tk.RIGHT)

    def _assistant_bind_entry_paste(self, w: tk.Entry) -> None:
        w.bind("<Control-v>", self._assistant_paste_from_clipboard)
        w.bind("<Control-V>", self._assistant_paste_from_clipboard)
        w.bind("<Shift-Insert>", self._assistant_paste_from_clipboard)

    def _assistant_paste_from_clipboard(self, event: tk.Event) -> str | None:
        w = event.widget
        try:
            clip = w.clipboard_get()
        except tk.TclError:
            return None
        try:
            if w.selection_present():
                w.delete(tk.SEL_FIRST, tk.SEL_LAST)
        except tk.TclError:
            pass
        w.insert(tk.INSERT, clip)
        return "break"

    def _assistant_send(self, _evt: object | None = None) -> None:
        if _assistant_llm is None:
            messagebox.showinfo("Ассистент", "Модуль ассистент_llm недоступен.")
            return
        if self._assistant_busy:
            return
        msg = (self._assistant_msg_var.get() or "").strip()
        if not msg:
            return
        self._assistant_msg_var.set("")
        self._assistant_append_chat("user", msg)
        self._assistant_busy = True
        try:
            self._assistant_btn_send.config(state=tk.DISABLED)
        except tk.TclError:
            pass

        def work() -> None:
            ctx = _assistant_llm.построить_контекст(
                pdf_basenames=[os.path.basename(p) for p in self._pdf_paths],
                axis_groups=list(self._axis_groups),
                goal=str(self._goal.get()),
                project_path=self._current_project_path,
                app_version=_APP_VERSION,
            )
            plan, err = _assistant_llm.запросить_план(_утил_dir, msg, ctx)
            self._root.after(0, lambda: self._assistant_on_llm_done(plan, err))

        threading.Thread(target=work, daemon=True).start()

    def _assistant_on_llm_done(self, plan: dict[str, Any] | None, err: str) -> None:
        self._assistant_busy = False
        try:
            self._assistant_btn_send.config(state=tk.NORMAL)
        except tk.TclError:
            pass
        if err:
            self._assistant_append_chat("err", err)
            return
        assert plan is not None
        am = str(plan.get("assistant_message") or "").strip()
        if am:
            self._assistant_append_chat("assistant", am)
        acts = plan.get("actions")
        if not isinstance(acts, list):
            acts = []
        self._assistant_execute_actions(acts, 0)

    @staticmethod
    def _axis_groups_from_numbers(nums: list[int]) -> list[dict[str, str]]:
        out: list[dict[str, str]] = []
        for n in nums:
            ns = str(int(n))
            fr = rf".*[_\-]{ns}([._\-]|\.pdf$|$)"
            out.append({"name": f"Ось {ns}", "file_re": fr, "mark_re": ""})
        return out

    def _assistant_execute_actions(self, acts: list[dict[str, Any]], idx: int) -> None:
        while idx < len(acts):
            act = acts[idx]
            if not isinstance(act, dict):
                idx += 1
                continue
            t = act.get("type")
            if t == "NONE":
                idx += 1
                continue
            if t == "SET_GOAL":
                g = act.get("goal")
                if g in ("metal", "sheet"):
                    self._goal.set(g)
                idx += 1
                continue
            if t == "SET_AXIS_RANGE":
                try:
                    a = int(act["from"])
                    b = int(act["to"])
                except Exception:
                    self._assistant_append_chat("err", "SET_AXIS_RANGE: неверные числа.")
                    return
                nums = list(range(a, b + 1))
                groups = self._axis_groups_from_numbers(nums)
                if bool(act.get("replace", True)):
                    self._axis_groups = groups
                else:
                    self._axis_groups.extend(groups)
                self._refresh_axes_grp_list()
                self._refresh_axes_summary()
                idx += 1
                continue
            if t == "SET_AXIS_LIST":
                axes_raw = act.get("axes")
                if not isinstance(axes_raw, list):
                    self._assistant_append_chat("err", "SET_AXIS_LIST: нужен axes[].")
                    return
                try:
                    nums = sorted({int(x) for x in axes_raw})
                except Exception:
                    self._assistant_append_chat("err", "SET_AXIS_LIST: оси должны быть целыми.")
                    return
                groups = self._axis_groups_from_numbers(nums)
                if bool(act.get("replace", True)):
                    self._axis_groups = groups
                else:
                    self._axis_groups.extend(groups)
                self._refresh_axes_grp_list()
                self._refresh_axes_summary()
                idx += 1
                continue
            if t == "SET_AXIS_CUSTOM":
                groups_raw = act.get("groups")
                if not isinstance(groups_raw, list):
                    self._assistant_append_chat("err", "SET_AXIS_CUSTOM: нужен groups[].")
                    return
                parsed: list[dict[str, str]] = []
                for g in groups_raw:
                    if not isinstance(g, dict):
                        continue
                    parsed.append(
                        {
                            "name": str(g.get("name") or "").strip(),
                            "file_re": str(g.get("file_re") or "").strip(),
                            "mark_re": str(g.get("mark_re") or "").strip(),
                        }
                    )
                validated = _normalize_axis_groups_json(parsed)
                if not validated and parsed:
                    self._assistant_append_chat("err", "SET_AXIS_CUSTOM: ни одна группа не прошла проверку regex.")
                    return
                if bool(act.get("replace", True)):
                    self._axis_groups = validated
                else:
                    self._axis_groups.extend(validated)
                self._refresh_axes_grp_list()
                self._refresh_axes_summary()
                idx += 1
                continue
            if t == "RUN_ANALYZE":
                if not self._pdf_paths:
                    self._assistant_append_chat("err", "Нет PDF — добавьте файлы или откройте проект.")
                    idx += 1
                    continue

                def cont() -> None:
                    self._assistant_execute_actions(acts, idx + 1)

                self._assistant_after_analyze_cb = cont
                self._start_analyze()
                return
            if t == "SAVE_PROJECT":
                ok, note = self._save_project(notify=False)
                self._assistant_append_chat("sys", note)
                idx += 1
                continue
            if t == "EXPORT_XLSX":
                p = act.get("path")
                ps = str(p).strip() if isinstance(p, str) else ""
                ok, note = self._export_xlsx(save_path=ps if ps else None, notify=False)
                self._assistant_append_chat("sys", note or ("Excel сохранён." if ok else "Экспорт не выполнен."))
                idx += 1
                continue
            idx += 1
        self._assistant_append_chat("sys", "Цепочка действий выполнена.")

    def _assistant_release_analyze_chain(self) -> None:
        cb = self._assistant_after_analyze_cb
        self._assistant_after_analyze_cb = None
        if callable(cb):
            self._root.after(120, cb)

    def _paint_banner(self, _event: object | None = None) -> None:
        c = self._banner_canvas
        if c is None:
            return
        c.delete("all")
        w = max(c.winfo_width(), 4)
        h = 72
        c1, c2 = _hub_theme.ACCENT_DIM, _hub_theme.ACCENT
        r1, g1, b1 = int(c1[1:3], 16), int(c1[3:5], 16), int(c1[5:7], 16)
        r2, g2, b2 = int(c2[1:3], 16), int(c2[3:5], 16), int(c2[5:7], 16)
        for i in range(w):
            t = i / max(w - 1, 1)
            r = int(r1 + (r2 - r1) * t)
            g = int(g1 + (g2 - g1) * t)
            b = int(b1 + (b2 - b1) * t)
            hx = f"#{r:02x}{g:02x}{b:02x}"
            c.create_line(i, 0, i, h + 2, fill=hx, width=1)
        if self._busy:
            cx = (math.sin(self._shimmer_phase) * 0.5 + 0.5) * max(w - 1, 1)
            for dx in range(-20, 22, 4):
                x = int(cx + dx)
                if 0 <= x < w:
                    c.create_line(x, 0, x, h + 2, fill="#7dd3fc", width=2)
        c.create_text(24, 22, anchor=tk.W, text=_APP_DISPLAY_NAME, fill="#ffffff", font=("Segoe UI", 17, "bold"))
        c.create_text(
            24,
            50,
            anchor=tk.W,
            text=f"v{_APP_VERSION}  ·  ведомость · лист PDF · отчёт Excel · ассистент LLM",
            fill="#e0f2fe",
            font=("Segoe UI", 9),
        )

    def _banner_shimmer_tick(self) -> None:
        if not self._busy:
            self._shimmer_after = None
            return
        self._shimmer_phase += 0.2
        self._paint_banner()
        self._shimmer_after = self._root.after(48, self._banner_shimmer_tick)

    def _set_progress(self, pct: int, message: str) -> None:
        p = max(0, min(100, int(pct)))
        self._pb["value"] = p
        self._lbl_pct.config(text=f"{p}%")
        self._lbl_stage.config(text=f"Этап: {message}")

    def _add_pdfs(self) -> None:
        paths = filedialog.askopenfilenames(filetypes=[("PDF", "*.pdf"), ("Все файлы", "*.*")])
        if not paths:
            return
        seen = set(os.path.normpath(p) for p in self._pdf_paths)
        for p in paths:
            n = os.path.normpath(p)
            if n not in seen:
                seen.add(n)
                self._pdf_paths.append(p)
        self._sync_list()

    def _clear_paths(self) -> None:
        self._pdf_paths.clear()
        self._results.clear()
        self._cross_file_validation.clear()
        self._sync_list()
        self._refresh_table()
        self._refresh_validation_panel()
        self._lbl_strategy.pack(anchor=tk.W, pady=(0, 4))
        self._lbl_reason.pack(anchor=tk.W, pady=(0, 6))
        self._lbl_strategy.config(text="Способ расчёта: —")
        self._lbl_reason.config(text="")
        self._lbl_total_sub.config(text="")
        self._set_progress(0, "—")

    def _sync_list(self) -> None:
        self._list.delete(0, tk.END)
        for p in self._pdf_paths:
            self._list.insert(tk.END, os.path.basename(p))
        self._lbl_count.config(text=f"Файлов: {len(self._pdf_paths)}")

    def _remote_catalog_settings(self) -> None:
        if _catalog_remote is None:
            messagebox.showinfo("Каталог", "Модуль каталога недоступен.")
            return
        cfg = _catalog_remote.read_config()
        top = tk.Toplevel(self._root)
        top.title("Каталог из интернета")
        top.transient(self._root)
        top.grab_set()
        pad = {"padx": 10, "pady": 6}
        tk.Label(
            top,
            text="HTTPS-URL CSV в том же формате, что профили_база.csv (ключ;м2_на_пм;кг_на_пм).\n"
            "В первых строках можно указать версию: # версия: 2026-02-01",
            justify=tk.LEFT,
        ).pack(anchor=tk.W, **pad)
        fr = tk.Frame(top)
        fr.pack(fill=tk.X, **pad)
        tk.Label(fr, text="URL:").grid(row=0, column=0, sticky=tk.W)
        url_var = tk.StringVar(value=cfg.get("url") or "")
        tk.Entry(fr, textvariable=url_var, width=64).grid(row=0, column=1, sticky=tk.EW)
        fr.columnconfigure(1, weight=1)
        en_var = tk.BooleanVar(value=bool(cfg.get("enabled")))
        tk.Checkbutton(top, text="Использовать удалённый каталог (кеш перекрывает локальную базу, но не проектный CSV)", variable=en_var).pack(
            anchor=tk.W, **pad
        )
        fo_var = tk.BooleanVar(value=bool(cfg.get("fetch_on_analyze")))
        tk.Checkbutton(top, text="Перед каждым анализом пытаться обновить кеш", variable=fo_var).pack(anchor=tk.W, padx=10)
        tfr = tk.Frame(top)
        tfr.pack(anchor=tk.W, **pad)
        tk.Label(tfr, text="Таймаут, с (5–120):").pack(side=tk.LEFT)
        to_var = tk.StringVar(value=str(int(cfg.get("timeout_sec") or 25)))
        tk.Entry(tfr, textvariable=to_var, width=6).pack(side=tk.LEFT, padx=(6, 0))
        st = _catalog_remote.catalog_summary_line()
        tk.Label(top, text=f"Сейчас: {st}", fg=_UI_TEXT_DIM, font=("Segoe UI", 8)).pack(anchor=tk.W, **pad)

        def save() -> None:
            try:
                to_v = max(5, min(120, int(str(to_var.get()).strip() or "25")))
            except ValueError:
                to_v = 25
            _catalog_remote.write_config(
                {
                    "url": url_var.get().strip(),
                    "enabled": en_var.get(),
                    "fetch_on_analyze": fo_var.get(),
                    "timeout_sec": to_v,
                }
            )
            top.destroy()
            messagebox.showinfo("Каталог", "Настройки сохранены (каталог_url.json рядом с утилитой).")

        bf = tk.Frame(top)
        bf.pack(fill=tk.X, **pad)
        ttk.Button(bf, text="Сохранить", command=save, style="Accent.TButton").pack(side=tk.RIGHT, padx=4)
        ttk.Button(bf, text="Отмена", command=top.destroy, style="Secondary.TButton").pack(side=tk.RIGHT)

    def _remote_catalog_fetch_now(self) -> None:
        if _catalog_remote is None:
            messagebox.showinfo("Каталог", "Модуль каталога недоступен.")
            return
        cfg = _catalog_remote.read_config()
        if not (cfg.get("url") or "").strip():
            messagebox.showinfo("Каталог", "Укажите URL в «Каталог из интернета: настройка…».")
            return

        def work() -> None:
            ok, msg = _catalog_remote.fetch_remote_catalog()
            self._root.after(0, lambda: self._remote_fetch_done(ok, msg))

        threading.Thread(target=work, daemon=True).start()

    def _remote_fetch_done(self, ok: bool, msg: str) -> None:
        if ok:
            messagebox.showinfo("Каталог", msg)
        else:
            messagebox.showerror("Каталог", msg)

    def _export_missing_keys_csv(self) -> None:
        if not self._results:
            messagebox.showinfo("Ключи", "Сначала выполните «Анализ».")
            return
        rows: list[dict[str, Any]] = []
        for pdf_path, res in self._results.items():
            if res.strategy != "metal_catalog":
                continue
            ml = getattr(res, "metal_lines", None) or []
            for item in metal_lines_missing_catalog_hints(ml):
                rows.append({"file": os.path.basename(pdf_path), **item})
        if not rows:
            messagebox.showinfo("Ключи", "Нет позиций без м²/м (или режим не «металл»).")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="ключи_дополнить_в_каталог.csv",
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(["файл", "позиция", "профиль_raw", "profile_key", "ключ_для_csv"])
                for r in rows:
                    raw = str(r.get("profile_raw") or "")
                    pk = r.get("profile_key") or ""
                    sug = _normalize_profile_key(raw) or (pk or "")
                    w.writerow([r["file"], r["position"], raw, pk, sug])
        except OSError as e:
            messagebox.showerror("Ключи", str(e))
            return
        messagebox.showinfo("Ключи", f"Сохранено {len(rows)} строк:\n{path}")

    def _start_analyze(self) -> None:
        if not self._pdf_paths or fitz is None or analyze_pdf_document is None:
            messagebox.showinfo("Анализ", "Добавьте хотя бы один PDF.")
            self._assistant_release_analyze_chain()
            return
        if self._busy:
            return
        self._busy = True
        self._btn_analyze.config(state=tk.DISABLED)
        self._shimmer_phase = 0.0
        self._paint_banner()
        if self._shimmer_after:
            try:
                self._root.after_cancel(self._shimmer_after)
            except Exception:
                pass
            self._shimmer_after = None
        self._banner_shimmer_tick()
        paths = list(self._pdf_paths)

        def ui_progress(pct: int, msg: str) -> None:
            self._root.after(0, lambda: self._set_progress(pct, msg))

        def work() -> None:
            new_results: dict[str, AnalyzeResult] = {}
            err: str | None = None
            n = len(paths)
            try:
                if _catalog_remote is not None:
                    cfg0 = _catalog_remote.read_config()
                    if cfg0.get("enabled") and cfg0.get("fetch_on_analyze") and (cfg0.get("url") or "").strip():
                        ui_progress(0, "Обновление каталога из интернета…")
                        ok_f, msg_f = _catalog_remote.maybe_fetch_before_analyze()
                        if not ok_f and msg_f:
                            self._root.after(
                                0,
                                lambda m=msg_f: messagebox.showwarning("Каталог из интернета", m),
                            )
                for fi, path in enumerate(paths):
                    base_lo = int(100 * fi / max(1, n))
                    base_hi = int(100 * (fi + 1) / max(1, n))

                    def make_prog(lo: int, hi: int, pdf_path: str) -> Any:
                        def _inner(p: int, m: str) -> None:
                            span = max(1, hi - lo)
                            g = lo + int(span * p / 100)
                            ui_progress(g, f"[{os.path.basename(pdf_path)}] {m}")

                        return _inner

                    prog_cb = make_prog(base_lo, base_hi, path)
                    doc = fitz.open(path)
                    try:
                        goal_g = self._goal.get()
                        cat_p = self._catalog_path
                        tol = _VALIDATION_TOLERANCE_STRICT_PCT
                        new_results[path] = analyze_pdf_document(  # type: ignore[misc]
                            doc,
                            prog_cb,
                            goal=goal_g,
                            catalog_path=cat_p,
                            source_path=path,
                            validation_tolerance_pct=tol,
                            metal_options={
                                "util_dir": _утил_dir,
                                "merged_llm_ui": bool(self._merged_llm_force.get()),
                                "allow_vision_ui": bool(self._allow_vision_ui.get()),
                                "project_json_path": self._current_project_path,
                            },
                        )
                    finally:
                        doc.close()
            except Exception as e:
                err = str(e)
            self._root.after(0, lambda: self._finish_analyze(new_results, err))

        threading.Thread(target=work, daemon=True).start()

    def _finish_analyze(self, new_results: dict[str, AnalyzeResult], err: str | None) -> None:
        self._busy = False
        self._btn_analyze.config(state=tk.NORMAL)
        if self._shimmer_after:
            try:
                self._root.after_cancel(self._shimmer_after)
            except Exception:
                pass
            self._shimmer_after = None
        self._paint_banner()
        if err:
            self._rep(True)
            messagebox.showerror("Анализ", err)
            self._set_progress(0, "Ошибка")
            self._assistant_release_analyze_chain()
            return
        self._results = new_results
        self._rep(False)
        self._set_progress(100, "Готово")
        # Для одного PDF — краткая строка «Способ…»; для нескольких — блок скрыт (без простыни «по файлам»).
        if len(self._results) == 1:
            self._lbl_strategy.pack(anchor=tk.W, pady=(0, 4))
            self._lbl_reason.pack(anchor=tk.W, pady=(0, 6))
            r = next(iter(self._results.values()))
            self._lbl_strategy.config(
                text=f"Способ: {self._strategy_ru(r.strategy)} (уверенность: {r.confidence})"
            )
            reason_txt = r.reason
            if r.strategy == "metal_catalog":
                ex = getattr(r, "explicit_paint_area_m2", None)
                if ex is not None:
                    reason_txt += f" Явная площадь в тексте PDF: {ex:.4f} м²."
                elif getattr(r, "explicit_paint_area_ambiguous", False):
                    reason_txt += " В тексте несколько явных площадей окраски — см. таблицу результатов (колонка «Детально»)."
                md_one = getattr(r, "merged_diagnostics", None) or {}
                bs = md_one.get("bom_source")
                if bs:
                    reason_txt += f" Источник ведомости: {bs}."
                sp = md_one.get("spatial_layout_diag") or {}
                if sp.get("rows_total"):
                    reason_txt += (
                        f' Таблица по координатам: {int(sp["rows_total"])} строк'
                        f", средн.увер. {float(sp.get('avg_confidence') or 0):.2f}."
                    )
            self._lbl_reason.config(text=reason_txt)
        else:
            self._lbl_strategy.pack_forget()
            self._lbl_reason.pack_forget()
        self._cross_file_validation.clear()
        if cross_file_mark_checks is not None and len(self._results) > 1:
            mb = {
                p: r.metal_lines
                for p, r in self._results.items()
                if r.strategy == "metal_catalog" and (getattr(r, "metal_lines", None) or [])
            }
            if len(mb) > 1:
                self._cross_file_validation = cross_file_mark_checks(
                    mb,
                    _CROSS_FILE_MARK_TOLERANCE_STRICT_PCT,
                    min_abs_delta_m2=_CROSS_FILE_MARK_MIN_ABS_DELTA_M2,
                )
        self._refresh_table()
        bad = False
        for r in self._results.values():
            for v in getattr(r, "metal_validation", []) or []:
                if v.get("severity") == "error":
                    bad = True
                    break
            if bad:
                break
        if bad:
            messagebox.showwarning(
                "Проверки",
                "В разборе есть ошибки валидации — см. таблицу результатов (колонка «Ошибка», текст «Детально»).",
            )
        self._refresh_axes_summary()
        self._assistant_release_analyze_chain()

    @staticmethod
    def _strategy_ru(s: str) -> str:
        return {
            "spec": "спецификация / текст",
            "geometry": "размеры в тексте",
            "fallback": "формат листа",
            "metal_catalog": "металл (ведомость + каталог CSV)",
        }.get(
            s, s
        )

    def _apply_tree_columns_sheet(self) -> None:
        self._tree.config(show="headings", columns=("file", "page", "area", "detail"))
        self._tree.heading("file", text="Файл")
        self._tree.heading("page", text="Лист")
        self._tree.heading("area", text="м²")
        self._tree.heading("detail", text="Детально")
        self._tree.column("file", width=180, minwidth=80)
        self._tree.column("page", width=52, minwidth=40)
        self._tree.column("area", width=96, minwidth=64)
        self._tree.column("detail", width=320, minwidth=120)

    def _apply_tree_columns_metal(self, *, flat: bool) -> None:
        tv = self._tree
        w_flat, w_tree = _read_metal_col_widths_from_disk()
        anch = tk.CENTER
        mw = 32

        def _configure(col_id: str, label_text: str, width_px: int, *, stretch: bool) -> None:
            tv.heading(col_id, text=label_text, anchor=anch)
            tv.column(col_id, width=int(width_px), minwidth=max(26, mw), anchor=anch, stretch=stretch)

        if flat:
            cols = tuple(c for c, _ in _METAL_COL_LABELS_FLAT)
            tv.config(show="headings", columns=cols)
            for cid, lab in _METAL_COL_LABELS_FLAT:
                # stretch=False: иначе столбец «prof» съедает всё свободное место и width из настроек почти не заметен.
                _configure(cid, lab, int(w_flat[cid]), stretch=False)
        else:
            tree_body = [(c, l) for c, l in _METAL_COL_LABELS_TREE if c != "#0"]
            cols = tuple(c for c, _ in tree_body)
            tv.config(show="tree headings", columns=cols)
            h0_lab = _METAL_COL_LABELS_TREE[0][1]
            _configure("#0", h0_lab, int(w_tree["#0"]), stretch=False)
            for cid, lab in tree_body:
                _configure(cid, lab, int(w_tree[cid]), stretch=False)

    def _metal_open_column_widths_dialog(self) -> None:
        wf, wt = _read_metal_col_widths_from_disk()
        dlg = tk.Toplevel(self._root)
        dlg.title("Ширина колонок таблицы (металл)")
        dlg.transient(self._root)
        dlg.resizable(False, False)
        fr = tk.Frame(dlg, padx=14, pady=12)
        fr.pack(fill=tk.BOTH, expand=True)

        tk.Label(
            fr,
            text=(
                "Ширина задаётся для каждого столбца целиком. Колонки не растягиваются автоматически — "
                "при нехватке места используйте горизонтальную прокрутку под таблицей. "
                "Настройки сохраняются в файл metal_table_prefs.json рядом с утилитой."
            ),
            wraplength=520,
            justify=tk.LEFT,
            font=("Segoe UI", 9),
        ).pack(anchor=tk.W, pady=(0, 12))

        nb = ttk.Notebook(fr)
        nb.pack(fill=tk.BOTH, expand=True)

        def _spinbox_row(parent: tk.Widget, r: int, cid: str, lbl: str, init: int, var_holder: dict[str, tk.StringVar]) -> None:
            var_holder[cid] = tk.StringVar(value=str(int(init)))
            tk.Label(parent, text=lbl, width=26, anchor=tk.W).grid(row=r, column=0, sticky=tk.W, pady=2)
            tk.Spinbox(
                parent,
                from_=28,
                to=900,
                width=8,
                textvariable=var_holder[cid],
            ).grid(row=r, column=1, sticky=tk.W, padx=(8, 0), pady=2)

        flat_vars: dict[str, tk.StringVar] = {}
        tab_f = tk.Frame(nb, padx=4, pady=8)
        nb.add(tab_f, text="Один список")
        for rf, (cid, lab) in enumerate(_METAL_COL_LABELS_FLAT):
            _spinbox_row(tab_f, rf, cid, lab, wf[cid], flat_vars)

        tree_vars: dict[str, tk.StringVar] = {}
        tab_t = tk.Frame(nb, padx=4, pady=8)
        nb.add(tab_t, text="Группы марок")
        for rt, (cid, lab) in enumerate(_METAL_COL_LABELS_TREE):
            _spinbox_row(tab_t, rt, cid, lab, wt[cid], tree_vars)

        def _reset_defaults() -> None:
            df, dt = _merge_metal_col_widths(None)
            for cid in df:
                flat_vars[cid].set(str(df[cid]))
            for cid in dt:
                tree_vars[cid].set(str(dt[cid]))

        def _collect() -> tuple[dict[str, int], dict[str, int]] | tuple[None, None]:
            out_f: dict[str, int] = {}
            out_t: dict[str, int] = {}
            try:
                for cid in _METAL_COL_DEFAULT_WIDTHS_FLAT:
                    out_f[cid] = max(28, min(900, int(float(str(flat_vars[cid].get()).replace(",", ".").strip()))))
                for cid in _METAL_COL_DEFAULT_WIDTHS_TREE:
                    out_t[cid] = max(28, min(900, int(float(str(tree_vars[cid].get()).replace(",", ".").strip()))))
            except ValueError:
                messagebox.showerror("Колонки", "Укажите целые числа ширины (28–900).", parent=dlg)
                return None, None
            return out_f, out_t

        btns = tk.Frame(fr)
        btns.pack(fill=tk.X, pady=(14, 0))
        ttk.Button(btns, text="Умолчания", command=_reset_defaults, style="Secondary.TButton").pack(side=tk.LEFT)
        ttk.Button(btns, text="Отмена", command=dlg.destroy, style="Secondary.TButton").pack(side=tk.RIGHT, padx=(8, 0))

        def _save() -> None:
            nf, nt = _collect()
            if nf is None:
                return
            try:
                _write_metal_col_widths_disk(nf, nt)
            except OSError as e:
                messagebox.showerror("Колонки", str(e), parent=dlg)
                return
            dlg.destroy()
            self._refresh_table()

        ttk.Button(btns, text="Применить", command=_save, style="Accent.TButton").pack(side=tk.RIGHT)

    def _schedule_metal_table_refresh(self, _evt: Any | None = None) -> None:
        aid = self._metal_table_after_id
        if aid:
            try:
                self._root.after_cancel(aid)
            except Exception:
                pass
        self._metal_table_after_id = self._root.after(150, self._metal_table_refresh_deferred)

    def _metal_table_refresh_deferred(self) -> None:
        self._metal_table_after_id = None
        self._refresh_table()

    def _sync_metal_result_tools(self, metal_active: bool) -> None:
        st_m = tk.NORMAL if metal_active else tk.DISABLED
        for w in self._metal_tool_widgets:
            try:
                w.configure(state=st_m)
            except tk.TclError:
                pass

    def _tree_double_click_copy_row(self, event: tk.Event) -> None:  # type: ignore[type-arg]
        iid = self._tree.identify_row(event.y)
        if not iid:
            return
        self._tree.selection_set(iid)
        vals = self._tree.item(iid, "values")
        text = self._tree.item(iid, "text") or ""
        line = text + "\t" + "\t".join(str(v) for v in vals)
        self._copy_line(line)

    def _tree_expand_all(self) -> None:
        for rid in self._tree.get_children(""):
            self._tree.item(rid, open=True)

    def _tree_collapse_marks(self) -> None:
        for rid in self._tree.get_children(""):
            self._tree.item(rid, open=False)

    def _merged_project_level_validation_warns(self) -> bool:
        codes = {"explicit_paint_orders_magnitude", "merged_pdf_uncertain_block"}
        for p in self._pdf_paths:
            res = self._results.get(p)
            if not res:
                continue
            md = getattr(res, "merged_diagnostics", None) or {}
            if md.get("forbid_clean_success"):
                return True
            for v in getattr(res, "metal_validation", []) or []:
                if str(v.get("code") or "") in codes:
                    return True
        return False

    def _export_merged_debug_json(self) -> None:
        if not self._results:
            messagebox.showinfo("Отладка", "Выполните сначала «Анализ».")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON", "*.json")],
            initialfile="akz_merged_debug.json",
        )
        if not path:
            return
        out: dict[str, Any] = {
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "pdf_paths": list(self._pdf_paths),
            "flags": {
                "merged_llm_ui": bool(self._merged_llm_force.get()),
                "allow_vision_ui": bool(self._allow_vision_ui.get()),
            },
            "files": {},
        }
        for p, r in self._results.items():
            out["files"][os.path.basename(p)] = {
                "merged_diagnostics": getattr(r, "merged_diagnostics", {}) or {},
                "metal_validation_tail": (getattr(r, "metal_validation", None) or [])[:80],
                "quality_metrics": getattr(r, "quality_metrics", {}) or {},
            }
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(out, f, ensure_ascii=False, indent=2)
        except OSError as e:
            messagebox.showerror("Отладка", str(e))
            return
        messagebox.showinfo("Отладка", f"Сохранено:\n{path}")

    def _metal_row_buckets(self) -> tuple[int, int, int, int]:
        """Позиций всего; без м²/м в CSV; предупр. сверки при наличии м²/м; жёсткие !."""
        n_pos = n_miss_m2 = n_warn_m2 = n_issue = 0
        for path in self._pdf_paths:
            res = self._results.get(path)
            if not res or res.strategy != "metal_catalog":
                continue
            for ml in getattr(res, "metal_lines", None) or []:
                n_pos += 1
                m2p = ml.get("m2_per_m")
                st = str(ml.get("row_status") or "ok")
                sig = _metal_line_profile_signature(ml)
                disp = _metal_line_profile_display_ui(ml)
                if (sig or disp.strip()) and m2p is None:
                    n_miss_m2 += 1
                if st == "issue":
                    n_issue += 1
                elif st == "warn" and m2p is not None:
                    n_warn_m2 += 1
        return n_pos, n_miss_m2, n_warn_m2, n_issue

    def _update_suspicion_label(self) -> None:
        n_pos, n_miss_m2, n_warn_m2, n_issue = self._metal_row_buckets()
        tol = _VALIDATION_TOLERANCE_STRICT_PCT
        try:
            total_val_txt = str(self._lbl_total_value.cget("text") or "").strip()
        except tk.TclError:
            total_val_txt = ""
        use_metal_ui = False
        for path in self._pdf_paths:
            res = self._results.get(path)
            if res and res.strategy == "metal_catalog" and (getattr(res, "metal_lines", None) or []):
                use_metal_ui = True
                break

        if not self._results:
            self._lbl_suspicion.config(text="", fg=_UI_TEXT_DIM, font=("Segoe UI", 9))
            return

        any_metal = any(
            self._results.get(p) and self._results[p].strategy == "metal_catalog" for p in self._pdf_paths
        )
        if not any_metal:
            self._lbl_suspicion.config(
                text=(
                    "Повышенный риск по позициям ведомости не считается "
                    "(в результатах нет режима металла). Колонка «Ошибка» относится только к ведомости."
                ),
                fg=_UI_TEXT_DIM,
                font=("Segoe UI", 9),
            )
            return

        if n_pos == 0:
            self._lbl_suspicion.config(
                text="Позиции ведомости не найдены — нечего сравнивать.",
                fg=_UI_TEXT_DIM,
                font=("Segoe UI", 9),
            )
            self._lbl_total_sub.config(
                text="Почему: нет блока «Спецификация деталей / Specification» или строки не распознались. Проверьте PDF.",
                fg=_UI_TEXT_DIM,
                font=("Segoe UI", 9),
            )
            return

        n_xf = len(self._cross_file_validation)
        merged_pdf_warn = self._merged_project_level_validation_warns()
        all_ok = (
            n_miss_m2 == 0
            and n_warn_m2 == 0
            and n_issue == 0
            and n_xf == 0
            and not merged_pdf_warn
        )

        if merged_pdf_warn and n_miss_m2 == 0 and n_issue == 0 and n_xf == 0:
            self._lbl_suspicion.config(
                text=(
                    "Формально по строкам таблицы нет «?» и символов в колонке «Ошибка», но для объединённого PDF сработала доп. проверка:\n"
                    "возможна ошибка границы BOM и/или расхождение с явной площадью в тексте PDF.\n"
                    "LLM для крупных файлов при необходимости запускается автоматически; для жёсткого вызова включите "
                    "«Принудительно LLM при ≥4 стр.» в «Ещё» → «Объединённые PDF (КМД)», при сомнении — «Отладка merged: сохранить JSON»."
                ),
                fg="#b45309",
                font=("Segoe UI", 9),
            )
            if total_val_txt and total_val_txt != "—":
                self._lbl_total_sub.config(
                    text=(f"Итог {total_val_txt} м² — перепроверьте при объединённом КМД."),
                    fg="#b45309",
                    font=("Segoe UI", 9),
                )
            else:
                self._lbl_total_sub.config(text="", fg=_UI_TEXT_DIM, font=("Segoe UI", 9))
            return

        if all_ok:
            self._lbl_suspicion.config(
                text=(
                    "Успешно. Расчёт завершён без замечаний.\n"
                    f"Позиций ведомости: {n_pos}. Для каждой найден коэффициент м²/м в CSV; сверка масс с ведомостью "
                    f"в допуске ±{tol}% (в колонке «Ошибка» нет «?» и «!»).\n"
                ),
                fg="#047857",
                font=("Segoe UI", 10),
            )
            if total_val_txt and total_val_txt != "—":
                metal_hint = (
                    " У строки марки — м² суммарно по всем отправочным комплектам; у позиций — м² на один комплект "
                    "(как в спецификации)."
                    if use_metal_ui
                    else ""
                )
                self._lbl_total_sub.config(
                    text=(
                        f"Итог {total_val_txt} м² — полный: все позиции участвовали в сумме площади АКЗ.{metal_hint}"
                    ),
                    fg="#047857",
                    font=("Segoe UI", 9),
                )
            else:
                self._lbl_total_sub.config(text="", fg=_UI_TEXT_DIM, font=("Segoe UI", 9))
            return

        blocks: list[str] = [
            "Обнаружены проблемы — прочитайте пункты ниже (что не так и почему).\n",
        ]
        if merged_pdf_warn:
            blocks.append(
                "• Объединённый PDF — повышенный риск неверной границы ведомости "
                "(см. также предупреждения в тексте ниже).\n\n"
            )
        if n_miss_m2:
            blocks.append(
                f"• Неполный справочник CSV ({n_miss_m2} поз.). Что: нет коэффициента м²/м.\n"
                f"  Почему: для обозначения профиля из ведомости нет подходящей строки в каталоге "
                f"(или текст в PDF распознан не так, как записано в CSV). Площадь этих строк — 0 м².\n"
                f"  Действие: «Ещё» → «Ключи без м²/м в каталоге (CSV)…» или дополните CSV.\n\n"
            )
        if n_warn_m2:
            warnc = Counter()
            for path in self._pdf_paths:
                res = self._results.get(path)
                if not res or res.strategy != "metal_catalog":
                    continue
                for v in getattr(res, "metal_validation", []) or []:
                    if v.get("severity") == "warning" and v.get("code"):
                        warnc[str(v["code"])] += 1
            head = (
                f"• Предупреждение сверки ({n_warn_m2} поз., в таблице «?»).\n"
                f"  Что: м²/м в каталоге есть, но масса/длина не сходятся с ведомостью сильнее {tol}%.\n"
                f"  Почему: неверный кг/м в CSV, другая марка/сечение или ошибка в массе в PDF.\n"
            )
            if warnc:
                top_w = sorted(warnc.items(), key=lambda kv: (-kv[1], kv[0]))[:8]
                head += (
                    f"  Частые коды сверки (все сообщения): "
                    f"{', '.join(f'{c} ({n})' for c, n in top_w)}.\n"
                    f"  Это сверка данных, не ошибка выполнения программы.\n\n"
                )
            else:
                head += "\n"
            blocks.append(head)
        if n_issue:
            blocks.append(
                f"• Ошибка сверки ({n_issue} поз., колонка «Ошибка»).\n"
                f"  Что: жёсткое расхождение контрольных соотношений.\n"
                f"  Почему: несоответствие массы и коэффициентов каталога фактическим данным ведомости.\n\n"
            )
        if n_xf:
            blocks.append(
                f"• Межфайловые замечания ({n_xf}).\n"
                f"  Что: при анализе нескольких PDF расхождения по маркам или массам между листами.\n"
                f"  Почему: разные ведомости/маркировка или неполный набор файлов.\n\n"
            )

        sev_fg = "#b91c1c" if (n_issue or n_miss_m2) else "#b45309"
        self._lbl_suspicion.config(text="".join(blocks).rstrip(), fg=sev_fg, font=("Segoe UI", 9))

        sub = []
        if total_val_txt and total_val_txt != "—":
            sub.append(
                f"Число «Итого по таблице» сейчас {total_val_txt} м² — в сумму входят только строки с заполненным м²/м; "
                f"где «—», площадь 0 м²."
            )
        else:
            sub.append("Итог не выведен — нет данных для суммы.")
        self._lbl_total_sub.config(
            text=" ".join(sub),
            fg=sev_fg,
            font=("Segoe UI", 9),
        )

    def _refresh_validation_panel(self) -> None:
        self._update_suspicion_label()

    @staticmethod
    def _file_sha256(path: str) -> str | None:
        if not path or not os.path.isfile(path):
            return None
        h = hashlib.sha256()
        try:
            with open(path, "rb") as f:
                for chunk in iter(lambda: f.read(65536), b""):
                    h.update(chunk)
            return h.hexdigest()[:16]
        except OSError:
            return None

    def _export_validation_json(self) -> None:
        if not self._results:
            messagebox.showinfo("JSON", "Сначала выполните «Анализ».")
            return
        initial = self._suggested_export_basename() + "_проверки.json"
        path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON", "*.json")],
            initialfile=initial,
        )
        if not path:
            return
        if self._catalog_path and str(self._catalog_path).strip():
            cat_paths = [str(self._catalog_path).strip()]
        else:
            try:
                cat_paths = list(default_catalog_paths())
            except Exception:
                p0 = default_catalog_path()
                cat_paths = [p0] if p0 and os.path.isfile(p0) else []
        cat_label = cat_paths[0] if len(cat_paths) == 1 else " | ".join(cat_paths)
        sha_parts = [self._file_sha256(p) for p in cat_paths]
        sha_combo = ";".join(h for h in sha_parts if h) or None
        blob: dict[str, Any] = {
            "app_version": _APP_VERSION,
            "validation_tolerance_pct": _VALIDATION_TOLERANCE_STRICT_PCT,
            "catalog_path": cat_label,
            "catalog_sha256_prefix": sha_combo,
            "cross_file_checks": self._cross_file_validation,
            "files": {},
        }
        if _catalog_remote is not None:
            blob["catalog_remote_meta"] = _catalog_remote.read_meta()
            rcfg = _catalog_remote.read_config()
            blob["catalog_remote_config"] = {
                "enabled": bool(rcfg.get("enabled")),
                "fetch_on_analyze": bool(rcfg.get("fetch_on_analyze")),
                "url_present": bool((rcfg.get("url") or "").strip()),
            }
        agg_codes: Counter[str] = Counter()
        agg_sev: Counter[str] = Counter()
        for pdf_path in self._pdf_paths:
            res = self._results.get(pdf_path)
            if not res:
                continue
            mv = getattr(res, "metal_validation", []) or []
            ml = getattr(res, "metal_lines", []) or []
            summ = summarize_metal_validation(mv)
            for row in summ.get("codes_top") or []:
                if isinstance(row, dict) and row.get("code") is not None:
                    try:
                        agg_codes[str(row["code"])] += int(row.get("count") or 0)
                    except (TypeError, ValueError):
                        pass
            for sk, sv in (summ.get("by_severity") or {}).items():
                try:
                    agg_sev[str(sk)] += int(sv)
                except (TypeError, ValueError):
                    pass
            miss_hints: list[dict[str, Any]] = []
            try:
                miss_hints = list(metal_lines_missing_catalog_hints(ml))
            except Exception:
                miss_hints = []
            blob["files"][pdf_path] = {
                "strategy": res.strategy,
                "merged_diagnostics": getattr(res, "merged_diagnostics", {}) or {},
                "metal_validation": mv,
                "validation_summary": summ,
                "quality_metrics": getattr(res, "quality_metrics", {}),
                "explicit_paint_area_m2": getattr(res, "explicit_paint_area_m2", None),
                "explicit_paint_area_ambiguous": getattr(res, "explicit_paint_area_ambiguous", False),
                "metal_lines": ml,
                "metal_lines_signals": _metal_export_row_signals(ml),
                "metal_lines_missing_catalog_hints": miss_hints,
                "shipment_qty_by_mark": dict(getattr(res, "shipment_qty_by_mark", None) or {}),
            }
        blob["validation_aggregate"] = {
            "by_severity": dict(sorted(agg_sev.items(), key=lambda kv: kv[0])),
            "codes_top": [
                {"code": c, "count": n}
                for c, n in sorted(agg_codes.items(), key=lambda kv: (-kv[1], kv[0]))[:48]
            ],
        }
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(blob, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("JSON", f"Сохранено:\n{path}")
            self._rep(False)
        except Exception as e:
            self._rep(True)
            messagebox.showerror("JSON", str(e))

    def _copy_summary(self) -> None:
        parts = [f"{_APP_DISPLAY_NAME} v{_APP_VERSION}"]
        tot = 0.0
        for path in self._pdf_paths:
            res = self._results.get(path)
            if not res:
                continue
            t = sum(float(p.area_m2) for p in res.per_page)
            tot += t
            qm = getattr(res, "quality_metrics", {}) or {}
            parts.append(
                f"{os.path.basename(path)}: {t:.4f} м², индекс {qm.get('confidence_score', '—')}, "
                f"+1%к кат ≈ +{qm.get('sensitivity_delta_m2_if_catalog_plus_1pct', '—')} м²"
            )
        parts.append(f"ИТОГО таблица: {tot:.4f} м²")
        self._root.clipboard_clear()
        self._root.clipboard_append("\n".join(parts))
        messagebox.showinfo("Буфер", "Сводка скопирована в буфер обмена.")

    def _tree_context_menu(self, event: tk.Event) -> None:  # type: ignore[type-arg]
        iid = self._tree.identify_row(event.y)
        if not iid:
            return
        self._tree.selection_set(iid)
        vals = self._tree.item(iid, "values")
        text = self._tree.item(iid, "text") or ""
        line = text + "\t" + "\t".join(str(v) for v in vals)
        m = tk.Menu(self._tree, tearoff=0)
        m.add_command(label="Копировать строку", command=lambda: self._copy_line(line))
        try:
            m.tk_popup(event.x_root, event.y_root)
        finally:
            m.grab_release()

    def _copy_line(self, line: str) -> None:
        self._root.clipboard_clear()
        self._root.clipboard_append(line)

    def _refresh_table(self) -> None:
        for i in self._tree.get_children():
            self._tree.delete(i)
        use_metal_ui = False
        for path in self._pdf_paths:
            res = self._results.get(path)
            if res and res.strategy == "metal_catalog" and (getattr(res, "metal_lines", None) or []):
                use_metal_ui = True
                break
        total = 0.0
        flt = self._metal_filter_var.get()
        flat_m = bool(self._metal_flat_list.get())
        if use_metal_ui:
            try:
                self._tree_hscroll.grid(row=1, column=0, columnspan=2, sticky="ew")
            except tk.TclError:
                pass

            def _len_cell_ml(ml_: dict[str, Any]) -> str:
                lm_ = ml_.get("length_mm")
                if lm_ is None:
                    return ""
                try:
                    return f"{float(lm_):.0f}"
                except (TypeError, ValueError):
                    return str(lm_)

            self._apply_tree_columns_metal(flat=flat_m)
            zi_global = 0
            for path in self._pdf_paths:
                res = self._results.get(path)
                if not res:
                    continue
                base = os.path.basename(path)
                mlns = getattr(res, "metal_lines", None) or []
                if res.strategy == "metal_catalog" and mlns:
                    by_mark: dict[str, list[dict[str, Any]]] = defaultdict(list)
                    for ml in mlns:
                        mk = str(ml.get("assembly_mark") or "Без марки")
                        by_mark[mk].append(ml)
                    for mark in sorted(by_mark.keys(), key=lambda x: (len(x), x)):
                        lines_raw = sorted(by_mark[mark], key=_metal_line_position_sort)
                        vis_lines = [
                            ml for ml in lines_raw if _metal_filter_matches_row(mark, ml, flt)
                        ]
                        if not vis_lines:
                            continue
                        subtot = sum(float(x.get("area_m2") or 0) for x in vis_lines)
                        total += subtot
                        sq = max(1, int(lines_raw[0].get("shipment_qty") or 1))

                        if flat_m:
                            for ml in vis_lines:
                                zi_global += 1
                                zt = "z_even" if zi_global % 2 == 0 else "z_odd"
                                pr = _metal_line_profile_display_ui(ml)[:160]
                                if not pr:
                                    pr = "—"
                                apc = _metal_line_area_m2_per_piece(ml)
                                m2full = float(ml.get("area_m2") or 0)
                                row_st = str(ml.get("row_status") or "ok")
                                fl = "!" if row_st == "issue" else ("?" if row_st == "warn" else "")
                                tags = (zt, row_st) if row_st in ("issue", "warn") else (zt,)
                                shp = max(1, int(ml.get("shipment_qty") or 1))
                                self._tree.insert(
                                    "",
                                    tk.END,
                                    text="",
                                    values=(
                                        mark,
                                        str(ml.get("position") or ""),
                                        pr,
                                        _len_cell_ml(ml),
                                        str(ml.get("qty") or ""),
                                        str(shp),
                                        f"{m2full:.4f}",
                                        f"{apc:.4f}",
                                        fl,
                                    ),
                                    tags=tags,
                                )
                        else:
                            n_vis = len(vis_lines)
                            grp_prof = (
                                f"{n_vis} поз., раскройте"
                                if n_vis > 1
                                else "см. ниже"
                            )
                            mid = self._tree.insert(
                                "",
                                tk.END,
                                text=f" {mark}  ·  {sq}× отпр.",
                                values=(
                                    grp_prof,
                                    "",
                                    "",
                                    f"{subtot:.4f}",
                                    "",
                                    "",
                                ),
                                tags=("depth1",),
                                open=False,
                            )
                            for zi, ml in enumerate(vis_lines):
                                zt = "z_even" if zi % 2 == 0 else "z_odd"
                                pr = _metal_line_profile_display_ui(ml)[:160]
                                if not pr:
                                    pr = "—"
                                apc = _metal_line_area_m2_per_piece(ml)
                                m2full = float(ml.get("area_m2") or 0)
                                row_st = str(ml.get("row_status") or "ok")
                                fl = "!" if row_st == "issue" else ("?" if row_st == "warn" else "")
                                tags = (zt, row_st) if row_st in ("issue", "warn") else (zt,)
                                pos_lbl = str(ml.get("position") or "").strip()
                                self._tree.insert(
                                    mid,
                                    tk.END,
                                    text=(f" № {pos_lbl}" if pos_lbl else " —"),
                                    values=(
                                        pr,
                                        _len_cell_ml(ml),
                                        str(ml.get("qty", "")),
                                        f"{m2full:.4f}",
                                        f"{apc:.4f}",
                                        fl,
                                    ),
                                    tags=tags,
                                )
                else:
                    s = sum(pr.area_m2 for pr in res.per_page)
                    total += s
                    if flat_m:
                        self._tree.insert(
                            "",
                            tk.END,
                            text="",
                            values=(
                                f"[{base}]",
                                "—",
                                f"режим «лист PDF» ({self._strategy_ru(res.strategy)})",
                                "",
                                "",
                                "",
                                f"{s:.4f}",
                                "",
                                "",
                            ),
                            tags=("depth0",),
                        )
                    else:
                        self._tree.insert(
                            "",
                            tk.END,
                            text=f" {base}",
                            values=(
                                f"лист PDF · {self._strategy_ru(res.strategy)}",
                                "",
                                "",
                                f"{s:.4f}",
                                "",
                                "",
                            ),
                            tags=("depth0",),
                        )
            if total > 0:
                if flat_m:
                    self._tree.insert(
                        "",
                        tk.END,
                        text="",
                        values=(
                            "Σ итого",
                            "",
                            "",
                            "",
                            "",
                            "",
                            f"{total:.4f}",
                            "",
                            "",
                        ),
                        tags=("total_row",),
                    )
                else:
                    self._tree.insert(
                        "",
                        tk.END,
                        text=" Σ итого",
                        values=("", "", "", f"{total:.4f}", "", ""),
                        tags=("total_row",),
                    )
        else:
            try:
                self._tree_hscroll.grid_remove()
            except tk.TclError:
                pass
            self._apply_tree_columns_sheet()
            sheet_i = 0
            for path in self._pdf_paths:
                res = self._results.get(path)
                if not res:
                    continue
                base = os.path.basename(path)
                for pr in res.per_page:
                    total += pr.area_m2
                    zt = "z_even" if sheet_i % 2 == 0 else "z_odd"
                    sheet_i += 1
                    self._tree.insert(
                        "",
                        tk.END,
                        values=(base, str(pr.page_index + 1), f"{pr.area_m2:.4f}", pr.detail[:200]),
                        tags=(zt,),
                    )
        if total:
            self._lbl_total_value.config(text=f"{total:.4f}")
            self._lbl_total_unit.config(text="м²")
            if use_metal_ui and flt.strip():
                self._lbl_total_sub.config(
                    text=f"Сумма только по строкам, попавшим в фильтр («{flt.strip()}»).",
                    fg=_UI_TEXT_DIM,
                    font=("Segoe UI", 9),
                )
            elif use_metal_ui:
                self._lbl_total_sub.config(text="", fg=_UI_TEXT_DIM, font=("Segoe UI", 9))
            else:
                self._lbl_total_sub.config(text="", fg=_UI_TEXT_DIM, font=("Segoe UI", 9))
        else:
            self._lbl_total_value.config(text="—")
            self._lbl_total_unit.config(text="")
            if use_metal_ui and flt.strip():
                self._lbl_total_sub.config(
                    text="По фильтру не осталось строк — очистите поле или измените подстроку.",
                    fg=_UI_TEXT_DIM,
                    font=("Segoe UI", 9),
                )
            else:
                self._lbl_total_sub.config(text="", fg=_UI_TEXT_DIM, font=("Segoe UI", 9))
        self._sync_metal_result_tools(use_metal_ui)
        self._refresh_validation_panel()
        self._refresh_axes_summary()

    def _load_axis_groups_disk(self) -> None:
        path = _axis_groups_json_path()
        self._axis_groups = []
        try:
            if not os.path.isfile(path):
                return
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            raw = data.get("groups")
            if not isinstance(raw, list):
                return
            for it in raw:
                if not isinstance(it, dict):
                    continue
                name = str(it.get("name") or "").strip() or "Группа"
                fr = str(it.get("file_re") or "").strip()
                mr = str(it.get("mark_re") or "").strip()
                self._axis_groups.append({"name": name, "file_re": fr, "mark_re": mr})
        except Exception:
            self._axis_groups = []

    def _save_axis_groups_disk(self) -> None:
        path = _axis_groups_json_path()
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"groups": self._axis_groups}, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showwarning("Оси и зоны", f"Не удалось сохранить правила:\n{e}")

    def _get_km_kmd_base(self) -> dict[str, Any]:
        if self._km_kmd_cache is not None:
            return self._km_kmd_cache
        path = _km_kmd_base_path()
        data: dict[str, Any] = {}
        try:
            if os.path.isfile(path):
                with open(path, encoding="utf-8") as f:
                    raw = json.load(f)
                if isinstance(raw, dict):
                    data = raw
        except Exception:
            data = {}
        self._km_kmd_cache = data
        return data

    def _show_km_kmd_reference(self) -> None:
        base = self._get_km_kmd_base()
        top = tk.Toplevel(self._root)
        top.title("База КМ и КМД")
        top.transient(self._root)
        top.geometry("760x560")
        fr = tk.Frame(top, padx=10, pady=8)
        fr.pack(fill=tk.BOTH, expand=True)
        tk.Label(fr, text=base.get("title") or "КМ / КМД", font=("Segoe UI", 12, "bold")).pack(anchor=tk.W)
        intro = base.get("intro") or ""
        if intro:
            tk.Label(fr, text=intro, font=("Segoe UI", 9), fg=_UI_TEXT_DIM, wraplength=720, justify=tk.LEFT).pack(
                anchor=tk.W, pady=(6, 8)
            )
        nb = ttk.Notebook(fr)
        nb.pack(fill=tk.BOTH, expand=True, pady=8)
        tab_gl = tk.Frame(nb, bg=_UI_LIST_BG)
        nb.add(tab_gl, text="Термины")
        tx1 = tk.Text(
            tab_gl,
            wrap=tk.WORD,
            height=18,
            font=("Segoe UI", 10),
            relief=tk.FLAT,
            padx=6,
            pady=6,
            bg=_hub_theme.TEXT_AREA_BG,
            fg=_UI_TEXT,
            insertbackground=_UI_TEXT,
            selectbackground=_hub_theme.SIDEBAR_SELECT_BG,
            selectforeground=_hub_theme.SIDEBAR_SELECT_FG,
            highlightthickness=0,
        )
        sb1 = ttk.Scrollbar(tab_gl, command=tx1.yview)
        tx1.configure(yscrollcommand=sb1.set)
        tx1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb1.pack(side=tk.RIGHT, fill=tk.Y)
        lines: list[str] = []
        for it in base.get("glossary") or []:
            if not isinstance(it, dict):
                continue
            ab = str(it.get("abbr") or "")
            full = str(it.get("full") or "")
            role = str(it.get("role") or "")
            akz = str(it.get("akz") or "")
            lines.append(f"{ab} — {full}\n\n{role}\n\nАКЗ: {akz}\n\n{'—' * 40}\n\n")
        tx1.insert(tk.END, "".join(lines) if lines else "Нет данных. Проверьте файл км_кмд_база.json рядом с утилитой.")
        tx1.configure(state=tk.DISABLED)

        tab_hi = tk.Frame(nb, bg=_UI_LIST_BG)
        nb.add(tab_hi, text="Имена файлов")
        tx2 = tk.Text(
            tab_hi,
            wrap=tk.WORD,
            height=18,
            font=("Segoe UI", 10),
            relief=tk.FLAT,
            padx=6,
            pady=6,
            bg=_hub_theme.TEXT_AREA_BG,
            fg=_UI_TEXT,
            insertbackground=_UI_TEXT,
            selectbackground=_hub_theme.SIDEBAR_SELECT_BG,
            selectforeground=_hub_theme.SIDEBAR_SELECT_FG,
            highlightthickness=0,
        )
        sb2 = ttk.Scrollbar(tab_hi, command=tx2.yview)
        tx2.configure(yscrollcommand=sb2.set)
        tx2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb2.pack(side=tk.RIGHT, fill=tk.Y)
        htxt = "\n".join(f"• {x}" for x in (base.get("filename_hints") or [])) or "—"
        tx2.insert(tk.END, htxt)
        tx2.configure(state=tk.DISABLED)

        tab_pr = tk.Frame(nb, bg=_UI_LIST_BG)
        nb.add(tab_pr, text="Шаблоны зон")
        tx3 = tk.Text(
            tab_pr,
            wrap=tk.WORD,
            height=18,
            font=("Consolas", 9),
            relief=tk.FLAT,
            padx=6,
            pady=6,
            bg=_hub_theme.TEXT_AREA_BG,
            fg=_UI_TEXT,
            insertbackground=_UI_TEXT,
            selectbackground=_hub_theme.SIDEBAR_SELECT_BG,
            selectforeground=_hub_theme.SIDEBAR_SELECT_FG,
            highlightthickness=0,
        )
        sb3 = ttk.Scrollbar(tab_pr, command=tx3.yview)
        tx3.configure(yscrollcommand=sb3.set)
        tx3.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb3.pack(side=tk.RIGHT, fill=tk.Y)
        pr_txt: list[str] = []
        for zp in base.get("zone_presets") or []:
            if not isinstance(zp, dict):
                continue
            pr_txt.append(f"«{zp.get('title', '')}»\n{zp.get('description', '')}\n")
            for g in zp.get("groups") or []:
                if isinstance(g, dict):
                    pr_txt.append(f"  • {g.get('name')}: {g.get('file_re')}\n")
            pr_txt.append("\n")
        tx3.insert(tk.END, "".join(pr_txt) if pr_txt else "Нет пресетов в базе.")
        tx3.configure(state=tk.DISABLED)

        foot = "\n".join(f"• {x}" for x in (base.get("links_read") or []))
        if foot.strip():
            tk.Label(fr, text=foot, font=("Segoe UI", 9), fg=_UI_TEXT_DIM, wraplength=720, justify=tk.LEFT).pack(
                anchor=tk.W, pady=(6, 0)
            )
        ttk.Button(fr, text="Закрыть", command=top.destroy).pack(anchor=tk.E, pady=(8, 0))

    def _apply_km_kmd_presets_dialog(self) -> None:
        base = self._get_km_kmd_base()
        presets = [p for p in (base.get("zone_presets") or []) if isinstance(p, dict)]
        if not presets:
            messagebox.showinfo(
                "КМ / КМД",
                f"В файле «{os.path.basename(_km_kmd_base_path())}» нет шаблонов zone_presets или файл не найден.",
            )
            return
        top = tk.Toplevel(self._root)
        top.title("Шаблоны зон КМ / КМД")
        top.transient(self._root)
        top.grab_set()
        pv = {"padx": 8, "pady": 4}
        tk.Label(
            top,
            text="Выберите набор правил. Его можно вставить в таблицу зон целиком или добавить в конец.",
            wraplength=480,
            justify=tk.LEFT,
        ).grid(row=0, column=0, columnspan=2, sticky=tk.W, **pv)
        lb = tk.Listbox(top, height=min(12, len(presets) + 2), width=64, font=("Segoe UI", 9))
        for i, zp in enumerate(presets):
            lb.insert(tk.END, f"{zp.get('title', f'Набор {i + 1}')}")
        lb.grid(row=1, column=0, columnspan=2, sticky=tk.NSEW, **pv)
        top.columnconfigure(0, weight=1)
        mode = tk.StringVar(value="replace")
        tk.Radiobutton(top, text="Заменить текущие зоны", variable=mode, value="replace").grid(
            row=2, column=0, sticky=tk.W, **pv
        )
        tk.Radiobutton(top, text="Добавить в конец списка", variable=mode, value="append").grid(
            row=3, column=0, sticky=tk.W, **pv
        )

        def apply() -> None:
            sel = lb.curselection()
            if not sel:
                messagebox.showinfo("КМ / КМД", "Выберите строку в списке.")
                return
            zp = presets[int(sel[0])]
            raw_groups = zp.get("groups")
            if not isinstance(raw_groups, list):
                messagebox.showerror("КМ / КМД", "В шаблоне нет массива groups.")
                return
            new_groups: list[dict[str, str]] = []
            for g in raw_groups:
                if not isinstance(g, dict):
                    continue
                name = str(g.get("name") or "").strip() or "Зона"
                fr = str(g.get("file_re") or "").strip()
                mr = str(g.get("mark_re") or "").strip()
                if not fr:
                    continue
                try:
                    re.compile(fr, re.I)
                except re.error as e:
                    messagebox.showerror("КМ / КМД", f"Ошибка regex для «{name}»:\n{e}")
                    return
                if mr:
                    try:
                        re.compile(mr, re.I)
                    except re.error as e:
                        messagebox.showerror("КМ / КМД", f"Ошибка regex марки для «{name}»:\n{e}")
                        return
                new_groups.append({"name": name, "file_re": fr, "mark_re": mr})
            if not new_groups:
                messagebox.showwarning("КМ / КМД", "Не удалось собрать ни одной зоны из шаблона.")
                return
            if self._axis_groups and mode.get() == "replace":
                if not messagebox.askyesno("КМ / КМД", "Удалить текущие зоны и вставить выбранный шаблон?"):
                    return
                self._axis_groups = list(new_groups)
            else:
                self._axis_groups.extend(new_groups)
            self._refresh_axes_grp_list()
            self._refresh_axes_summary()
            top.destroy()
            messagebox.showinfo(
                "КМ / КМД",
                "Правила зон обновлены. Чтобы сохранить общий шаблон для всех проектов: меню «Ещё» → «Зоны: сохранить общий шаблон…».",
            )

        bf = tk.Frame(top)
        bf.grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(bf, text="Применить", command=apply, style="Accent.TButton").pack(side=tk.LEFT, padx=6)
        ttk.Button(bf, text="Отмена", command=top.destroy, style="Secondary.TButton").pack(side=tk.LEFT)

    def _refresh_axes_grp_list(self) -> None:
        tree = self._axes_grp_tree
        if tree is None:
            return
        for i in tree.get_children():
            tree.delete(i)
        for i, g in enumerate(self._axis_groups):
            tree.insert(
                "",
                tk.END,
                iid=str(i),
                values=(g.get("name", ""), g.get("file_re", ""), g.get("mark_re", "")),
            )

    def _axis_group_add_dialog(self) -> None:
        self._axis_group_edit_dialog(None)

    def _axis_group_edit_selected(self) -> None:
        tree = self._axes_grp_tree
        if tree is None:
            return
        sel = tree.selection()
        if not sel:
            messagebox.showinfo("Оси и зоны", "Выберите строку в таблице правил.")
            return
        try:
            idx = int(sel[0])
        except ValueError:
            return
        self._axis_group_edit_dialog(idx)

    def _axis_group_delete_selected(self) -> None:
        tree = self._axes_grp_tree
        if tree is None:
            return
        sel = tree.selection()
        if not sel:
            messagebox.showinfo("Оси и зоны", "Выберите строку для удаления.")
            return
        try:
            idx = int(sel[0])
            if 0 <= idx < len(self._axis_groups):
                del self._axis_groups[idx]
        except ValueError:
            return
        self._refresh_axes_grp_list()
        self._refresh_axes_summary()

    def _axis_group_edit_dialog(self, index: int | None) -> None:
        is_new = index is None
        cur = {"name": "", "file_re": "", "mark_re": ""}
        if not is_new and index is not None and 0 <= index < len(self._axis_groups):
            cur = dict(self._axis_groups[index])
        top = tk.Toplevel(self._root)
        top.title("Зона / ось" if is_new else "Изменить зону")
        top.transient(self._root)
        top.grab_set()
        pv = {"padx": 8, "pady": 4}
        v_name = tk.StringVar(value=cur.get("name", ""))
        v_fr = tk.StringVar(value=cur.get("file_re", ""))
        v_mr = tk.StringVar(value=cur.get("mark_re", ""))
        tk.Label(top, text="Название (для отчёта):").grid(row=0, column=0, sticky=tk.W, **pv)
        tk.Entry(top, textvariable=v_name, width=48).grid(row=0, column=1, sticky=tk.EW, **pv)
        tk.Label(top, text="Regex по имени файла:").grid(row=1, column=0, sticky=tk.W, **pv)
        tk.Entry(top, textvariable=v_fr, width=48).grid(row=1, column=1, sticky=tk.EW, **pv)
        tk.Label(top, text="Regex по марке (пусто = все строки):").grid(row=2, column=0, sticky=tk.W, **pv)
        tk.Entry(top, textvariable=v_mr, width=48).grid(row=2, column=1, sticky=tk.EW, **pv)
        top.columnconfigure(1, weight=1)
        hint = tk.Label(
            top,
            text="Пример для «оси 2» в имени: .*[_\\-]2([._\\-]|\\.pdf$|$)",
            fg=_UI_TEXT_DIM,
            font=("Segoe UI", 8),
            justify=tk.LEFT,
        )
        hint.grid(row=3, column=0, columnspan=2, sticky=tk.W, padx=8, pady=4)

        def ok() -> None:
            name = (v_name.get() or "").strip() or "Зона"
            fr = (v_fr.get() or "").strip()
            if not fr:
                messagebox.showwarning("Оси и зоны", "Укажите регулярное выражение для имени файла.")
                return
            try:
                re.compile(fr, re.I)
            except re.error as e:
                messagebox.showerror("Оси и зоны", f"Ошибка в regex файла:\n{e}")
                return
            mr = (v_mr.get() or "").strip()
            if mr:
                try:
                    re.compile(mr, re.I)
                except re.error as e:
                    messagebox.showerror("Оси и зоны", f"Ошибка в regex марки:\n{e}")
                    return
            row = {"name": name, "file_re": fr, "mark_re": mr}
            if is_new:
                self._axis_groups.append(row)
            else:
                self._axis_groups[int(index)] = row  # type: ignore[arg-type]
            self._refresh_axes_grp_list()
            self._refresh_axes_summary()
            top.destroy()

        bf = tk.Frame(top)
        bf.grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(bf, text="OK", command=ok, style="Accent.TButton").pack(side=tk.LEFT, padx=6)
        ttk.Button(bf, text="Отмена", command=top.destroy, style="Secondary.TButton").pack(side=tk.LEFT)

    def _axis_preset_123(self) -> None:
        if self._axis_groups and not messagebox.askyesno(
            "Пресет",
            "Очистить текущие правила и вставить три зоны «Ось 1», «Ось 2», «Ось 3»?",
        ):
            return
        self._axis_groups = []
        for n in ("1", "2", "3"):
            fr = rf".*[_\-]{n}([._\-]|\.pdf$|$)"
            self._axis_groups.append({"name": f"Ось {n}", "file_re": fr, "mark_re": ""})
        self._refresh_axes_grp_list()
        self._refresh_axes_summary()
        messagebox.showinfo(
            "Пресет",
            "Добавлены три группы. При необходимости отредактируйте regex под ваши имена файлов "
            "и нажмите «Сохранить правила».",
        )

    def _compute_axis_breakdown(self) -> tuple[list[dict[str, Any]], float]:
        """Возвращает (строки по зонам, суммарно м² по всем учтённым строкам)."""
        rows_out: list[dict[str, Any]] = []
        tot_check = 0.0
        UNMATCHED = "Вне заданных зон"
        groups = self._axis_groups

        def bucket(name: str) -> dict[str, Any]:
            return {"name": name, "files": set(), "rows": 0, "m2": 0.0, "kg": 0.0}

        buckets: dict[str, dict[str, Any]] = {}

        for path in self._pdf_paths:
            base = os.path.basename(path)
            gi = -1
            for i, g in enumerate(groups):
                fr = (g.get("file_re") or "").strip()
                if not fr:
                    continue
                try:
                    if re.search(fr, base, re.I):
                        gi = i
                        break
                except re.error:
                    continue
            gname = groups[gi]["name"] if 0 <= gi < len(groups) else UNMATCHED
            if gname not in buckets:
                buckets[gname] = bucket(gname)
            b = buckets[gname]
            b["files"].add(base)

            res = self._results.get(path)
            if not res:
                continue
            mk_rx = ""
            if 0 <= gi < len(groups):
                mk_rx = (groups[gi].get("mark_re") or "").strip()
            mark_pat: re.Pattern[str] | None = None
            if mk_rx:
                try:
                    mark_pat = re.compile(mk_rx, re.I)
                except re.error:
                    mark_pat = None

            if res.strategy == "metal_catalog":
                mlns = getattr(res, "metal_lines", None) or []
                for ml in mlns:
                    if mark_pat:
                        am = str(ml.get("assembly_mark") or "")
                        if not mark_pat.search(am):
                            continue
                    b["rows"] += 1
                    am2 = float(ml.get("area_m2") or 0)
                    b["m2"] += am2
                    tot_check += am2
                    mtt = ml.get("mass_kg_total")
                    if mtt is not None:
                        b["kg"] += float(mtt)
            else:
                for pr in getattr(res, "per_page", ()) or ():
                    b["rows"] += 1
                    am2 = float(pr.area_m2)
                    b["m2"] += am2
                    tot_check += am2

        order_names: list[str] = [g["name"] for g in groups] + ([UNMATCHED] if UNMATCHED in buckets else [])
        seen = set()
        for nm in order_names:
            if nm in buckets and nm not in seen:
                seen.add(nm)
                d = buckets[nm]
                rows_out.append(
                    {
                        "name": nm,
                        "nfc": len(d["files"]),
                        "nrow": int(d["rows"]),
                        "m2": float(d["m2"]),
                        "kg": float(d["kg"]),
                    }
                )
        for nm, d in buckets.items():
            if nm not in seen:
                rows_out.append(
                    {
                        "name": nm,
                        "nfc": len(d["files"]),
                        "nrow": int(d["rows"]),
                        "m2": float(d["m2"]),
                        "kg": float(d["kg"]),
                    }
                )
        return rows_out, tot_check

    def _refresh_axes_summary(self) -> None:
        tree = self._axes_sum_tree
        lbl = self._lbl_axes_hint
        if tree is None:
            return
        for i in tree.get_children():
            tree.delete(i)
        if not self._axis_groups:
            tree.insert("", tk.END, values=("—", "—", "—", "—", "—"))
            if lbl is not None:
                lbl.config(
                    text="Добавьте хотя бы одну зону или нажмите «Пресет: Оси 1–3», затем выполните «Анализ»."
                )
            return
        rows, tot = self._compute_axis_breakdown()
        if not self._pdf_paths:
            tree.insert("", tk.END, values=("—", "—", "—", "—", "—"))
            if lbl is not None:
                lbl.config(text="Загрузите PDF и выполните «Анализ» на вкладке «Расчёт АКЗ».")
            return
        if not rows:
            tree.insert("", tk.END, values=("—", "—", "—", "—", "—"))
            if lbl is not None:
                lbl.config(text="Нет данных для свода.")
            return
        gtot = 0.0
        for r in rows:
            tree.insert(
                "",
                tk.END,
                values=(
                    r["name"],
                    str(r["nfc"]),
                    str(r["nrow"]),
                    f'{r["m2"]:.4f}',
                    f'{r["kg"]:.2f}' if r["kg"] > 1e-9 else "—",
                ),
            )
            gtot += r["m2"]
        if lbl is not None:
            grand = 0.0
            for path in self._pdf_paths:
                res = self._results.get(path)
                if res and res.strategy == "metal_catalog":
                    mlns = getattr(res, "metal_lines", None) or []
                    grand += sum(float(x.get("area_m2") or 0) for x in mlns)
            if grand > 1e-9:
                lbl.config(
                    text=(
                        f"Σ м² по зонам (с учётом фильтра марки, если задан): {gtot:.4f} м². "
                        f"Для сравнения: Σ м² по всем строкам ведомости (вкладка «Расчёт»): {grand:.4f} м²."
                    )
                )
            else:
                lbl.config(text=f"Σ м² по зонам: {gtot:.4f} м².")

    def _export_axes_csv(self) -> None:
        if not self._axis_groups:
            messagebox.showinfo("Оси и зоны", "Нет правил групп — нечего экспортировать.")
            return
        rows, _ = self._compute_axis_breakdown()
        initial = self._suggested_export_basename() + "_оси_зоны.csv"
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=initial,
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(["Зона", "Файлов", "Строк_ведомости", "S_м2_АКЗ", "Масса_кг"])
                for r in rows:
                    w.writerow([r["name"], r["nfc"], r["nrow"], f'{r["m2"]:.6f}'.replace(".", ","), f'{r["kg"]:.3f}'.replace(".", ",")])
        except Exception as e:
            messagebox.showerror("Оси и зоны", str(e))
            return
        messagebox.showinfo("Оси и зоны", f"Свод сохранён:\n{path}")

    def _export_title_block(self) -> dict[str, str]:
        o = (self._report_title_override.get() or "").strip()
        block = {"title": "", "code": "", "organization": "", "source": ""}
        for pdf_path in self._pdf_paths:
            res = self._results.get(pdf_path)
            if not res:
                continue
            rh = getattr(res, "report_header", None) or {}
            if rh.get("title"):
                block["title"] = str(rh.get("title", ""))
                block["code"] = str(rh.get("code", ""))
                block["organization"] = str(rh.get("organization", ""))
                block["source"] = str(rh.get("source", ""))
                break
        if o:
            block["title"] = o
            block["source"] = "вручную"
        elif not block["title"] and self._pdf_paths:
            block["title"] = os.path.splitext(os.path.basename(self._pdf_paths[0]))[0]
            block["source"] = "имя_файла"
        return block

    def _suggested_export_basename(self) -> str:
        """Базовое имя файла экспорта: из поля «Название для Excel-отчёта», иначе от первого PDF."""
        o = (self._report_title_override.get() or "").strip()
        if o:
            for c in '\\/:*?"<>|':
                o = o.replace(c, "_")
            o = o.strip(" .")
            if o:
                return o
        first = next(iter(self._pdf_paths)) if self._pdf_paths else ""
        if first:
            return os.path.splitext(os.path.basename(first))[0]
        return "АКЗ"

    def _fill_pto_akz_sheet(
        self,
        ws,
        *,
        grand_total_m2: float,
        title_block: dict[str, str],
        cat_disp: str,
        mode_txt: str,
        border: Border,
        accent_fill: PatternFill,
        head_fill: PatternFill,
        hdr_font: Font,
        meta_lbl_font: Font,
        meta_val_font: Font,
        small_font: Font,
        wrap: Alignment,
    ) -> None:
        """Лист «АКЗ_ПТО»: сводка для отдела ПТО, площади, качество разбора, шаблон расчёта ЛКМ."""
        last_col_l = "H"
        r = 1
        ws.merge_cells(f"A{r}:{last_col_l}{r}")
        t = ws.cell(row=r, column=1, value="АКЗ и ЛКМ — справка для отдела ПТО")
        t.font = Font(name="Calibri", size=16, bold=True, color="FFFFFFFF")
        t.fill = accent_fill
        t.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[r].height = 30
        r += 1
        ws.merge_cells(f"A{r}:{last_col_l}{r}")
        sub = ws.cell(
            row=r,
            column=1,
            value=(
                "Ниже приведены расчётные площади антикоррозионной защиты (АКЗ) по ведомости и каталогу м²/п.м "
                "или по геометрии листа PDF. Нормы расхода лакокрасочных материалов (ЛКМ), полная система "
                "покрытия (грунт / промежуточные / финиш) и объёмы закупки определяются по ТУ, ТД и локальным "
                "нормам — блок в конце листа служит только шаблоном для переноса этих норм."
            ),
        )
        sub.font = small_font
        sub.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        ws.row_dimensions[r].height = 52
        r += 2

        meta_pairs = [
            ("Объект / проект", title_block.get("title") or "—"),
            ("Шифр / обозначение", title_block.get("code") or "—"),
            ("Организация (штамп PDF)", title_block.get("organization") or "—"),
            ("Дата отчёта", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Режим расчёта площади", mode_txt),
            ("Каталог м²/п.м (профили)", cat_disp),
            ("Допуск сверки: сумма строк vs явная площадь в тексте PDF", f"{_VALIDATION_TOLERANCE_STRICT_PCT} %"),
            ("Версия утилиты", _APP_VERSION),
        ]
        for lab, val in meta_pairs:
            ws.cell(row=r, column=1, value=lab).font = meta_lbl_font
            ws.cell(row=r, column=1).border = border
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            c = ws.cell(row=r, column=2, value=val)
            c.font = meta_val_font
            c.border = border
            c.alignment = wrap
            ws.merge_cells(f"B{r}:{last_col_l}{r}")
            r += 1

        r += 1
        ws.merge_cells(f"A{r}:{last_col_l}{r}")
        sec = ws.cell(row=r, column=1, value="Ключевые показатели по проекту")
        sec.font = hdr_font
        sec.fill = head_fill
        sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sec.border = border
        ws.row_dimensions[r].height = 22
        r += 1

        n_pdf = len(self._pdf_paths)
        n_pos = 0
        sum_mass = 0.0
        any_metal = False
        for pdf_path in self._pdf_paths:
            res = self._results.get(pdf_path)
            if not res:
                continue
            if res.strategy == "metal_catalog":
                mlns = getattr(res, "metal_lines", None) or []
                if mlns:
                    any_metal = True
                    n_pos += len(mlns)
                    for ml in mlns:
                        mt = ml.get("mass_kg_total")
                        if mt is not None:
                            sum_mass += float(mt)

        ship_note_parts: list[str] = []
        for pdf_path in self._pdf_paths:
            res = self._results.get(pdf_path)
            if not res or res.strategy != "metal_catalog":
                continue
            sqm = getattr(res, "shipment_qty_by_mark", None) or {}
            if sqm:
                bn = os.path.basename(pdf_path)
                ship_note_parts.append(
                    bn + ": " + "; ".join(f"{mk} — {n} шт." for mk, n in sorted(sqm.items()))
                )
        ship_note = " | ".join(ship_note_parts) if ship_note_parts else "—"

        area_value_row = 0
        kv_rows: list[tuple[str, Any]] = [
            ("Σ площадь АКЗ (все файлы), м²", round(grand_total_m2, 6)),
            ("Количество PDF в расчёте", n_pdf),
            (
                "Позиций ведомости (всего)",
                n_pos if any_metal else "— (нет режима ведомости)",
            ),
            (
                "Отправочные комплекты по маркам (из PDF)",
                ship_note,
            ),
            (
                "Σ масса по ведомости, кг",
                round(sum_mass, 3) if any_metal and sum_mass > 1e-9 else "—",
            ),
        ]
        for lab, val in kv_rows:
            ws.cell(row=r, column=1, value=lab).font = meta_lbl_font
            ws.cell(row=r, column=1).border = border
            b = ws.cell(row=r, column=2, value=val)
            b.font = meta_val_font
            b.border = border
            if isinstance(val, float):
                b.number_format = "0.000000"
            elif isinstance(val, int):
                b.number_format = "0"
            if "Σ площадь АКЗ" in lab:
                area_value_row = r
            ws.merge_cells(f"B{r}:{last_col_l}{r}")
            r += 1

        r += 1
        ws.merge_cells(f"A{r}:{last_col_l}{r}")
        h2 = ws.cell(row=r, column=1, value="По файлам: площадь АКЗ и контроль разбора")
        h2.font = hdr_font
        h2.fill = head_fill
        h2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        h2.border = border
        r += 1

        hdrs = [
            "Файл PDF",
            "S АКЗ, м²",
            "Поз. ведомости",
            "Явная S в тексте PDF, м²",
            "Примечание по площади в PDF",
            "Индекс качества",
            "Ошибок",
            "Предупр.",
        ]
        hdr_row_files = r
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = hdr_font
            c.fill = head_fill
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 38
        r += 1

        for pdf_path in self._pdf_paths:
            res = self._results.get(pdf_path)
            if not res:
                continue
            bn = os.path.basename(pdf_path)
            mlns = getattr(res, "metal_lines", None) or []
            if res.strategy == "metal_catalog" and mlns:
                file_tot = sum(float(x.get("area_m2") or 0) for x in mlns)
                npr: Any = len(mlns)
            else:
                file_tot = sum(float(pr.area_m2) for pr in res.per_page)
                npr = "—"

            expl = getattr(res, "explicit_paint_area_m2", None)
            expl_amb = bool(getattr(res, "explicit_paint_area_ambiguous", False))
            if expl_amb:
                expl_note = "Несколько значений в тексте — проверьте вручную"
            elif expl is not None:
                expl_note = "Найдена в тексте; сверка с суммой строк — см. «Сводка»"
            else:
                expl_note = "Не выделена из текста / режим без явной площади"

            qm = getattr(res, "quality_metrics", None) or {}
            qscore = qm.get("confidence_score")
            if isinstance(qscore, (int, float)):
                qdisp: Any = int(qscore)
            else:
                qdisp = "—"
            if res.strategy == "metal_catalog" and qm:
                err_c = int(qm.get("error_count") or 0)
                warn_c = int(qm.get("warning_count") or 0)
            else:
                err_c = sum(
                    1 for v in getattr(res, "metal_validation", []) or [] if v.get("severity") == "error"
                )
                warn_c = sum(
                    1 for v in getattr(res, "metal_validation", []) or [] if v.get("severity") == "warning"
                )

            vals_row: list[Any] = [
                bn,
                round(file_tot, 6),
                npr,
                round(float(expl), 6) if expl is not None else "—",
                expl_note,
                qdisp,
                err_c,
                warn_c,
            ]
            for col, val in enumerate(vals_row, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left" if col in (1, 5) else "right",
                    vertical="top",
                    wrap_text=True,
                )
                if col in (2, 4) and isinstance(val, (int, float)):
                    cell.number_format = "0.000000"
            r += 1

        r += 1
        ws.merge_cells(f"A{r}:{last_col_l}{r}")
        agg_title = ws.cell(row=r, column=1, value="Сводка покрытия каталогом и чувствительность (режим ведомости)")
        agg_title.font = hdr_font
        agg_title.fill = head_fill
        agg_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        agg_title.border = border
        r += 1

        if any_metal:
            cov2: list[float] = []
            covk: list[float] = []
            sens: list[float] = []
            for pdf_path in self._pdf_paths:
                res = self._results.get(pdf_path)
                if not res or res.strategy != "metal_catalog":
                    continue
                qm2 = getattr(res, "quality_metrics", None) or {}
                if isinstance(qm2.get("coverage_m2_pct"), (int, float)):
                    cov2.append(float(qm2["coverage_m2_pct"]))
                if isinstance(qm2.get("coverage_kg_pct"), (int, float)):
                    covk.append(float(qm2["coverage_kg_pct"]))
                if isinstance(qm2.get("sensitivity_delta_m2_if_catalog_plus_1pct"), (int, float)):
                    sens.append(float(qm2["sensitivity_delta_m2_if_catalog_plus_1pct"]))
            row_agg = (
                f"Средняя доля позиций с м²/п.м в каталоге: {sum(cov2) / len(cov2):.1f}% ({len(cov2)} файл.)"
                if cov2
                else "—"
            )
            row_kg = f"Средняя доля позиций с кг/п.м: {sum(covk) / len(covk):.1f}%" if covk else "—"
            row_sens = (
                f"Изменение Σ м² при +1% к коэфф. м²/п.м (ориентир): до ~{sum(sens):.4f} м² по файлам с метрикой"
                if sens
                else "—"
            )
            for txt in (row_agg, row_kg, row_sens):
                ws.merge_cells(f"A{r}:{last_col_l}{r}")
                c = ws.cell(row=r, column=1, value=txt)
                c.font = meta_val_font
                c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
                c.border = border
                r += 1
        else:
            ws.merge_cells(f"A{r}:{last_col_l}{r}")
            c = ws.cell(row=r, column=1, value="Показатели каталога актуальны для режима «металл» (ведомость + CSV).")
            c.font = meta_val_font
            c.alignment = wrap
            c.border = border
            r += 1

        r += 1
        ws.merge_cells(f"A{r}:{last_col_l}{r}")
        cov_title2 = ws.cell(row=r, column=1, value="Шаблон: расход ЛКМ (заполнить нормами из ТУ / калькуляции)")
        cov_title2.font = hdr_font
        cov_title2.fill = head_fill
        cov_title2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cov_title2.border = border
        r += 1

        foot = (
            "Итоговая потребность в ЛКМ с учётом отходов, растворителей и фактической влажности поверхности "
            "оформляется отдельной ведомостью расхода. Данный блок — ориентир по массе готовой смеси."
        )
        ws.merge_cells(f"A{r}:{last_col_l}{r}")
        nf = ws.cell(row=r, column=1, value=foot)
        nf.font = small_font
        nf.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.row_dimensions[r].height = 32
        r += 1

        r_area = r
        ws.cell(row=r, column=1, value="Площадь АКЗ для расчёта, м² (связь с итогом выше)").font = meta_lbl_font
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=f"=B{area_value_row}").font = meta_val_font
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=2).number_format = "0.000000"
        ws.merge_cells(f"B{r}:{last_col_l}{r}")
        r += 1
        r_lkm_sys = r
        ws.cell(row=r, column=1, value="Система ЛКМ (описание, ГОСТ, ТУ, производитель)").font = meta_lbl_font
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=None).font = meta_val_font
        ws.cell(row=r, column=2).border = border
        ws.merge_cells(f"B{r}:{last_col_l}{r}")
        r += 1
        r_norm = r
        ws.cell(row=r, column=1, value="Норма расхода готовой смеси, кг/м² (один слой, из ТУ)").font = meta_lbl_font
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=None).font = meta_val_font
        ws.cell(row=r, column=2).border = border
        ws.merge_cells(f"B{r}:{last_col_l}{r}")
        r += 1
        r_layers = r
        ws.cell(row=r, column=1, value="Число слоёв по технологии").font = meta_lbl_font
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=None).font = meta_val_font
        ws.cell(row=r, column=2).border = border
        ws.merge_cells(f"B{r}:{last_col_l}{r}")
        r += 1
        r_kg = r
        ws.cell(row=r, column=1, value="Ориентировочный расход готовой краски, кг").font = meta_lbl_font
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=f"=B{r_area}*B{r_norm}*B{r_layers}").font = meta_val_font
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=2).number_format = "0.00"
        ws.merge_cells(f"B{r}:{last_col_l}{r}")
        r += 1
        r_pct = r
        ws.cell(row=r, column=1, value="Запас (непредвиденные работы), % — при необходимости").font = meta_lbl_font
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=None).font = meta_val_font
        ws.cell(row=r, column=2).border = border
        ws.merge_cells(f"B{r}:{last_col_l}{r}")
        r += 1
        r_tot = r
        ws.cell(row=r, column=1, value="Итого ЛКМ с учётом запаса, кг").font = meta_lbl_font
        ws.cell(row=r, column=1).border = border
        ws.cell(
            row=r,
            column=2,
            value=f"=IF(ISBLANK(B{r_pct}),B{r_kg},B{r_kg}*(1+B{r_pct}/100))",
        ).font = meta_val_font
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=2).number_format = "0.00"
        ws.merge_cells(f"B{r}:{last_col_l}{r}")
        r += 1

        ws.column_dimensions["A"].width = 44
        for col in range(2, 9):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.freeze_panes = f"A{hdr_row_files + 1}"

    def _export_xlsx(self, save_path: str | None = None, *, notify: bool = True) -> tuple[bool, str]:
        if not _OPENPYXL_OK or Workbook is None:
            msg = (
                "Для отчёта Excel (.xlsx) установите пакет openpyxl:\n\npip install openpyxl\n\n"
                "См. также УСТАНОВКА.txt"
            )
            if notify:
                messagebox.showinfo(_APP_DISPLAY_NAME, msg)
            return False, msg
        if not self._results:
            msg = "Сначала выполните «Анализ»."
            if notify:
                messagebox.showinfo(_APP_DISPLAY_NAME, msg)
            return False, msg
        path = (save_path or "").strip()
        if not path:
            initial = self._suggested_export_basename() + "_АКЗ.xlsx"
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Книга Excel", "*.xlsx")],
                initialfile=initial,
                title="Сохранить результат",
            )
            if not path:
                return False, ""

        parent_dir = os.path.dirname(os.path.abspath(path))
        if parent_dir:
            try:
                os.makedirs(parent_dir, exist_ok=True)
            except OSError:
                pass

        hdr_font = Font(name="Calibri", size=11, bold=True)
        title_font_big = Font(name="Calibri", size=16, bold=True, color="FFFFFFFF")
        meta_val_font = Font(name="Calibri", size=11)
        meta_lbl_font = Font(name="Calibri", size=11, bold=True, color="FF334155")
        small_font = Font(name="Calibri", size=10)
        head_fill = PatternFill(start_color="FFE8EEF4", end_color="FFE8EEF4", fill_type="solid")
        accent_fill = PatternFill(start_color="FF1565A8", end_color="FF1565A8", fill_type="solid")
        fill_issue = PatternFill(start_color="FFFFE0E0", end_color="FFFFE0E0", fill_type="solid")
        fill_warn = PatternFill(start_color="FFFFF9C4", end_color="FFFFF9C4", fill_type="solid")
        thin = Side(style="thin", color="FFB8C5D8")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        wrap = Alignment(wrap_text=True, vertical="top")
        wrap_vc = Alignment(wrap_text=True, vertical="center")
        title_block = self._export_title_block()
        cat_disp = _catalog_display_for_report(self._catalog_path)

        wb = Workbook()
        assert wb is not None
        ws_sum = wb.active
        ws_sum.title = "Сводка"

        ncols = 7
        last_let = get_column_letter(ncols)
        ws_sum.merge_cells(f"A1:{last_let}1")
        c1 = ws_sum.cell(row=1, column=1, value=f"{_APP_DISPLAY_NAME} — сводный отчёт")
        c1.font = title_font_big
        c1.fill = accent_fill
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_sum.row_dimensions[1].height = 32

        mode_txt = "Металл по ведомости (каталог м²/п.м)" if self._goal.get() == "metal" else "Площадь листа PDF"
        meta_rows = [
            ("Утилита", _APP_DISPLAY_NAME),
            ("Объект / проект", title_block["title"]),
            ("Шифр / обозначение", title_block["code"] if title_block.get("code") else "—"),
            ("Организация (штамп)", title_block["organization"] if title_block.get("organization") else "—"),
            ("Дата отчёта", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Версия", _APP_VERSION),
            ("Режим расчёта", mode_txt),
            ("Каталог м²/п.м", cat_disp),
            ("Название объекта: источник", title_block["source"] or "—"),
        ]
        r = 2
        for lab, val in meta_rows:
            a = ws_sum.cell(row=r, column=1, value=lab)
            a.font = meta_lbl_font
            a.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            a.border = border
            b = ws_sum.cell(row=r, column=2, value=val)
            b.font = meta_val_font
            b.alignment = wrap
            b.border = border
            ws_sum.merge_cells(f"B{r}:{last_let}{r}")
            r += 1
        meta_end = r - 1

        if len(self._pdf_paths) > 1:
            a = ws_sum.cell(row=r, column=1, value="Файлы в проекте")
            a.font = meta_lbl_font
            a.alignment = Alignment(horizontal="right", vertical="top")
            a.border = border
            b = ws_sum.cell(row=r, column=2, value="; ".join(os.path.basename(p) for p in self._pdf_paths))
            b.font = small_font
            b.alignment = wrap
            b.border = border
            ws_sum.merge_cells(f"B{r}:{last_let}{r}")
            r += 1
            meta_end = r - 1

        r += 1
        hdr_row = r
        sum_headers = ["Файл", "Лист / строк ведомости", "Площадь_м2", "Примечание", "Режим", "Уверенность", "Обоснование"]
        for col, h in enumerate(sum_headers, 1):
            cell = ws_sum.cell(row=hdr_row, column=col, value=h)
            cell.font = hdr_font
            cell.fill = head_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_sum.row_dimensions[hdr_row].height = 28

        tot = 0.0
        r = hdr_row + 1
        for pdf_path in self._pdf_paths:
            res = self._results.get(pdf_path)
            if not res:
                continue
            bn = os.path.basename(pdf_path)
            mlns = getattr(res, "metal_lines", []) or []
            if res.strategy == "metal_catalog" and mlns:
                file_tot = sum(float(x.get("area_m2") or 0) for x in mlns)
                tot += file_tot
                npos = len(mlns)
                note = (
                    f"Ведомость: {npos} поз. Подробная таблица — лист «Ведомость»; по маркам — «По маркам»."
                )
                row_vals: list[Any] = [
                    bn,
                    f"{npos} поз.",
                    round(file_tot, 6),
                    note,
                    res.strategy,
                    res.confidence,
                    _short_text(res.reason, 650),
                ]
                for col, val in enumerate(row_vals, 1):
                    cell = ws_sum.cell(row=r, column=col, value=val)
                    cell.border = border
                    cell.alignment = wrap
                    if col == 3:
                        cell.number_format = "0.000000"
                r += 1
            else:
                for pr in res.per_page:
                    tot += float(pr.area_m2)
                    row_vals = [
                        bn,
                        str(pr.page_index + 1),
                        round(float(pr.area_m2), 6),
                        _short_text(pr.detail, 650),
                        res.strategy,
                        res.confidence,
                        _short_text(res.reason, 650),
                    ]
                    for col, val in enumerate(row_vals, 1):
                        cell = ws_sum.cell(row=r, column=col, value=val)
                        cell.border = border
                        cell.alignment = wrap
                        if col == 3:
                            cell.number_format = "0.000000"
                    r += 1

        tot_row = r
        for col in range(1, ncols + 1):
            c = ws_sum.cell(row=tot_row, column=col)
            c.border = border
        ws_sum.cell(row=tot_row, column=1, value="ИТОГО, м²").font = hdr_font
        ws_sum.cell(row=tot_row, column=3, value=round(tot, 6)).font = hdr_font
        ws_sum.cell(row=tot_row, column=3).number_format = "0.000000"

        def _sev_label_sum(s: str) -> str:
            return {"error": "ошибка", "warning": "предупреждение", "info": "инфо"}.get(s, s or "—")

        vr = tot_row + 2
        sec_title = ws_sum.cell(row=vr, column=1, value="Контроль и валидация")
        sec_title.font = Font(name="Calibri", size=12, bold=True, color="FF0F172A")
        sec_title.fill = head_fill
        sec_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_sum.merge_cells(f"A{vr}:{last_let}{vr}")
        for c in range(1, ncols + 1):
            ws_sum.cell(row=vr, column=c).border = border
        vr += 1
        vh = vr
        for col, h in enumerate(["Уровень", "Файл", "Сообщение"], 1):
            cell = ws_sum.cell(row=vh, column=col, value=h)
            cell.font = hdr_font
            cell.fill = head_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_sum.merge_cells(f"C{vh}:{last_let}{vh}")
        vr = vh + 1
        for pdf_path in self._pdf_paths:
            res = self._results.get(pdf_path)
            if not res:
                continue
            bn = os.path.basename(pdf_path)
            qm = getattr(res, "quality_metrics", None) or {}
            sc = qm.get("confidence_score")
            if isinstance(sc, (int, float)):
                ws_sum.cell(row=vr, column=1, value="инфо")
                ws_sum.cell(row=vr, column=2, value=bn)
                ws_sum.cell(row=vr, column=3, value=f"Индекс качества: {int(sc)}/100; Σ м²: {qm.get('total_area_m2', '—')}")
                ws_sum.merge_cells(f"C{vr}:{last_let}{vr}")
                for c in range(1, 4):
                    ws_sum.cell(row=vr, column=c).border = border
                    ws_sum.cell(row=vr, column=c).alignment = wrap
                vr += 1
            for v in getattr(res, "metal_validation", []) or []:
                ws_sum.cell(row=vr, column=1, value=_sev_label_sum(str(v.get("severity", ""))))
                ws_sum.cell(row=vr, column=2, value=bn)
                ws_sum.cell(row=vr, column=3, value=str(v.get("message", ""))[:4000])
                ws_sum.merge_cells(f"C{vr}:{last_let}{vr}")
                for c in range(1, 4):
                    ws_sum.cell(row=vr, column=c).border = border
                    ws_sum.cell(row=vr, column=c).alignment = wrap
                vr += 1
        for v in self._cross_file_validation:
            ws_sum.cell(row=vr, column=1, value=_sev_label_sum(str(v.get("severity", ""))))
            ws_sum.cell(row=vr, column=2, value="(несколько PDF)")
            ws_sum.cell(row=vr, column=3, value=str(v.get("message", ""))[:4000])
            ws_sum.merge_cells(f"C{vr}:{last_let}{vr}")
            for c in range(1, 4):
                ws_sum.cell(row=vr, column=c).border = border
                ws_sum.cell(row=vr, column=c).alignment = wrap
            vr += 1
        if vr == vh + 1:
            ws_sum.cell(row=vr, column=1, value="—")
            ws_sum.cell(row=vr, column=2, value="")
            ws_sum.cell(row=vr, column=3, value="Нет записей проверок.")
            ws_sum.merge_cells(f"C{vr}:{last_let}{vr}")
            vr += 1

        ws_sum.freeze_panes = f"A{hdr_row + 1}"
        if r > hdr_row + 1:
            tbl = Table(displayName="TblSumOkr", ref=f"A{hdr_row}:{last_let}{r - 1}")
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws_sum.add_table(tbl)

        ws_sum.column_dimensions["A"].width = 28
        ws_sum.column_dimensions["B"].width = 16
        ws_sum.column_dimensions["C"].width = 14
        ws_sum.column_dimensions["D"].width = 42
        ws_sum.column_dimensions["E"].width = 16
        ws_sum.column_dimensions["F"].width = 14
        ws_sum.column_dimensions["G"].width = 48
        for rr in range(2, meta_end + 1):
            ws_sum.row_dimensions[rr].height = max(ws_sum.row_dimensions[rr].height or 0, 18)

        ws_pto = wb.create_sheet("АКЗ_ПТО", 1)
        self._fill_pto_akz_sheet(
            ws_pto,
            grand_total_m2=tot,
            title_block=title_block,
            cat_disp=cat_disp,
            mode_txt=mode_txt,
            border=border,
            accent_fill=accent_fill,
            head_fill=head_fill,
            hdr_font=hdr_font,
            meta_lbl_font=meta_lbl_font,
            meta_val_font=meta_val_font,
            small_font=small_font,
            wrap=wrap,
        )

        # --- Ведомость ---
        ws_m = wb.create_sheet("Ведомость", 2)
        ws_m.merge_cells("A1:M1")
        t = ws_m.cell(row=1, column=1, value=f"{_APP_DISPLAY_NAME} — ведомость (все файлы)")
        t.font = Font(name="Calibri", size=14, bold=True, color="FFFFFFFF")
        t.fill = accent_fill
        t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_m.row_dimensions[1].height = 28
        m_hdr_row = 3
        m_cols = [
            "Файл",
            "марка",
            "отпр_шт",
            "позиция",
            "профиль",
            "шт",
            "м2_за_1_шт",
            "кг_на_пм",
            "м2_1_компл",
            "м2_все_отпр",
            "м2_из_массы",
            "дельта_м2",
            "проверка_строки",
        ]

        for col, h in enumerate(m_cols, 1):
            cell = ws_m.cell(row=m_hdr_row, column=col, value=h)
            cell.font = hdr_font
            cell.fill = head_fill
            cell.border = border

        mr = m_hdr_row + 1
        has_metal = False
        for pdf_path in self._pdf_paths:
            res = self._results.get(pdf_path)
            if not res or res.strategy != "metal_catalog":
                continue
            mlns = getattr(res, "metal_lines", []) or []
            if not mlns:
                continue
            has_metal = True
            bn = os.path.basename(pdf_path)
            for ml in mlns:
                kgp = ml.get("kg_per_m")
                aim = ml.get("area_implied_m2")
                dlm = ml.get("delta_area_m2")
                row_status = str(ml.get("row_status") or "")
                area_one = _metal_line_area_m2_one_assembly(ml)
                area_full = float(ml.get("area_m2") or 0)
                area_pc = _metal_line_area_m2_per_piece(ml)
                vals: list[Any] = [
                    bn,
                    str(ml.get("assembly_mark") or ""),
                    int(ml.get("shipment_qty") or 1),
                    ml.get("position", ""),
                    _metal_line_profile_display_ui(ml),
                    ml.get("qty", ""),
                    area_pc,
                    float(kgp) if kgp is not None else None,
                    area_one,
                    area_full,
                    float(aim) if aim is not None else None,
                    float(dlm) if dlm is not None else None,
                    row_status,
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws_m.cell(row=mr, column=col, value=val)
                    cell.border = border
                    if col in (3,) and val is not None:
                        cell.number_format = "0"
                    if col in (7, 8, 9, 10, 11, 12) and val is not None:
                        cell.number_format = "0.000000"
                if row_status == "issue":
                    for col in range(1, 14):
                        ws_m.cell(row=mr, column=col).fill = fill_issue
                elif row_status == "warn":
                    for col in range(1, 14):
                        ws_m.cell(row=mr, column=col).fill = fill_warn
                mr += 1

        if not has_metal:
            ws_m.cell(row=m_hdr_row + 1, column=1, value="Нет строк ведомости (режим металла или пустой разбор).")

        ws_m.freeze_panes = "A4"
        last_m = mr - 1 if has_metal else m_hdr_row + 1
        if has_metal and last_m > m_hdr_row:
            tbl2 = Table(displayName="TblMetal", ref=f"A{m_hdr_row}:M{last_m}")
            tbl2.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            ws_m.add_table(tbl2)
        for col in range(1, 14):
            letter = get_column_letter(col)
            ws_m.column_dimensions[letter].width = min(28, max(8, len(m_cols[col - 1]) + 3))

        n_mk = 7
        mk_last = get_column_letter(n_mk)
        ws_mk = wb.create_sheet("По маркам", 3)
        ws_mk.merge_cells(f"A1:{mk_last}1")
        hmk = ws_mk.cell(row=1, column=1, value=f"{_APP_DISPLAY_NAME} — по маркам сборки")
        hmk.font = Font(name="Calibri", size=15, bold=True, color="FFFFFFFF")
        hmk.fill = accent_fill
        hmk.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_mk.row_dimensions[1].height = 30
        ws_mk.merge_cells(f"A2:{mk_last}2")
        sub_mk = ws_mk.cell(
            row=2,
            column=1,
            value=(
                "Файл → марка сборки → позиции ведомости. У марки — м² с учётом всех отпр. комплектов; у позиций — "
                "м² на 1 комплект и «м² за 1 шт.» (по колонке шт)."
            ),
        )
        sub_mk.font = Font(name="Calibri", size=10, color="FF475569")
        sub_mk.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws_mk.row_dimensions[2].height = 22

        mk_col_titles = ["Поз.", "Сечение", "Шт", "м² за 1 шт.", "кг/м", "м² (1 к-т)", "!"]
        hr_mk = 4
        for col, h in enumerate(mk_col_titles, 1):
            c = ws_mk.cell(row=hr_mk, column=col, value=h)
            c.font = hdr_font
            c.fill = head_fill
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_mk.freeze_panes = f"A{hr_mk + 1}"
        ws_mk.row_dimensions[hr_mk].height = 24

        file_banner_fill = PatternFill(start_color="FFE0F2FE", end_color="FFE0F2FE", fill_type="solid")
        mark_banner_fill = PatternFill(start_color="FFE2E8F0", end_color="FFE2E8F0", fill_type="solid")
        rmk = hr_mk + 1
        has_grouped = False

        for pdf_path in self._pdf_paths:
            res = self._results.get(pdf_path)
            if not res or res.strategy != "metal_catalog":
                continue
            mlns = getattr(res, "metal_lines", []) or []
            if not mlns:
                continue
            has_grouped = True
            bn = os.path.basename(pdf_path)
            pidx = (res.per_page[0].page_index + 1) if res.per_page else 1
            ws_mk.merge_cells(f"A{rmk}:{mk_last}{rmk}")
            cr = ws_mk.cell(row=rmk, column=1, value=f"▸  {bn}  ·  лист чертежа {pidx}")
            cr.font = Font(name="Calibri", size=12, bold=True, color="FF1E3A5F")
            cr.fill = file_banner_fill
            cr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cr.border = border
            rmk += 1

            by_mark: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for ml in mlns:
                mk = str(ml.get("assembly_mark") or "Без марки")
                by_mark[mk].append(ml)
            for mark in sorted(by_mark.keys(), key=lambda x: (len(x), x)):
                lines = by_mark[mark]
                subtot = sum(float(x.get("area_m2") or 0) for x in lines)
                subtot_one = sum(_metal_line_area_m2_one_assembly(x) for x in lines)
                seen_sig = set()
                names = []
                for x in sorted(lines, key=_metal_line_position_sort):
                    disp_m = _metal_line_profile_display_ui(x)
                    if not disp_m:
                        continue
                    sig = _metal_line_profile_signature(x)
                    if not sig:
                        sig = disp_m.lower()
                    if sig in seen_sig:
                        continue
                    seen_sig.add(sig)
                    names.append(disp_m)
                comp = "; ".join(names[:10])
                if len(names) > 10:
                    comp += "; …"
                ws_mk.merge_cells(f"A{rmk}:{mk_last}{rmk}")
                sq0 = int(lines[0].get("shipment_qty") or 1) if lines else 1
                cm = ws_mk.cell(
                    row=rmk,
                    column=1,
                    value=(
                        f"Марка {mark}  ·  отпр. {sq0} шт.  —  всего {subtot:.4f} м² "
                        f"(~{subtot_one:.4f} на 1 компл.)   ·   {comp or '—'}"
                    ),
                )
                cm.font = Font(name="Calibri", size=11, bold=True, color="FF0F172A")
                cm.fill = mark_banner_fill
                cm.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
                cm.border = border
                rmk += 1

                for ml in lines:
                    kgp = ml.get("kg_per_m")
                    pr = _metal_line_profile_display_ui(ml)[:52]
                    if not pr:
                        pr = "—"
                    ar = _metal_line_area_m2_one_assembly(ml)
                    apc = _metal_line_area_m2_per_piece(ml)
                    row_status = str(ml.get("row_status") or "")
                    fl = "!" if row_status == "issue" else ("?" if row_status == "warn" else "")
                    vals_r = (
                        str(ml.get("position", "")),
                        pr,
                        ml.get("qty", ""),
                        apc,
                        float(kgp) if kgp is not None else None,
                        ar,
                        fl,
                    )
                    for col, val in enumerate(vals_r, 1):
                        cell = ws_mk.cell(row=rmk, column=col, value=val)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="right" if col not in (2, 7) else "left", vertical="center")
                        if col in (4, 5, 6) and val is not None and isinstance(val, (int, float)):
                            cell.number_format = "0.000000"
                        elif col == 3 and val is not None and isinstance(val, (int, float)):
                            cell.number_format = "0"
                    if row_status == "issue":
                        for col in range(1, n_mk + 1):
                            ws_mk.cell(row=rmk, column=col).fill = fill_issue
                    elif row_status == "warn":
                        for col in range(1, n_mk + 1):
                            ws_mk.cell(row=rmk, column=col).fill = fill_warn
                    rmk += 1
            rmk += 1

        if not has_grouped:
            ws_mk.merge_cells(f"A{rmk}:{mk_last}{rmk}")
            ws_mk.cell(row=rmk, column=1, value="Нет данных: режим «металл» и непустая ведомость после «Анализ».").font = meta_val_font

        ws_mk.column_dimensions["A"].width = 8
        ws_mk.column_dimensions["B"].width = 36
        ws_mk.column_dimensions["C"].width = 8
        ws_mk.column_dimensions["D"].width = 10
        ws_mk.column_dimensions["E"].width = 12
        ws_mk.column_dimensions["F"].width = 12
        ws_mk.column_dimensions["G"].width = 5

        try:
            wb.properties.title = _APP_DISPLAY_NAME
            wb.properties.subject = "Площадь АКЗ — ПТО, сводка, ведомость, марки"
        except Exception:
            pass
        for nm, rgb in (
            ("Сводка", "FF1565A8"),
            ("АКЗ_ПТО", "FF0F766E"),
            ("Ведомость", "FF0D9488"),
            ("По маркам", "FF7C3AED"),
        ):
            if nm in wb.sheetnames:
                try:
                    wb[nm].sheet_properties.tabColor = rgb
                except Exception:
                    pass

        try:
            wb.save(path)
            ok_msg = f"Результат сохранён в Excel:\n{path}"
            if notify:
                messagebox.showinfo(_APP_DISPLAY_NAME, ok_msg)
            self._rep(False)
            return True, ok_msg
        except Exception as e:
            self._rep(True)
            if notify:
                messagebox.showerror(_APP_DISPLAY_NAME, str(e))
            return False, str(e)

    def _export_csv(self) -> None:
        if not self._results:
            messagebox.showinfo("CSV", "Сначала выполните «Анализ».")
            return
        initial = self._suggested_export_basename() + "_окраска.csv"
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=initial,
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=";")
                tb = self._export_title_block()
                w.writerow(["Отчёт", _APP_DISPLAY_NAME, f"версия {_APP_VERSION}", "", "", "", ""])
                w.writerow(
                    [
                        "Объект",
                        tb["title"],
                        "Шифр",
                        tb.get("code") or "",
                        "Источник названия",
                        tb.get("source") or "",
                        "Дата",
                        datetime.now().strftime("%Y-%m-%d %H:%M"),
                    ]
                )
                w.writerow([])
                w.writerow(["Файл", "Лист", "Площадь_м2", "Детально", "Режим", "Уверенность", "Обоснование_режима"])
                tot = 0.0
                for pdf_path in self._pdf_paths:
                    res = self._results.get(pdf_path)
                    if not res:
                        continue
                    bn = os.path.basename(pdf_path)
                    for pr in res.per_page:
                        tot += pr.area_m2
                        w.writerow(
                            [
                                bn,
                                pr.page_index + 1,
                                f"{pr.area_m2:.6f}".replace(".", ","),
                                pr.detail.replace(";", ","),
                                res.strategy,
                                res.confidence,
                                res.reason.replace(";", ","),
                            ]
                        )
                    mlns = getattr(res, "metal_lines", []) or []
                    if res.strategy == "metal_catalog" and mlns:
                        w.writerow([])
                        w.writerow(["Секция", "Детализация ведомости", os.path.basename(pdf_path)] + [""] * 10)
                        w.writerow(
                            [
                                "Файл",
                                "марка",
                                "отпр_шт",
                                "позиция",
                                "профиль",
                                "шт",
                                "м2_за_1_шт",
                                "кг_на_пм",
                                "м2_1_компл",
                                "м2_все_отпр",
                                "м2_из_массы",
                                "дельта_м2",
                                "проверка_строки",
                            ]
                        )
                        for ml in mlns:
                            kgp = ml.get("kg_per_m")
                            aim = ml.get("area_implied_m2")
                            dlm = ml.get("delta_area_m2")
                            a1 = _metal_line_area_m2_one_assembly(ml)
                            af = float(ml.get("area_m2") or 0)
                            apc = _metal_line_area_m2_per_piece(ml)
                            w.writerow(
                                [
                                    bn,
                                    str(ml.get("assembly_mark") or "").replace(";", ","),
                                    int(ml.get("shipment_qty") or 1),
                                    ml.get("position", ""),
                                    _metal_line_profile_display_ui(ml).replace(";", ","),
                                    ml.get("qty", ""),
                                    f'{apc:.6f}'.replace(".", ","),
                                    (
                                        f'{float(kgp):.6f}'.replace(".", ",")
                                        if kgp is not None
                                        else ""
                                    ),
                                    f'{a1:.6f}'.replace(".", ","),
                                    f'{af:.6f}'.replace(".", ","),
                                    (f'{float(aim):.6f}'.replace(".", ",") if aim is not None else ""),
                                    (f'{float(dlm):.6f}'.replace(".", ",") if dlm is not None else ""),
                                    str(ml.get("row_status") or ""),
                                ]
                            )
                    w.writerow([])
                w.writerow(["ИТОГО_м2", f"{tot:.6f}".replace(".", ",")])
            messagebox.showinfo("CSV", f"Сохранено:\n{path}")
            self._rep(False)
        except Exception as e:
            self._rep(True)
            messagebox.showerror("CSV", str(e))

    def _project_dict(self) -> dict[str, Any]:
        results: dict[str, Any] = {}
        for p, r in self._results.items():
            ep = getattr(r, "explicit_paint_area_m2", None)
            rh = getattr(r, "report_header", None) or {}
            results[p] = {
                "strategy": r.strategy,
                "confidence": r.confidence,
                "reason": r.reason,
                "report_header": dict(rh),
                "per_page": [
                    {"page_index": pr.page_index, "area_m2": pr.area_m2, "detail": pr.detail} for pr in r.per_page
                ],
                "metal_lines": getattr(r, "metal_lines", []),
                "metal_validation": getattr(r, "metal_validation", []),
                "quality_metrics": getattr(r, "quality_metrics", {}),
                "explicit_paint_area_m2": ep,
                "explicit_paint_area_ambiguous": getattr(r, "explicit_paint_area_ambiguous", False),
                "shipment_qty_by_mark": dict(getattr(r, "shipment_qty_by_mark", None) or {}),
                "merged_diagnostics": dict(getattr(r, "merged_diagnostics", None) or {}),
            }
        return {
            "version": 4,
            "app_version": _APP_VERSION,
            "goal": self._goal.get(),
            "catalog_path": self._catalog_path,
            "validation_tolerance_pct": _VALIDATION_TOLERANCE_STRICT_PCT,
            "strict_validation": True,
            "report_title_override": (self._report_title_override.get() or "").strip(),
            "paths": list(self._pdf_paths),
            "axis_groups": list(self._axis_groups),
            "results": results,
        }

    def _save_project(self, *, notify: bool = True) -> tuple[bool, str]:
        if not self._pdf_paths:
            msg = "Список файлов пуст."
            if notify:
                messagebox.showinfo("Проект", msg)
            return False, msg
        if len(self._pdf_paths) == 1:
            path = _путь_проекта_для_pdf(self._pdf_paths[0])
        else:
            path = _путь_проекта_multi(self._pdf_paths[0])
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self._project_dict(), f, ensure_ascii=False, indent=2)
            self._current_project_path = path
            _recent_projects_touch(path)
            msg = f"Сохранено:\n{path}"
            if notify:
                messagebox.showinfo("Проект", msg)
            self._rep(False)
            return True, msg
        except Exception as e:
            self._rep(True)
            if notify:
                messagebox.showerror("Проект", str(e))
            return False, str(e)

    def _restore_results_from_project_raw(self, raw_res: dict[str, Any]) -> None:
        self._results.clear()
        if AnalyzeResult is None or PageResult is None:
            return
        norm_map: dict[str, dict[str, Any]] = {}
        for k, v in raw_res.items():
            if isinstance(k, str) and isinstance(v, dict):
                norm_map[os.path.normpath(k)] = v
        for p in self._pdf_paths:
            block = raw_res.get(p)
            if not isinstance(block, dict):
                block = norm_map.get(os.path.normpath(p))
            if not isinstance(block, dict):
                continue
            pps = []
            for row in block.get("per_page", []):
                pps.append(
                    PageResult(
                        int(row.get("page_index", 0)),
                        float(row.get("area_m2", 0)),
                        str(row.get("detail", "")),
                    )
                )
            ep = block.get("explicit_paint_area_m2")
            ep_f: float | None = None
            if ep is not None:
                try:
                    ep_f = float(ep)
                except (TypeError, ValueError):
                    ep_f = None
            self._results[p] = AnalyzeResult(
                strategy=str(block.get("strategy", "fallback")),
                confidence=str(block.get("confidence", "")),
                reason=str(block.get("reason", "")),
                per_page=pps,
                metal_lines=list(block.get("metal_lines") or []),
                metal_validation=list(block.get("metal_validation") or []),
                quality_metrics=dict(block.get("quality_metrics") or {}),
                explicit_paint_area_m2=ep_f,
                explicit_paint_area_ambiguous=bool(block.get("explicit_paint_area_ambiguous", False)),
                report_header=dict(block.get("report_header") or {}),
                shipment_qty_by_mark=dict(block.get("shipment_qty_by_mark") or {}),
                merged_diagnostics=dict(block.get("merged_diagnostics") or {}),
            )

    def _recompute_cross_file_validation_after_restore(self) -> None:
        self._cross_file_validation.clear()
        if cross_file_mark_checks is None or len(self._results) <= 1:
            return
        mb = {
            p: r.metal_lines
            for p, r in self._results.items()
            if r.strategy == "metal_catalog" and (getattr(r, "metal_lines", None) or [])
        }
        if len(mb) > 1:
            self._cross_file_validation = cross_file_mark_checks(
                mb,
                _CROSS_FILE_MARK_TOLERANCE_STRICT_PCT,
                min_abs_delta_m2=_CROSS_FILE_MARK_MIN_ABS_DELTA_M2,
            )

    def _apply_loaded_project_payload(
        self,
        data: dict[str, Any],
        *,
        pdf_paths_override: list[str] | None,
        json_source_path: str | None,
        show_loaded_dialog: bool,
    ) -> str:
        """Применить данные сессии из JSON.

        Возвращает:
          \"full\" — полная подгрузка с таблицей результатов;
          \"paths_only\" — список PDF и зоны выставлены, но движка AnalyzeResult нет (зависимости);
          \"none\" — отмена (старый формат, нет путей и т.п.).
        """
        ver = int(data.get("version", 1))
        if ver < 2:
            messagebox.showwarning(
                "Проект",
                "Это проект старого формата (ручные полигоны). Он не поддерживается в авто-режиме. Сохраните новый проект после анализа.",
            )
            return "none"

        if pdf_paths_override is not None:
            self._pdf_paths = list(pdf_paths_override)
            self._sync_list()
        else:
            paths = data.get("paths") or []
            if not paths:
                messagebox.showerror("Проект", "В файле нет списка путей.")
                return "none"
            missing = [p for p in paths if not os.path.isfile(p)]
            if missing:
                messagebox.showwarning(
                    "Проект",
                    "Некоторые PDF не найдены по сохранённым путям. Загружены только доступные.",
                )
            self._pdf_paths = [p for p in paths if os.path.isfile(p)]
            if not self._pdf_paths:
                messagebox.showerror("Проект", "Нет доступных PDF по путям из файла.")
                return "none"
            self._sync_list()

        if ver >= 3:
            g = data.get("goal")
            if g in ("metal", "sheet"):
                self._goal.set(g)
            cp = data.get("catalog_path")
            if cp and os.path.isfile(cp):
                self._catalog_path = cp
            ro = data.get("report_title_override")
            if isinstance(ro, str):
                self._report_title_override.set(ro)

        ag = data.get("axis_groups")
        if isinstance(ag, list):
            self._axis_groups = _normalize_axis_groups_json(ag)
        elif ver >= 4:
            self._axis_groups = []
        else:
            self._load_axis_groups_disk()
        self._refresh_axes_grp_list()
        self._refresh_axes_summary()

        self._restore_results_from_project_raw(data.get("results") or {})
        if AnalyzeResult is None:
            if json_source_path:
                self._current_project_path = os.path.normpath(json_source_path)
                _recent_projects_touch(json_source_path)
            self._assistant_update_project_status_label()
            self._rep(False)
            return "paths_only"

        self._recompute_cross_file_validation_after_restore()
        self._refresh_table()
        if len(self._results) == 1:
            r = next(iter(self._results.values()))
            self._lbl_strategy.config(
                text=f"Способ: {self._strategy_ru(r.strategy)} (уверенность: {r.confidence})"
            )
            self._lbl_reason.config(text=r.reason)
        else:
            self._lbl_strategy.config(text="Проект загружен (несколько файлов).")
            self._lbl_reason.config(text="")

        self._set_progress(100, "Проект из файла")

        if json_source_path:
            self._current_project_path = os.path.normpath(json_source_path)
            _recent_projects_touch(json_source_path)

        if show_loaded_dialog:
            messagebox.showinfo("Проект", "Проект загружен.")

        self._assistant_update_project_status_label()
        self._rep(False)
        return "full"

    def _load_project_dialog(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("JSON проекта", "*.json"), ("Все", "*.*")])
        if not path:
            return
        self._load_project_from_path(path)

    def _load_project_from_path(self, path: str) -> None:
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            applied = self._apply_loaded_project_payload(
                data,
                pdf_paths_override=None,
                json_source_path=os.path.normpath(path),
                show_loaded_dialog=False,
            )
            if applied == "full":
                messagebox.showinfo("Проект", "Проект загружен.")
        except Exception as e:
            self._rep(True)
            messagebox.showerror("Проект", str(e))


def запустить(родитель: tk.Misc) -> None:
    if not _DEPS_OK:
        txt = "Не установлены зависимости.\n\n"
        try:
            with open(_путь_установки(), "r", encoding="utf-8") as f:
                txt += f.read()
        except Exception:
            txt += "См. файл УСТАНОВКА.txt в папке утилиты."
        _shell.подготовить_стиль(родитель)
        _, body = _shell.карточка_утилиты(родитель)
        tk.Label(
            body,
            text=txt,
            justify=tk.LEFT,
            wraplength=520,
            fg=_hub_theme.ERROR_FG,
            bg=_hub_theme.CARD,
            font=_hub_theme.FONT_BASE,
        ).pack(anchor=tk.W)
        try:
            родитель.report_tab_error(True)  # type: ignore[attr-defined]
        except Exception:
            pass
        return
    try:
        _, body = _shell.карточка_утилиты(родитель)
        ОкраскаПоPdfПанель(body)
        try:
            родитель.report_tab_error(False)  # type: ignore[attr-defined]
        except Exception:
            pass
    except Exception as e:
        try:
            родитель.report_tab_error(True)  # type: ignore[attr-defined]
        except Exception:
            pass
        try:
            for w in родитель.winfo_children():
                w.destroy()
        except Exception:
            pass
        _shell.подготовить_стиль(родитель)
        _, err_body = _shell.карточка_утилиты(родитель)
        tk.Label(
            err_body,
            text=f"Ошибка запуска утилиты:\n{e}",
            fg=_hub_theme.ERROR_FG,
            bg=_hub_theme.CARD,
            justify=tk.LEFT,
            font=_hub_theme.FONT_BASE,
        ).pack(anchor=tk.W)


def автосохранить() -> None:
    """Хаб может вызывать раз в 30 с — для этой утилиты не требуется."""
    pass
