# Генератор отчетов в Excel
import logging
import os
from datetime import datetime
from config import Config


class ExcelWriter:
    """Класс для создания отчетов в Excel."""
    
    def __init__(self):
        """Инициализация ExcelWriter."""
        self.config = Config
        self.openpyxl = None
        self._import_openpyxl()
        logging.info("ExcelWriter инициализирован")
    
    def _import_pdfplumber(self):
        """Импорт библиотеки openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            self.openpyxl = openpyxl
            self.styles = {
                "font": Font,
                "alignment": Alignment,
                "fill": PatternFill,
                "border": Border,
                "side": Side,
            }
            logging.info("openpyxl загружен успешно")
        except ImportError:
            logging.warning("openpyxl не установлен, используем заглушки")
            self.openpyxl = None
    
    def _import_openpyxl(self):
        """Импорт библиотеки openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            self.openpyxl = openpyxl
            self.styles_module = type('styles', (), {
                'Font': Font,
                'Alignment': Alignment,
                'PatternFill': PatternFill,
                'Border': Border,
                'Side': Side,
            })
            logging.info("openpyxl загружен успешно")
        except ImportError:
            logging.warning("openpyxl не установлен, используем заглушки")
            self.openpyxl = None
    
    def create_report(self, data, filename):
        """
        Создать общий отчет Excel.
        
        Args:
            data: словарь с данными отчета
            filename: имя файла для сохранения
        
        Returns:
            путь к созданному файлу
        """
        try:
            logging.info(f"Создание отчета: {filename}")
            
            # Создаем папку для отчетов, если нет
            os.makedirs(self.config.REPORTS_FOLDER, exist_ok=True)
            
            output_path = os.path.join(self.config.REPORTS_FOLDER, filename)
            
            if self.openpyxl is None:
                logging.warning("openpyxl не доступен, создаем CSV")
                return self._create_csv_report(data, output_path)
            
            # Создаем книгу
            wb = self.openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчет"
            
            # Заголовок
            ws['A1'] = data.get('название', 'Отчет')
            ws['A1'].font = self.styles_module.Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            
            # Дата
            ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws.merge_cells('A2:E2')
            
            # Данные
            row = 4
            for key, value in data.items():
                if key != 'название':
                    ws[f'A{row}'] = str(key)
                    ws[f'B{row}'] = str(value)
                    row += 1
            
            # Форматирование
            self.format_excel(ws)
            
            # Сохранение
            wb.save(output_path)
            logging.info(f"Отчет сохранен: {output_path}")
            return output_path
            
        except Exception as e:
            logging.error(f"Ошибка создания отчета: {e}")
            return None
    
    def _create_csv_report(self, data, output_path):
        """Создание отчета в формате CSV."""
        csv_path = output_path.replace('.xlsx', '.csv')
        
        try:
            with open(csv_path, 'w', encoding='utf-8') as f:
                f.write(f"{data.get('название', 'Отчет')}\n")
                f.write(f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n")
                for key, value in data.items():
                    if key != 'название':
                        f.write(f"{key}: {value}\n")
            
            logging.info(f"CSV отчет сохранен: {csv_path}")
            return csv_path
            
        except Exception as e:
            logging.error(f"Ошибка создания CSV: {e}")
            return None
    
    def write_specification(self, data):
        """
        Создать спецификацию металла.
        
        Args:
            data: словарь с данными спецификации
        
        Returns:
            путь к созданному файлу
        """
        try:
            logging.info("Создание спецификации металла")
            
            os.makedirs(self.config.REPORTS_FOLDER, exist_ok=True)
            
            filename = f"specification_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(self.config.REPORTS_FOLDER, filename)
            
            if self.openpyxl is None:
                logging.warning("openpyxl не доступен")
                return None
            
            wb = self.openpyxl.Workbook()
            ws = wb.active
            ws.title = "Спецификация"
            
            # Заголовок
            ws['A1'] = "СПЕЦИФИКАЦИЯ МЕТАЛЛОКОНСТРУКЦИЙ"
            ws['A1'].font = self.styles_module.Font(size=14, bold=True)
            ws.merge_cells('A1:F1')
            
            ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y')}"
            ws.merge_cells('A2:F2')
            
            # Заголовки таблицы
            headers = self.config.EXCEL_HEADERS["спецификация"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = self.styles_module.Font(bold=True)
                cell.fill = self.styles_module.PatternFill(start_color="CCCCCC", 
                                                          end_color="CCCCCC", 
                                                          fill_type="solid")
            
            # Данные спецификации
            row = 5
            positions = data.get('позиции', [])
            
            for i, item in enumerate(positions, 1):
                ws.cell(row=row, column=1, value=i)  # Поз.
                ws.cell(row=row, column=2, value=item.get('наименование', ''))
                ws.cell(row=row, column=3, value=item.get('обозначение', ''))
                ws.cell(row=row, column=4, value=item.get('количество', ''))
                ws.cell(row=row, column=5, value=item.get('материал', ''))
                ws.cell(row=row, column=6, value=item.get('примечание', ''))
                row += 1
            
            # Итого
            row += 1
            ws.cell(row=row, column=1, value="Итого:")
            ws.cell(row=row, column=4, value=len(positions))
            
            # Форматирование
            self.format_excel(ws)
            
            wb.save(output_path)
            logging.info(f"Спецификация сохранена: {output_path}")
            return output_path
            
        except Exception as e:
            logging.error(f"Ошибка создания спецификации: {e}")
            return None
    
    def write_cutting_plan(self, data):
        """
        Создать карту раскроя.
        
        Args:
            data: словарь с данными раскроя
        
        Returns:
            путь к созданному файлу
        """
        try:
            logging.info("Создание карты раскроя")
            
            os.makedirs(self.config.REPORTS_FOLDER, exist_ok=True)
            
            filename = f"cutting_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(self.config.REPORTS_FOLDER, filename)
            
            if self.openpyxl is None:
                logging.warning("openpyxl не доступен")
                return None
            
            wb = self.openpyxl.Workbook()
            ws = wb.active
            ws.title = "Карта раскроя"
            
            # Заголовок
            ws['A1'] = "КАРТА РАСКРОЯ"
            ws['A1'].font = self.styles_module.Font(size=14, bold=True)
            ws.merge_cells('A1:E1')
            
            ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y')}"
            ws.merge_cells('A2:E2')
            
            # Заголовки таблицы
            headers = self.config.EXCEL_HEADERS["раскрой"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = self.styles_module.Font(bold=True)
                cell.fill = self.styles_module.PatternFill(start_color="CCCCCC", 
                                                          end_color="CCCCCC", 
                                                          fill_type="solid")
            
            # Данные
            row = 5
            details = data.get('детали', [])
            
            for i, detail in enumerate(details, 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=detail.get('название', ''))
                ws.cell(row=row, column=3, value=detail.get('ширина', ''))
                ws.cell(row=row, column=4, value=detail.get('длина', ''))
                ws.cell(row=row, column=5, value=detail.get('количество', ''))
                ws.cell(row=row, column=6, value=detail.get('площадь', ''))
                row += 1
            
            # Итого
            row += 1
            ws.cell(row=row, column=1, value="Итого листов:")
            ws.cell(row=row, column=2, value=data.get('всего_листов', ''))
            
            # Форматирование
            self.format_excel(ws)
            
            wb.save(output_path)
            logging.info(f"Карта раскроя сохранена: {output_path}")
            return output_path
            
        except Exception as e:
            logging.error(f"Ошибка создания карты раскроя: {e}")
            return None
    
    def format_excel(self, worksheet):
        """
        Форматирование листа Excel.
        
        Args:
            worksheet: лист для форматирования
        """
        try:
            if self.openpyxl is None:
                return
            
            # Автоширина колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Рамки для всех ячеек
            thin_border = self.styles_module.Border(
                left=self.styles_module.Side(style='thin'),
                right=self.styles_module.Side(style='thin'),
                top=self.styles_module.Side(style='thin'),
                bottom=self.styles_module.Side(style='thin')
            )
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = self.styles_module.Alignment(horizontal='left', 
                                                                  vertical='center')
            
            logging.info("Форматирование Excel завершено")
            
        except Exception as e:
            logging.error(f"Ошибка форматирования: {e}")
